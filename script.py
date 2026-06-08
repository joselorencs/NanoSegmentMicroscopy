#!/usr/bin/env python3
"""
PARTICLE_AI.py

Step-1 implementation for SEM/TEM/AFM particle segmentation with SAM:
- Import image (jpg/png/tif/tiff)
- Run SAM automatic masks
- Navigate image (pan/zoom via matplotlib toolbar)
- Select particles by:
  - Individual click
  - Rectangular ROI
  - Circular ROI
- Show selected particle areas on a right-side table in um^2

Notes:
- Requires SAM checkpoint (.pth) and Python packages `torch` + `segment_anything`.
- Pixel-to-micron scale must be provided as um/px to compute physical area.
"""

from __future__ import annotations

import argparse
import csv
import colorsys
import math
import os
import sys
import threading
import time
import warnings
from dataclasses import dataclass
from typing import Callable, Dict, List, Optional, Set, Tuple

import numpy as np
import tkinter as tk
from PIL import Image, ImageTk, ImageDraw
from tkinter import colorchooser, filedialog, messagebox, simpledialog, ttk

import matplotlib

matplotlib.use("TkAgg")
PLOT_TEXT_SIZE = 18
matplotlib.rcParams.update(
    {
        "axes.titlesize": PLOT_TEXT_SIZE,
        "axes.labelsize": PLOT_TEXT_SIZE,
        "xtick.labelsize": PLOT_TEXT_SIZE,
        "ytick.labelsize": PLOT_TEXT_SIZE,
    }
)
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from matplotlib.transforms import Bbox
from matplotlib.lines import Line2D
from matplotlib.patches import Circle, Ellipse, Rectangle
from matplotlib.widgets import SpanSelector

SAM_IMPORT_ERROR: Optional[Exception] = None
SAM_AVAILABLE = False
try:
    import torch
    from segment_anything import SamAutomaticMaskGenerator, sam_model_registry

    SAM_AVAILABLE = True
except Exception as exc:  # pragma: no cover - depends on local env
    SAM_IMPORT_ERROR = exc

GMM_IMPORT_ERROR: Optional[Exception] = None
GMM_AVAILABLE = False
try:
    from scipy.stats import gaussian_kde, kurtosis, skew, t as student_t
    from sklearn.mixture import GaussianMixture

    GMM_AVAILABLE = True
except Exception as exc:  # pragma: no cover - depends on local env
    GMM_IMPORT_ERROR = exc

FIT_IMPORT_ERROR: Optional[Exception] = None
FIT_AVAILABLE = False
try:
    from scipy.optimize import curve_fit

    FIT_AVAILABLE = True
except Exception as exc:  # pragma: no cover - depends on local env
    FIT_IMPORT_ERROR = exc

NND_IMPORT_ERROR: Optional[Exception] = None
NND_AVAILABLE = False
try:
    from scipy.spatial import cKDTree
    from scipy.stats import gaussian_kde

    NND_AVAILABLE = True
except Exception as exc:  # pragma: no cover - depends on local env
    NND_IMPORT_ERROR = exc

VORONOI_IMPORT_ERROR: Optional[Exception] = None
VORONOI_AVAILABLE = False
try:
    from scipy.spatial import Voronoi

    VORONOI_AVAILABLE = True
except Exception as exc:  # pragma: no cover - depends on local env
    VORONOI_IMPORT_ERROR = exc

DBSCAN_IMPORT_ERROR: Optional[Exception] = None
DBSCAN_AVAILABLE = False
try:
    from scipy.spatial import ConvexHull, QhullError
    from sklearn.cluster import DBSCAN

    DBSCAN_AVAILABLE = True
except Exception as exc:  # pragma: no cover - depends on local env
    DBSCAN_IMPORT_ERROR = exc

XLSX_IMPORT_ERROR: Optional[Exception] = None
XLSX_AVAILABLE = False
try:
    from openpyxl import Workbook

    XLSX_AVAILABLE = True
except Exception as exc:  # pragma: no cover - depends on local env
    XLSX_IMPORT_ERROR = exc

CV2_IMPORT_ERROR: Optional[Exception] = None
CV2_AVAILABLE = False
try:
    import cv2

    CV2_AVAILABLE = True
except Exception as exc:  # pragma: no cover - depends on local env
    CV2_IMPORT_ERROR = exc

SKIMAGE_IMPORT_ERROR: Optional[Exception] = None
SKIMAGE_AVAILABLE = False
try:
    from skimage.measure import perimeter_crofton

    SKIMAGE_AVAILABLE = True
except Exception as exc:  # pragma: no cover - depends on local env
    SKIMAGE_IMPORT_ERROR = exc

SUPPORTED_EXT = (".jpg", ".jpeg", ".png", ".tif", ".tiff")


@dataclass
class ParticleMask:
    mask_id: int
    segmentation: np.ndarray  # bool [H, W]
    area_px: int
    centroid_xy: Tuple[float, float]  # (x, y)
    bbox_xywh: Tuple[int, int, int, int]
    image_index: int = 0
    offset_xy: Tuple[int, int] = (0, 0)
    feret_px: Optional[float] = None


def load_rgb_image(path: str) -> np.ndarray:
    img = Image.open(path).convert("RGB")
    return np.asarray(img, dtype=np.uint8)


def dependency_help_text() -> str:
    return (
        "Missing dependencies for SAM.\n\n"
        "Install (CPU example):\n"
        "1) pip install torch torchvision --index-url https://download.pytorch.org/whl/cpu\n"
        "2) pip install git+https://github.com/facebookresearch/segment-anything.git\n"
        "3) pip install matplotlib pillow numpy\n\n"
        f"Import error: {SAM_IMPORT_ERROR}"
    )


def build_particles_from_sam_dicts(masks: List[Dict]) -> List[ParticleMask]:
    particles: List[ParticleMask] = []
    for idx, m in enumerate(masks, start=1):
        seg = np.asarray(m["segmentation"], dtype=bool)
        area_px = int(m.get("area", int(seg.sum())))
        ys, xs = np.nonzero(seg)
        if len(xs) == 0:
            continue
        cx = float(xs.mean())
        cy = float(ys.mean())
        bbox = m.get("bbox", [0, 0, 0, 0])
        particle = ParticleMask(
            mask_id=idx,
            segmentation=seg,
            area_px=area_px,
            centroid_xy=(cx, cy),
            bbox_xywh=(int(bbox[0]), int(bbox[1]), int(bbox[2]), int(bbox[3])),
        )
        particles.append(particle)
    return particles


def filter_particles_by_image_size(
    particles: List[ParticleMask],
    image_hw: Tuple[int, int],
    max_fraction: float = 0.8,
) -> List[ParticleMask]:
    """Remove masks that are too large relative to the full image."""
    if not particles:
        return particles

    img_h, img_w = image_hw
    if img_h <= 0 or img_w <= 0:
        return particles

    area_limit = float(max_fraction) * float(img_h * img_w)
    length_limit = float(max_fraction) * float(max(img_h, img_w))
    width_limit = float(max_fraction) * float(min(img_h, img_w))

    filtered: List[ParticleMask] = []
    for p in particles:
        bbox_w = float(max(1, int(p.bbox_xywh[2])))
        bbox_h = float(max(1, int(p.bbox_xywh[3])))
        length_px = max(bbox_w, bbox_h)
        width_px = min(bbox_w, bbox_h)
        area_px = float(max(0, int(p.area_px)))

        # Discard likely background masks that span most of the image.
        if area_px >= area_limit or length_px >= length_limit or width_px >= width_limit:
            continue
        filtered.append(p)
    return filtered


def _resize_image_for_sam(image_rgb: np.ndarray, max_side: int) -> Tuple[np.ndarray, Tuple[int, int]]:
    h, w = image_rgb.shape[:2]
    if max_side <= 0:
        return image_rgb, (h, w)
    current_max = max(h, w)
    if current_max <= max_side:
        return image_rgb, (h, w)
    scale = float(max_side) / float(current_max)
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    resized = np.asarray(
        Image.fromarray(image_rgb).resize((new_w, new_h), Image.Resampling.BILINEAR),
        dtype=np.uint8,
    )
    return resized, (h, w)


def _upsample_masks_to_original_size(masks: List[Dict], output_hw: Tuple[int, int]) -> List[Dict]:
    out_h, out_w = output_hw
    out_masks: List[Dict] = []
    for m in masks:
        seg = np.asarray(m["segmentation"], dtype=np.uint8) * 255
        seg_up = Image.fromarray(seg).resize((out_w, out_h), Image.Resampling.NEAREST)
        seg_bool = np.asarray(seg_up, dtype=np.uint8) > 0

        ys, xs = np.nonzero(seg_bool)
        if len(xs) == 0:
            continue
        x_min, x_max = int(xs.min()), int(xs.max())
        y_min, y_max = int(ys.min()), int(ys.max())

        m2 = dict(m)
        m2["segmentation"] = seg_bool
        m2["area"] = int(seg_bool.sum())
        m2["bbox"] = [x_min, y_min, int(x_max - x_min + 1), int(y_max - y_min + 1)]
        out_masks.append(m2)
    return out_masks


def run_sam(
    image_rgb: np.ndarray,
    checkpoint_path: str,
    model_type: str = "vit_b",
    max_side: int = 1280,
    fast_mode: bool = True,
    progress_callback: Optional[Callable[[str, float], None]] = None,
    cancel_event: Optional[threading.Event] = None,
) -> List[ParticleMask]:
    if not SAM_AVAILABLE:
        raise RuntimeError(dependency_help_text())
    if not os.path.isfile(checkpoint_path):
        raise FileNotFoundError(f"SAM checkpoint not found: {checkpoint_path}")
    if model_type not in ("vit_b", "vit_l", "vit_h"):
        raise ValueError("model_type must be one of: vit_b, vit_l, vit_h")

    if progress_callback is not None:
        progress_callback("Loading SAM model...", 5.0)
    device = "cuda" if torch.cuda.is_available() else "cpu"
    sam = sam_model_registry[model_type](checkpoint=checkpoint_path)
    sam.to(device=device)

    if cancel_event is not None and cancel_event.is_set():
        raise RuntimeError("SAM cancelled by user.")

    if progress_callback is not None:
        progress_callback("Preparing image for SAM...", 15.0)
    input_image, original_hw = _resize_image_for_sam(image_rgb, max_side=max_side)

    # Fast defaults for CPU; switch off fast_mode for higher quality.
    points_per_side = 24 if fast_mode else 32
    crop_n_layers = 0 if fast_mode else 1

    if progress_callback is not None:
        progress_callback("Configuring mask generator...", 25.0)
    mask_generator = SamAutomaticMaskGenerator(
        model=sam,
        points_per_side=points_per_side,
        pred_iou_thresh=0.80,
        stability_score_thresh=0.90,
        crop_n_layers=crop_n_layers,
        crop_n_points_downscale_factor=2,
        min_mask_region_area=20,
    )

    if cancel_event is not None and cancel_event.is_set():
        raise RuntimeError("SAM cancelled by user.")

    if progress_callback is not None:
        progress_callback("Generating masks...", 35.0)
    masks = mask_generator.generate(input_image)
    if cancel_event is not None and cancel_event.is_set():
        raise RuntimeError("SAM cancelled by user.")

    if progress_callback is not None:
        progress_callback("Post-processing masks...", 90.0)
    if input_image.shape[:2] != original_hw:
        masks = _upsample_masks_to_original_size(masks, output_hw=original_hw)
    masks = sorted(masks, key=lambda d: int(d.get("area", 0)), reverse=True)
    if progress_callback is not None:
        progress_callback("SAM finished.", 100.0)
    particles = build_particles_from_sam_dicts(masks)
    particles = filter_particles_by_image_size(particles, image_hw=original_hw, max_fraction=0.8)
    return particles


def quick_test(image_path: str, checkpoint_path: str, model_type: str, um_per_px: float) -> int:
    image_rgb = load_rgb_image(image_path)
    particles = run_sam(image_rgb=image_rgb, checkpoint_path=checkpoint_path, model_type=model_type)
    if len(particles) == 0:
        print("SAM test completed: 0 particle masks found.")
        return 0

    areas_um2 = np.array([p.area_px * (um_per_px ** 2) for p in particles], dtype=float)
    print(f"SAM test completed: {len(particles)} masks found.")
    print(f"Area um^2 stats -> min: {areas_um2.min():.6f}, median: {np.median(areas_um2):.6f}, max: {areas_um2.max():.6f}")
    print("Top 10 areas um^2:")
    for i, area in enumerate(areas_um2[:10], start=1):
        print(f"  #{i}: {area:.6f}")
    return len(particles)


class ParticleAIApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("NanoSegment AI")
        self.root.geometry("1500x900")
        self.base_window_width = 1500
        self.base_window_height = 900

        # State
        self.left_panel_container: Optional[ttk.Frame] = None
        self.left_panel_canvas: Optional[tk.Canvas] = None
        self.left_panel_scroll: Optional[ttk.Scrollbar] = None
        self.left_panel_inner: Optional[ttk.Frame] = None
        self.left_panel_window: Optional[int] = None
        self.image_path: Optional[str] = None
        self.image_rgb: Optional[np.ndarray] = None
        self.image_paths: List[str] = []
        self.image_rgbs: List[np.ndarray] = []
        self.image_offsets: Dict[int, Tuple[int, int]] = {}
        self.image_cell_size: int = 0
        self.view_mode: int = 1
        self.view_mode_icons: Dict[int, ImageTk.PhotoImage] = {}
        self.view_mode_buttons: Dict[int, tk.Button] = {}
        self.particles: List[ParticleMask] = []
        self.particles_by_image: Dict[int, List[ParticleMask]] = {}
        self.selected_ids: Set[int] = set()
        self.mask_pick_map: Optional[np.ndarray] = None  # [H, W] -> mask_id or -1
        self.mode: str = "none"  # none | navigate | select_one | roi_rect | roi_circle | deselect_rect | measure_line | measure_angle | oval_section | rect_section | scale_bar_rect
        self.autocalibration_pending: bool = False
        self.autocalibration_prev_mode: str = "none"
        self.show_more_data: bool = False
        self.show_more_btn: Optional[ttk.Button] = None

        self.drag_start: Optional[Tuple[float, float]] = None
        self.roi_artist: Optional[object] = None
        self.measure_line: Optional[Tuple[float, float, float, float]] = None
        self.measure_label: Optional[str] = None
        self.angle_points: List[Tuple[float, float]] = []
        self.angle_label: Optional[str] = None
        self.angle_preview_point: Optional[Tuple[float, float]] = None
        self.section_shape: Optional[str] = None
        self.section_bbox: Optional[Tuple[float, float, float, float]] = None
        self.section_label: Optional[str] = None
        self.flip_window: Optional[tk.Toplevel] = None
        self.selection_color_rgb: Tuple[int, int, int] = (227, 41, 158)
        self.selection_alpha: float = 0.40
        self.color_settings_window: Optional[tk.Toplevel] = None
        self.color_wheel_img: Optional[Image.Image] = None
        self.color_wheel_tk: Optional[ImageTk.PhotoImage] = None
        self.color_wheel_canvas: Optional[tk.Canvas] = None
        self.color_r_var: Optional[tk.StringVar] = None
        self.color_g_var: Optional[tk.StringVar] = None
        self.color_b_var: Optional[tk.StringVar] = None
        self.color_preview_canvas: Optional[tk.Canvas] = None
        self.sam_running: bool = False
        self.sam_job_id: int = 0
        self.sam_cancel_event: Optional[threading.Event] = None
        self.sam_cancel_requested: bool = False
        self.sam_started_at: float = 0.0
        self.sam_estimated_total_s: float = 0.0
        self.sam_progress_after_id: Optional[str] = None
        self.hist_window: Optional[tk.Toplevel] = None
        self.hist_fig: Optional[Figure] = None
        self.hist_ax = None
        self.hist_canvas: Optional[FigureCanvasTkAgg] = None
        self.hist_property_var: Optional[tk.StringVar] = None
        self.hist_click_cid: Optional[int] = None
        self.hist_bins: Optional[np.ndarray] = None
        self.hist_values: Optional[np.ndarray] = None
        self.hist_value_ids: Optional[np.ndarray] = None
        self.hist_patches: List = []
        self.hist_active_bin_idx: Optional[int] = None
        self.highlighted_ids: Set[int] = set()
        self.hist_icon_img: Optional[ImageTk.PhotoImage] = None
        self.basic_stats_icon_img: Optional[ImageTk.PhotoImage] = None
        self.basic_stats_window: Optional[tk.Toplevel] = None
        self.basic_stats_property_var: Optional[tk.StringVar] = None
        self.basic_stats_basis_var: Optional[tk.StringVar] = None
        self.basic_stats_value_vars: Dict[str, tk.StringVar] = {}
        self.gmm_icon_img: Optional[ImageTk.PhotoImage] = None
        self.gmm_window: Optional[tk.Toplevel] = None
        self.gmm_property_var: Optional[tk.StringVar] = None
        self.gmm_basis_var: Optional[tk.StringVar] = None
        self.gmm_tree: Optional[ttk.Treeview] = None
        self.gmm_help_btn: Optional[tk.Button] = None
        self.gmm_help_window: Optional[tk.Toplevel] = None
        self.gmm_help_tracking: bool = False
        self.bootstrap_icon_img: Optional[ImageTk.PhotoImage] = None
        self.bootstrap_window: Optional[tk.Toplevel] = None
        self.bootstrap_property_var: Optional[tk.StringVar] = None
        self.bootstrap_n_var: Optional[tk.StringVar] = None
        self.bootstrap_basis_var: Optional[tk.StringVar] = None
        self.bootstrap_stat_vars: Dict[str, tk.StringVar] = {}
        self.bootstrap_fig: Optional[Figure] = None
        self.bootstrap_ax = None
        self.bootstrap_canvas: Optional[FigureCanvasTkAgg] = None
        self.bootstrap_help_btn: Optional[tk.Button] = None
        self.bootstrap_help_window: Optional[tk.Toplevel] = None
        self.bootstrap_help_tracking: bool = False
        self.dbscan_icon_img: Optional[ImageTk.PhotoImage] = None
        self.dbscan_window: Optional[tk.Toplevel] = None
        self.dbscan_eps_var: Optional[tk.StringVar] = None
        self.dbscan_min_samples_var: Optional[tk.StringVar] = None
        self.dbscan_basis_var: Optional[tk.StringVar] = None
        self.dbscan_summary_vars: Dict[str, tk.StringVar] = {}
        self.dbscan_tree: Optional[ttk.Treeview] = None
        self.dbscan_export_headers: Tuple[str, ...] = ()
        self.dbscan_export_rows: List[Tuple] = []
        self.dbscan_summary_snapshot: Dict[str, str] = {}
        self.dbscan_help_btn: Optional[tk.Button] = None
        self.dbscan_help_window: Optional[tk.Toplevel] = None
        self.dbscan_help_tracking: bool = False
        self.threshold_icon_img: Optional[ImageTk.PhotoImage] = None
        self.threshold_window: Optional[tk.Toplevel] = None
        self.threshold_property_var: Optional[tk.StringVar] = None
        self.threshold_min_var: Optional[tk.StringVar] = None
        self.threshold_max_var: Optional[tk.StringVar] = None
        self.threshold_unit_var: Optional[tk.StringVar] = None
        self.threshold_basis_var: Optional[tk.StringVar] = None
        self.threshold_fig: Optional[Figure] = None
        self.threshold_ax = None
        self.threshold_canvas: Optional[FigureCanvasTkAgg] = None
        self.threshold_span: Optional[SpanSelector] = None
        self.threshold_values: Optional[np.ndarray] = None
        self.threshold_ids: Optional[np.ndarray] = None
        self.nnd_icon_img: Optional[ImageTk.PhotoImage] = None
        self.nnd_summary_window: Optional[tk.Toplevel] = None
        self.nnd_plot_window: Optional[tk.Toplevel] = None
        self.nnd_basis_var: Optional[tk.StringVar] = None
        self.nnd_mc_var: Optional[tk.StringVar] = None
        self.nnd_summary_vars: Dict[str, tk.StringVar] = {}
        self.nnd_map_fig: Optional[Figure] = None
        self.nnd_map_ax = None
        self.nnd_map_hist_ax = None
        self.nnd_map_canvas: Optional[FigureCanvasTkAgg] = None
        self.nnd_map_colorbar = None
        self.nnd_map_hist_values: Optional[np.ndarray] = None
        self.nnd_map_hist_bins: Optional[int] = None
        self.nnd_map_kde_line = None
        self.nnd_map_kde_annotation = None
        self.nnd_map_hover_cid: Optional[int] = None
        self.nnd_map_kde_enabled: bool = False
        self.nnd_map_kde_x: Optional[np.ndarray] = None
        self.nnd_map_kde_y: Optional[np.ndarray] = None
        self.nnd_map_kde_btn: Optional[ttk.Button] = None
        self.nnd_hist_fig: Optional[Figure] = None
        self.nnd_hist_ax = None
        self.nnd_hist_canvas: Optional[FigureCanvasTkAgg] = None
        self.nnd_results: Optional[Dict[str, object]] = None
        self.nnd_help_btn: Optional[tk.Button] = None
        self.nnd_help_window: Optional[tk.Toplevel] = None
        self.nnd_help_tracking: bool = False
        self.ttest_icon_img: Optional[ImageTk.PhotoImage] = None
        self.ttest_window: Optional[tk.Toplevel] = None
        self.ttest_property_var: Optional[tk.StringVar] = None
        self.ttest_type_var: Optional[tk.StringVar] = None
        self.ttest_basis_var: Optional[tk.StringVar] = None
        self.ttest_tree: Optional[ttk.Treeview] = None
        self.ttest_export_headers: Tuple[str, ...] = ()
        self.ttest_export_rows: List[Tuple] = []
        self.ttest_help_btn: Optional[tk.Button] = None
        self.ttest_help_window: Optional[tk.Toplevel] = None
        self.ttest_help_tracking: bool = False
        self.false_color_icon_img: Optional[ImageTk.PhotoImage] = None
        self.false_color_window: Optional[tk.Toplevel] = None
        self.false_color_property_var: Optional[tk.StringVar] = None
        self.false_color_gradient_var: Optional[tk.StringVar] = None
        self.false_color_fig: Optional[Figure] = None
        self.false_color_ax = None
        self.false_color_canvas: Optional[FigureCanvasTkAgg] = None
        self.false_color_colorbar = None
        self.false_color_gradient_icons: List[ImageTk.PhotoImage] = []
        self.violin_icon_img: Optional[ImageTk.PhotoImage] = None
        self.violin_window: Optional[tk.Toplevel] = None
        self.violin_property_var: Optional[tk.StringVar] = None
        self.violin_fig: Optional[Figure] = None
        self.violin_ax = None
        self.violin_canvas: Optional[FigureCanvasTkAgg] = None
        self.violin_help_btn: Optional[tk.Button] = None
        self.violin_help_window: Optional[tk.Toplevel] = None
        self.violin_help_tracking: bool = False
        self.bivariate_icon_img: Optional[ImageTk.PhotoImage] = None
        self.bivariate_window: Optional[tk.Toplevel] = None
        self.bivariate_x_var: Optional[tk.StringVar] = None
        self.bivariate_y_var: Optional[tk.StringVar] = None
        self.bivariate_fig: Optional[Figure] = None
        self.bivariate_ax = None
        self.bivariate_canvas: Optional[FigureCanvasTkAgg] = None
        self.bivariate_ids: Optional[np.ndarray] = None
        self.bivariate_scatter = None
        self.bivariate_pick_cid: Optional[int] = None
        self.bivariate_x_data: Optional[np.ndarray] = None
        self.bivariate_y_data: Optional[np.ndarray] = None
        self.bivariate_fit_line = None
        self.bivariate_fit_text = None
        self.bivariate_fit_info_window: Optional[tk.Toplevel] = None
        self.bivariate_gmm_icon_img: Optional[ImageTk.PhotoImage] = None
        self.bivariate_gmm_window: Optional[tk.Toplevel] = None
        self.bivariate_gmm_x_var: Optional[tk.StringVar] = None
        self.bivariate_gmm_y_var: Optional[tk.StringVar] = None
        self.bivariate_gmm_distinction_var: Optional[tk.StringVar] = None
        self.bivariate_gmm_fig: Optional[Figure] = None
        self.bivariate_gmm_ax = None
        self.bivariate_gmm_canvas: Optional[FigureCanvasTkAgg] = None
        self.bivariate_gmm_scatter_groups: List[Dict[str, object]] = []
        self.bivariate_gmm_pick_cid: Optional[int] = None
        self.bivariate_gmm_x_data: Optional[np.ndarray] = None
        self.bivariate_gmm_y_data: Optional[np.ndarray] = None
        self.bivariate_gmm_fit_line = None
        self.bivariate_gmm_fit_text = None
        self.bivariate_gmm_fit_info_window: Optional[tk.Toplevel] = None
        self.overlay_icon_img: Optional[ImageTk.PhotoImage] = None
        self.overlay_prompt_window: Optional[tk.Toplevel] = None
        self.overlay_opacity_window: Optional[tk.Toplevel] = None
        self.overlay_base_image: Optional[np.ndarray] = None
        self.help_icon_img: Optional[ImageTk.PhotoImage] = None
        self.iou_window: Optional[tk.Toplevel] = None
        self.iou_canvas: Optional[tk.Canvas] = None
        self.iou_base_image: Optional[Image.Image] = None
        self.iou_overlay_image: Optional[Image.Image] = None
        self.iou_display_tk: Optional[ImageTk.PhotoImage] = None
        self.iou_mask: Optional[Image.Image] = None
        self.iou_scale: float = 1.0
        self.iou_brush_size_var: Optional[tk.StringVar] = None
        self.iou_brush_size_scale: Optional[ttk.Scale] = None
        self.iou_brush_color: Tuple[int, int, int] = (255, 0, 0)
        self.iou_brush_preview: Optional[tk.Canvas] = None
        self.iou_result_var: Optional[tk.StringVar] = None
        self.iou_precision_var: Optional[tk.StringVar] = None
        self.iou_recall_var: Optional[tk.StringVar] = None
        self.iou_f1_var: Optional[tk.StringVar] = None
        self.iou_ba_mean_var: Optional[tk.StringVar] = None
        self.iou_ba_sd_var: Optional[tk.StringVar] = None
        self.iou_ba_loa_var: Optional[tk.StringVar] = None
        self.iou_ba_data = None
        self.iou_ba_fig: Optional[Figure] = None
        self.iou_ba_ax = None
        self.iou_ba_canvas: Optional[FigureCanvasTkAgg] = None
        self.iou_draw_enabled: bool = False
        self.iou_brush_mode: str = "brush"
        self.iou_mode_var: Optional[tk.StringVar] = None
        self.iou_tool_buttons: Dict[str, ttk.Button] = {}
        self.iou_selected_snapshot: Set[int] = set()
        self.iou_last_pos: Optional[Tuple[float, float]] = None

        # Correct parameter editing (length/width)
        self.correct_params_active: bool = False
        self.correct_param_name: Optional[str] = None
        self.correct_param_var: Optional[tk.StringVar] = None
        self.correct_params_panel: Optional[ttk.Frame] = None
        self.correct_params_btn: Optional[ttk.Button] = None
        self.correct_lines: Dict[int, Line2D] = {}
        self.correct_line_data: Dict[int, Dict[str, object]] = {}
        self.correct_drag: Optional[Dict[str, object]] = None
        self.length_overrides_px: Dict[int, float] = {}
        self.width_overrides_px: Dict[int, float] = {}
        self.correct_overrides_backup: Optional[Tuple[Dict[int, float], Dict[int, float]]] = None

        self._build_ui()
        self._connect_events()

    def _build_ui(self) -> None:
        self.main_frame = ttk.Frame(self.root, padding=8)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.left_panel_container = ttk.Frame(self.main_frame, width=160)
        self.left_panel_container.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 8))
        self.left_panel_container.pack_propagate(False)

        self.left_panel_canvas = tk.Canvas(self.left_panel_container, highlightthickness=0)
        self.left_panel_scroll = ttk.Scrollbar(
            self.left_panel_container, orient="vertical", command=self.left_panel_canvas.yview
        )
        self.left_panel_canvas.configure(yscrollcommand=self.left_panel_scroll.set)
        self.left_panel_container.grid_rowconfigure(0, weight=1)
        self.left_panel_container.grid_columnconfigure(0, weight=1)
        self.left_panel_canvas.grid(row=0, column=0, sticky="nsew")
        self.left_panel_scroll.grid(row=0, column=1, sticky="ns")

        self.left_panel_inner = ttk.Frame(self.left_panel_canvas)
        self.left_panel_window = self.left_panel_canvas.create_window(
            (0, 0),
            window=self.left_panel_inner,
            anchor="nw",
            width=160,
        )
        self.left_panel_inner.bind(
            "<Configure>",
            lambda _evt: self.left_panel_canvas.configure(scrollregion=self.left_panel_canvas.bbox("all")),
        )
        self.left_panel_canvas.bind(
            "<Configure>",
            lambda _evt: self.left_panel_canvas.itemconfigure(self.left_panel_window, width=_evt.width),
        )

        self.left_panel_inner.bind("<Enter>", self._bind_left_panel_mousewheel)
        self.left_panel_inner.bind("<Leave>", self._unbind_left_panel_mousewheel)
        self.left_panel = self.left_panel_inner

        self.center_panel = ttk.Frame(self.main_frame)
        self.center_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))

        self.right_panel = ttk.Frame(self.main_frame, width=340)
        self.right_panel.pack(side=tk.LEFT, fill=tk.Y)

        # Left panel controls
        ttk.Label(self.left_panel, text="IMAGE LOADING", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 6))
        ttk.Label(
            self.left_panel,
            text="(Up to 4 images with the same scale bar)",
            font=("Segoe UI", 8),
            foreground="#666666",
        ).pack(anchor="w", pady=(0, 6))
        ttk.Button(self.left_panel, text="IMPORT", command=self.on_import_image).pack(fill=tk.X, pady=(0, 10))

        view_frame = ttk.Frame(self.left_panel)
        view_frame.pack(fill=tk.X, pady=(0, 10))
        self._init_view_mode_icons()
        view_modes = {
            1: "single window mode",
            2: "2 window mode",
            3: "3 window mode",
            4: "4 window mode",
        }
        for mode in (1, 2, 3, 4):
            btn = tk.Button(
                view_frame,
                image=self.view_mode_icons.get(mode),
                width=28,
                height=28,
                relief="raised",
                bd=1,
                command=lambda m=mode: self.set_view_mode(m),
            )
            btn.pack(side=tk.LEFT, padx=2)
            self._attach_tooltip(btn, view_modes.get(mode, "view mode"))
            self.view_mode_buttons[mode] = btn
        self._update_view_mode_buttons()

        ttk.Separator(self.left_panel).pack(fill=tk.X, pady=6)
        ttk.Label(self.left_panel, text="AI/SCALE SETTINGS", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(4, 6))

        ckpt_frame = ttk.Frame(self.left_panel)
        ckpt_frame.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(ckpt_frame, text="Checkpoint (.pth):").pack(anchor="w")
        self.checkpoint_var = tk.StringVar(value=os.path.join(os.getcwd(), "sam_vit_b_01ec64.pth"))
        ckpt_entry = ttk.Entry(ckpt_frame, textvariable=self.checkpoint_var)
        ckpt_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=(2, 0))
        ttk.Button(ckpt_frame, text="...", width=3, command=self.on_select_checkpoint).pack(side=tk.LEFT, padx=(4, 0))

        model_frame = ttk.Frame(self.left_panel)
        model_frame.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(model_frame, text="Model type:").pack(side=tk.LEFT)
        self.model_type_var = tk.StringVar(value="vit_b")
        model_combo = ttk.Combobox(
            model_frame,
            textvariable=self.model_type_var,
            values=("vit_b", "vit_l", "vit_h"),
            width=8,
            state="readonly",
        )
        model_combo.pack(side=tk.LEFT, padx=(6, 0))

        scale_frame = ttk.Frame(self.left_panel)
        scale_frame.pack(fill=tk.X, pady=(8, 0))
        ttk.Label(scale_frame, text="Scale (um/px):").pack(side=tk.LEFT)
        self.um_per_px_var = tk.StringVar(value="0.01")
        ttk.Entry(scale_frame, textvariable=self.um_per_px_var, width=12).pack(side=tk.LEFT, padx=(6, 0))

        max_side_frame = ttk.Frame(self.left_panel)
        max_side_frame.pack(fill=tk.X, pady=(8, 0))
        ttk.Label(max_side_frame, text="Max side px (AI):").pack(side=tk.LEFT)
        self.max_side_var = tk.StringVar(value="1280")
        ttk.Entry(max_side_frame, textvariable=self.max_side_var, width=8).pack(side=tk.LEFT, padx=(6, 0))

        ttk.Button(self.left_panel, text="autocalibration", command=self.on_autocalibration).pack(
            fill=tk.X,
            pady=(6, 0),
        )

        self.fast_mode_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.left_panel, text="Fast mode (CPU)", variable=self.fast_mode_var).pack(anchor="w", pady=(6, 0))

        self.sam_run_btn = ttk.Button(self.left_panel, text="RUN AI SEGMENTATION", command=self.on_run_sam)
        self.sam_run_btn.pack(fill=tk.X, pady=(8, 4))

        self.sam_progress_var = tk.DoubleVar(value=0.0)
        self.sam_progress = ttk.Progressbar(
            self.left_panel,
            variable=self.sam_progress_var,
            mode="determinate",
            maximum=100.0,
        )
        self.sam_progress.pack(fill=tk.X, pady=(0, 4))

        self.sam_cancel_btn = ttk.Button(
            self.left_panel,
            text="CANCEL SEGMENTATION",
            command=self.on_cancel_sam,
            state="disabled",
        )
        self.sam_cancel_btn.pack(fill=tk.X, pady=(0, 10))

        ttk.Separator(self.left_panel).pack(fill=tk.X, pady=6)
        ttk.Label(self.left_panel, text="AI SEGMENTATION TOOLS", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(4, 6))
        ttk.Button(self.left_panel, text="NAVIGATE", command=lambda: self.set_mode("navigate")).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="SELECT PARTICLE", command=lambda: self.set_mode("select_one")).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="CIRCLE AREA", command=lambda: self.set_mode("roi_circle")).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="RECTANGLE AREA", command=lambda: self.set_mode("roi_rect")).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="DE-SELECTOR", command=lambda: self.set_mode("deselect_rect")).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="CLEAR SELECTION", command=self.clear_selection).pack(fill=tk.X, pady=(8, 2))
        ttk.Button(self.left_panel, text="AI EVALUATION", command=self.open_iou_tool).pack(fill=tk.X, pady=(2, 2))
        ttk.Button(self.left_panel, text="COLOR SETTINGS", command=self.open_color_settings).pack(fill=tk.X, pady=(2, 2))

        ttk.Separator(self.left_panel).pack(fill=tk.X, pady=(10, 6))
        ttk.Label(self.left_panel, text="MEASUREMENT TOOLS", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(2, 6))
        ttk.Button(self.left_panel, text="MEASURE LINE", command=lambda: self.set_mode("measure_line")).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="MEASURE ANGLE", command=lambda: self.set_mode("measure_angle")).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="OVAL SECTION", command=lambda: self.set_mode("oval_section")).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="RECTANGULAR SECTION", command=lambda: self.set_mode("rect_section")).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="FLIP IMAGE", command=self.open_flip_image).pack(fill=tk.X, pady=2)
        ttk.Button(self.left_panel, text="CLEAR MEASURE", command=self.clear_measurement).pack(fill=tk.X, pady=(2, 2))

        self.status_var = tk.StringVar(value="Ready. Import image, then run SAM.")
        ttk.Label(self.left_panel, textvariable=self.status_var, wraplength=150, foreground="#444444").pack(anchor="w", pady=(12, 0))

        # Top horizontal tools bar for future square tool buttons.
        self.top_tools_bar = ttk.Frame(self.center_panel)
        self.top_tools_bar.pack(fill=tk.X, pady=(0, 8))

        self.save_image_btn = ttk.Button(self.top_tools_bar, text="Save Image", command=self.on_save_image)
        self.save_image_btn.pack(side=tk.LEFT, padx=(0, 8), pady=(0, 2))

        self.basic_stats_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_basic_statistical_parameters,
            cursor="hand2",
        )
        self.basic_stats_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_basic_stats_icon()
        self._attach_tooltip(self.basic_stats_btn, "Basic statistical parameters")

        self.gmm_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_particle_population_analysis_gmm,
            cursor="hand2",
        )
        self.gmm_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_gmm_icon()
        self._attach_tooltip(self.gmm_btn, "Particle Population Analysis (GMM)")

        self.bootstrap_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_bootstrapping,
            cursor="hand2",
        )
        self.bootstrap_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_bootstrap_icon()
        self._attach_tooltip(self.bootstrap_btn, "Bootstrapping")

        self.dbscan_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_dbscan_clustering,
            cursor="hand2",
        )
        self.dbscan_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_dbscan_icon()
        self._attach_tooltip(self.dbscan_btn, "DBSCAN Clustering")

        self.threshold_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_threshold_tool,
            cursor="hand2",
        )
        self.threshold_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_threshold_icon()
        self._attach_tooltip(self.threshold_btn, "Threshold")

        self.hist_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_histogram,
            cursor="hand2",
        )
        self.hist_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_histogram_icon()
        self._attach_tooltip(self.hist_btn, "Histogram")

        self.nnd_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_nearest_neighbor_analysis,
            cursor="hand2",
        )
        self.nnd_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_nnd_icon()
        self._attach_tooltip(self.nnd_btn, "Nearest-Neighbor Analysis")

        self.ttest_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_compare_particle_groups_ttest,
            cursor="hand2",
        )
        self.ttest_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_ttest_icon()
        self._attach_tooltip(self.ttest_btn, "Compare Particle Groups (t-test)")

        self.false_color_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_false_color_map,
            cursor="hand2",
        )
        self.false_color_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_false_color_icon()
        self._attach_tooltip(self.false_color_btn, "False color map")

        self.violin_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_violin_box_plots,
            cursor="hand2",
        )
        self.violin_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_violin_icon()
        self._attach_tooltip(self.violin_btn, "Violin + Box Plots")

        self.bivariate_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_bivariate_analysis,
            cursor="hand2",
        )
        self.bivariate_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_bivariate_icon()
        self._attach_tooltip(self.bivariate_btn, "Bivariate Analysis")

        self.bivariate_gmm_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_bivariate_gmm_analysis,
            cursor="hand2",
        )
        self.bivariate_gmm_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_bivariate_gmm_icon()
        self._attach_tooltip(self.bivariate_gmm_btn, "Bivariate Analysis + GMM")

        self.overlay_btn = tk.Button(
            self.top_tools_bar,
            relief="raised",
            bd=1,
            command=self.open_overlay_tool,
            cursor="hand2",
        )
        self.overlay_btn.pack(side=tk.LEFT, padx=(0, 6), pady=(0, 2))
        self._load_overlay_icon()
        self._attach_tooltip(self.overlay_btn, "Overlay")

        # Center panel with matplotlib canvas
        self.figure = Figure(figsize=(10, 8), dpi=100)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title("No image loaded")
        self.ax.axis("off")

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.center_panel)
        self.canvas_widget = self.canvas.get_tk_widget()
        self.canvas_widget.pack(fill=tk.BOTH, expand=True)

        # Keep toolbar visible for standard pan/zoom behavior.
        toolbar_frame = ttk.Frame(self.center_panel)
        toolbar_frame.pack(fill=tk.X)
        self.toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
        self.toolbar.update()

        # Right panel table
        right_header = ttk.Frame(self.right_panel)
        right_header.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(right_header, text="DATA STREAM", font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT, anchor="w")
        self.show_more_btn = ttk.Button(right_header, text="Show more data", command=self.on_show_more_data)
        self.show_more_btn.pack(side=tk.LEFT, padx=(8, 0))
        self.summary_var = tk.StringVar(value="Selected: 0 | Total area: 0.000000 nm²")
        ttk.Label(right_header, textvariable=self.summary_var).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(right_header, text="Save Data", command=self.on_save_data).pack(side=tk.RIGHT)

        right_controls = ttk.Frame(self.right_panel)
        right_controls.pack(fill=tk.X, pady=(0, 8))
        self.correct_params_btn = ttk.Button(
            right_controls, text="Correct parameters", command=self.toggle_correct_parameters
        )
        self.correct_params_btn.pack(side=tk.LEFT)

        self.correct_params_panel = ttk.Frame(right_controls)
        self.correct_params_panel.pack(side=tk.LEFT, padx=(8, 0))
        self.correct_params_panel.pack_forget()
        self.correct_param_var = tk.StringVar(value="Select parameter")
        param_btn = ttk.Menubutton(self.correct_params_panel, textvariable=self.correct_param_var, width=16)
        param_menu = tk.Menu(param_btn, tearoff=0)
        param_menu.add_command(label="Length", command=lambda: self._set_correct_param("length"))
        param_menu.add_command(label="Width", command=lambda: self._set_correct_param("width"))
        param_btn["menu"] = param_menu
        param_btn.pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(self.correct_params_panel, text="Save changes", command=self._save_correct_parameters).pack(
            side=tk.LEFT
        )
        ttk.Button(self.correct_params_panel, text="Cancel", command=self._cancel_correct_parameters).pack(
            side=tk.LEFT, padx=(6, 0)
        )
        self.tree = ttk.Treeview(
            self.right_panel,
            columns=("id", "area_um2", "length_nm", "width_nm"),
            show="headings",
            height=35,
        )
        self.tree.heading("id", text="ID")
        self.tree.heading("area_um2", text="AREA (nm²)")
        self.tree.heading("length_nm", text="LENGTH (nm)")
        self.tree.heading("width_nm", text="WIDTH (nm)")
        self.tree.column("id", width=70, anchor="center")
        self.tree.column("area_um2", width=120, anchor="e")
        self.tree.column("length_nm", width=110, anchor="e")
        self.tree.column("width_nm", width=110, anchor="e")
        self._configure_data_stream_columns(show_more=False)
        tree_scroll = ttk.Scrollbar(self.right_panel, orient=tk.VERTICAL, command=self.tree.yview)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    def _connect_events(self) -> None:
        self.cid_press = self.canvas.mpl_connect("button_press_event", self.on_mouse_press)
        self.cid_move = self.canvas.mpl_connect("motion_notify_event", self.on_mouse_move)
        self.cid_release = self.canvas.mpl_connect("button_release_event", self.on_mouse_release)
        self.cid_scroll = self.canvas.mpl_connect("scroll_event", self.on_mouse_scroll)

    def toggle_correct_parameters(self) -> None:
        if self.correct_params_active:
            messagebox.showinfo("Correct parameters", "Use 'Save changes' to finish.")
            return
        if not self.selected_ids:
            messagebox.showerror("Correct parameters", "Select particles first.")
            return
        self.correct_params_active = True
        self.correct_param_name = None
        self.correct_overrides_backup = (self.length_overrides_px.copy(), self.width_overrides_px.copy())
        self.correct_line_data = {}
        self.correct_lines = {}
        if self.correct_param_var is not None:
            self.correct_param_var.set("Select parameter")
        if self.correct_params_panel is not None:
            self.correct_params_panel.pack(side=tk.LEFT, padx=(8, 0))
        if self.correct_params_btn is not None:
            try:
                self.correct_params_btn.configure(state="disabled")
            except Exception:
                pass
        self.status_var.set("Correct parameters: choose Length or Width.")
        self.render_image()

    def _set_correct_param(self, param: str) -> None:
        if not self.correct_params_active:
            return
        if param not in ("length", "width"):
            return
        self.correct_param_name = param
        if self.correct_param_var is not None:
            self.correct_param_var.set("Length" if param == "length" else "Width")
        self.correct_line_data = {}
        self.correct_lines = {}
        self.render_image()

    def _save_correct_parameters(self) -> None:
        if not self.correct_params_active:
            return
        self.correct_params_active = False
        self.correct_param_name = None
        self.correct_overrides_backup = None
        self._clear_correct_lines()
        if self.correct_params_panel is not None:
            self.correct_params_panel.pack_forget()
        if self.correct_params_btn is not None:
            try:
                self.correct_params_btn.configure(state="normal")
            except Exception:
                pass
        self.status_var.set("Parameter corrections saved.")
        self.render_image()
        self.refresh_table()

    def _cancel_correct_parameters(self) -> None:
        if not self.correct_params_active:
            return
        if self.correct_overrides_backup is not None:
            self.length_overrides_px, self.width_overrides_px = self.correct_overrides_backup
        self.correct_overrides_backup = None
        self.correct_params_active = False
        self.correct_param_name = None
        self._clear_correct_lines()
        if self.correct_params_panel is not None:
            self.correct_params_panel.pack_forget()
        if self.correct_params_btn is not None:
            try:
                self.correct_params_btn.configure(state="normal")
            except Exception:
                pass
        self.status_var.set("Parameter corrections canceled.")
        self.render_image()
        self.refresh_table()

    def _clear_correct_lines(self) -> None:
        for line in self.correct_lines.values():
            try:
                line.remove()
            except Exception:
                pass
        self.correct_lines = {}
        self.correct_line_data = {}
        self.correct_drag = None

    def _line_endpoints(self, center: Tuple[float, float], orientation: str, length_px: float) -> Tuple[float, float, float, float]:
        cx, cy = center
        half = max(0.5, float(length_px) / 2.0)
        if orientation == "h":
            return (cx - half, cy, cx + half, cy)
        return (cx, cy - half, cx, cy + half)

    def _build_correct_lines(self) -> None:
        if not self.correct_params_active or self.correct_param_name not in ("length", "width"):
            return
        self.correct_lines = {}
        self.correct_line_data = {}
        if not self.selected_ids:
            return
        for pid in sorted(self.selected_ids):
            p = self._particle_by_id(pid)
            if p is None:
                continue
            horiz, vert, cx_local, cy_local = self._center_line_components(p)
            cx, cy = self._particle_centroid_global(p)
            if self.correct_param_name == "length":
                base = max(horiz, vert)
                orientation = "h" if horiz >= vert else "v"
                length_px = self.length_overrides_px.get(pid, base)
            else:
                base = min(horiz, vert)
                orientation = "v" if horiz >= vert else "h"
                length_px = self.width_overrides_px.get(pid, base)
            length_px = max(1.0, float(length_px))
            self.correct_line_data[pid] = {
                "orientation": orientation,
                "center": (float(cx), float(cy)),
                "length": float(length_px),
            }
            x0, y0, x1, y1 = self._line_endpoints((float(cx), float(cy)), orientation, float(length_px))
            color = "#00BCD4" if self.correct_param_name == "length" else "#FFB74D"
            line = self.ax.plot([x0, x1], [y0, y1], color=color, linewidth=2.4, solid_capstyle="round")[0]
            self.correct_lines[pid] = line

    def _handle_correct_press(self, event) -> bool:
        if not self.correct_params_active or self.correct_param_name not in ("length", "width"):
            return False
        if event.inaxes != self.ax or event.xdata is None or event.ydata is None:
            return False
        if not self.correct_line_data:
            self._build_correct_lines()
        x = float(event.xdata)
        y = float(event.ydata)
        tol = 6.0
        best = None
        for pid, data in self.correct_line_data.items():
            orientation = data["orientation"]
            cx, cy = data["center"]
            length_px = data["length"]
            x0, y0, x1, y1 = self._line_endpoints((cx, cy), orientation, length_px)
            # Endpoint distances
            d_start = math.hypot(x - x0, y - y0)
            d_end = math.hypot(x - x1, y - y1)
            if d_start <= tol:
                best = (pid, "start", x0, y0, x1, y1, cx, cy, length_px, orientation)
                break
            if d_end <= tol:
                best = (pid, "end", x0, y0, x1, y1, cx, cy, length_px, orientation)
                break
            # Line distance
            if orientation == "h":
                if min(x0, x1) - tol <= x <= max(x0, x1) + tol and abs(y - cy) <= tol:
                    best = (pid, "move", x0, y0, x1, y1, cx, cy, length_px, orientation)
                    break
            else:
                if min(y0, y1) - tol <= y <= max(y0, y1) + tol and abs(x - cx) <= tol:
                    best = (pid, "move", x0, y0, x1, y1, cx, cy, length_px, orientation)
                    break
        if best is None:
            return False
        pid, part, x0, y0, x1, y1, cx, cy, length_px, orientation = best
        self.correct_drag = {
            "pid": pid,
            "part": part,
            "start_x": x,
            "start_y": y,
            "x0": x0,
            "y0": y0,
            "x1": x1,
            "y1": y1,
            "center": (cx, cy),
            "length": length_px,
            "orientation": orientation,
        }
        return True

    def _handle_correct_move(self, event) -> bool:
        if self.correct_drag is None:
            return False
        if event.inaxes != self.ax or event.xdata is None or event.ydata is None:
            return True
        x = float(event.xdata)
        y = float(event.ydata)
        pid = int(self.correct_drag["pid"])
        part = self.correct_drag["part"]
        orientation = self.correct_drag["orientation"]
        x0 = float(self.correct_drag["x0"])
        y0 = float(self.correct_drag["y0"])
        x1 = float(self.correct_drag["x1"])
        y1 = float(self.correct_drag["y1"])
        cx0, cy0 = self.correct_drag["center"]
        length_px = float(self.correct_drag["length"])

        if part == "move":
            dx = x - float(self.correct_drag["start_x"])
            dy = y - float(self.correct_drag["start_y"])
            cx = cx0 + dx
            cy = cy0 + dy
            new_length = length_px
        else:
            if orientation == "h":
                if part == "start":
                    x0 = x
                else:
                    x1 = x
                new_length = max(1.0, abs(x1 - x0))
                cx = 0.5 * (x0 + x1)
                cy = cy0
            else:
                if part == "start":
                    y0 = y
                else:
                    y1 = y
                new_length = max(1.0, abs(y1 - y0))
                cx = cx0
                cy = 0.5 * (y0 + y1)

        self.correct_line_data[pid] = {"orientation": orientation, "center": (cx, cy), "length": new_length}
        if pid in self.correct_lines:
            nx0, ny0, nx1, ny1 = self._line_endpoints((cx, cy), orientation, new_length)
            self.correct_lines[pid].set_data([nx0, nx1], [ny0, ny1])
            self.canvas.draw_idle()

        if self.correct_param_name == "length":
            self.length_overrides_px[pid] = float(new_length)
        else:
            self.width_overrides_px[pid] = float(new_length)
        self.refresh_table()
        return True

    def _handle_correct_release(self, _event) -> bool:
        if self.correct_drag is None:
            return False
        self.correct_drag = None
        return True

    def _bind_left_panel_mousewheel(self, _event: Optional[tk.Event] = None) -> None:
        if self.left_panel_canvas is None:
            return
        self.left_panel_canvas.bind_all("<MouseWheel>", self._on_left_panel_mousewheel)

    def _unbind_left_panel_mousewheel(self, _event: Optional[tk.Event] = None) -> None:
        if self.left_panel_canvas is None:
            return
        self.left_panel_canvas.unbind_all("<MouseWheel>")

    def _on_left_panel_mousewheel(self, event: tk.Event) -> None:
        if self.left_panel_canvas is None:
            return
        self.left_panel_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def on_show_more_data(self) -> None:
        if not self.show_more_data:
            self.show_more_data = True
            if self.show_more_btn is not None:
                try:
                    self.show_more_btn.configure(text="Show less data")
                except Exception:
                    pass
            self._configure_data_stream_columns(show_more=True)
            self._expand_window_for_more_data()
        else:
            self.show_more_data = False
            if self.show_more_btn is not None:
                try:
                    self.show_more_btn.configure(text="Show more data")
                except Exception:
                    pass
            self._configure_data_stream_columns(show_more=False)
            self._collapse_window_for_more_data()
        self.refresh_table()

    def _expand_window_for_more_data(self) -> None:
        try:
            self.root.update_idletasks()
            w = self.root.winfo_width()
            h = self.root.winfo_height()
            x = self.root.winfo_x()
            y = self.root.winfo_y()
            extra = 420
            new_w = w + extra
            new_x = max(0, x - extra)
            self.root.geometry(f"{new_w}x{h}+{new_x}+{y}")
        except Exception:
            pass
        try:
            self.right_panel.configure(width=760)
        except Exception:
            pass

    def _collapse_window_for_more_data(self) -> None:
        try:
            self.root.update_idletasks()
            w = self.root.winfo_width()
            h = self.root.winfo_height()
            x = self.root.winfo_x()
            y = self.root.winfo_y()
            extra = 420
            new_w = max(self.base_window_width, w - extra)
            shift = w - new_w
            new_x = x + shift
            self.root.geometry(f"{new_w}x{h}+{new_x}+{y}")
        except Exception:
            pass
        try:
            self.right_panel.configure(width=340)
        except Exception:
            pass

    def _configure_data_stream_columns(self, show_more: bool) -> None:
        if show_more:
            columns = (
                "id",
                "area_um2",
                "length_nm",
                "width_nm",
                "circularity",
                "eccentricity",
                "feret_nm",
            )
        else:
            columns = ("id", "area_um2", "length_nm", "width_nm")

        self.tree.configure(columns=columns)
        self.tree.heading("id", text="ID")
        self.tree.heading("area_um2", text="AREA (nm²)")
        self.tree.heading("length_nm", text="LENGTH (nm)")
        self.tree.heading("width_nm", text="WIDTH (nm)")
        self.tree.column("id", width=70, anchor="center")
        self.tree.column("area_um2", width=120, anchor="e")
        self.tree.column("length_nm", width=110, anchor="e")
        self.tree.column("width_nm", width=110, anchor="e")
        if show_more:
            self.tree.heading("circularity", text="CIRCULARITY")
            self.tree.heading("eccentricity", text="ECCENTRICITY")
            self.tree.heading("feret_nm", text="FERET DIAMETER (nm)")
            self.tree.column("circularity", width=110, anchor="e")
            self.tree.column("eccentricity", width=110, anchor="e")
            self.tree.column("feret_nm", width=140, anchor="e")

    def _init_view_mode_icons(self) -> None:
        if self.view_mode_icons:
            return
        for mode in (1, 2, 3, 4):
            self.view_mode_icons[mode] = self._make_view_mode_icon(mode)

    def _make_view_mode_icon(self, mode: int, size: int = 28) -> ImageTk.PhotoImage:
        pad = 3
        gap = 2
        img = Image.new("RGBA", (size, size), (255, 255, 255, 0))
        draw = ImageDraw.Draw(img)
        stroke = (70, 70, 70, 255)
        fill = (120, 120, 120, 255)

        def cell_rect(r: int, c: int) -> Tuple[int, int, int, int]:
            cell_size = (size - (2 * pad) - gap) // 2
            x0 = pad + c * (cell_size + gap)
            y0 = pad + r * (cell_size + gap)
            x1 = x0 + cell_size
            y1 = y0 + cell_size
            return x0, y0, x1, y1

        if mode == 1:
            draw.rectangle((pad, pad, size - pad, size - pad), outline=stroke, width=2, fill=fill)
        elif mode == 2:
            cell_w = (size - (2 * pad) - gap)
            half = cell_w // 2
            draw.rectangle((pad, pad, pad + half, size - pad), outline=stroke, width=1, fill=fill)
            draw.rectangle(
                (pad + half + gap, pad, size - pad, size - pad),
                outline=stroke,
                width=1,
                fill=fill,
            )
            draw.rectangle((pad, pad, size - pad, size - pad), outline=stroke, width=2)
        else:
            filled = {(0, 0), (0, 1), (1, 0), (1, 1)}
            if mode == 3:
                filled.remove((1, 1))
            for r in range(2):
                for c in range(2):
                    rect = cell_rect(r, c)
                    if (r, c) in filled:
                        draw.rectangle(rect, outline=stroke, width=1, fill=fill)
                    else:
                        draw.rectangle(rect, outline=stroke, width=1, fill=None)
            draw.rectangle((pad, pad, size - pad, size - pad), outline=stroke, width=2)

        return ImageTk.PhotoImage(img)

    def _update_view_mode_buttons(self) -> None:
        max_mode = len(self.image_rgbs)
        for mode, btn in self.view_mode_buttons.items():
            relief = "sunken" if mode == self.view_mode else "raised"
            try:
                state = "normal" if max_mode == 0 or mode <= max_mode else "disabled"
                btn.configure(relief=relief, state=state)
            except Exception:
                pass

    def _active_image_count(self) -> int:
        if not self.image_rgbs:
            return 0
        return max(1, min(self.view_mode, len(self.image_rgbs)))

    def _active_image_indices(self) -> List[int]:
        return list(range(self._active_image_count()))

    def _active_images(self) -> List[np.ndarray]:
        count = self._active_image_count()
        return self.image_rgbs[:count]

    def _compose_active_images(self) -> None:
        if not self.image_rgbs:
            self.image_rgb = None
            self.image_offsets = {}
            self.image_cell_size = 0
            return

        count = self._active_image_count()
        images = self.image_rgbs[:count]

        if count <= 1:
            rows, cols = 1, 1
        elif count == 2:
            rows, cols = 1, 2
        else:
            rows, cols = 2, 2

        cell_size = max(max(img.shape[0], img.shape[1]) for img in images)
        self.image_cell_size = int(cell_size)

        canvas_h = rows * cell_size
        canvas_w = cols * cell_size
        composite = np.zeros((canvas_h, canvas_w, 3), dtype=np.uint8)
        offsets: Dict[int, Tuple[int, int]] = {}

        for idx, img in enumerate(images):
            row = 0 if rows == 1 else idx // 2
            col = idx if rows == 1 else idx % 2
            h, w = img.shape[:2]
            off_x = int(col * cell_size + max(0, (cell_size - w) // 2))
            off_y = int(row * cell_size + max(0, (cell_size - h) // 2))
            composite[off_y : off_y + h, off_x : off_x + w] = img
            offsets[idx] = (off_x, off_y)

        # Draw visible divider lines between cells.
        divider_color = np.array([180, 180, 180], dtype=np.uint8)
        line_thickness = 2
        if cols > 1:
            x = cell_size
            composite[:, max(0, x - line_thickness // 2) : min(canvas_w, x + line_thickness // 2 + 1)] = divider_color
        if rows > 1:
            y = cell_size
            composite[max(0, y - line_thickness // 2) : min(canvas_h, y + line_thickness // 2 + 1), :] = divider_color

        self.image_rgb = composite
        self.image_offsets = offsets
        self.image_path = self.image_paths[0] if count == 1 and self.image_paths else None

    def _sync_particles_for_layout(self) -> None:
        active_indices = set(self._active_image_indices())
        self.particles = []
        for idx in active_indices:
            for p in self.particles_by_image.get(idx, []):
                p.offset_xy = self.image_offsets.get(idx, (0, 0))
                self.particles.append(p)

        valid_ids = {p.mask_id for p in self.particles}
        self.selected_ids &= valid_ids
        self.highlighted_ids &= self.selected_ids
        self._build_pick_map()

    def set_view_mode(self, mode: int) -> None:
        if mode not in (1, 2, 3, 4):
            return
        max_mode = min(4, len(self.image_rgbs)) if self.image_rgbs else 4
        self.view_mode = max(1, min(mode, max_mode))
        self._update_view_mode_buttons()
        if not self.image_rgbs:
            return
        self._compose_active_images()
        self._sync_particles_for_layout()
        self.render_image()
        self.refresh_table()

    def _load_basic_stats_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "BASICSTATISTICALPARAMETERS.png")
        if not os.path.isfile(icon_path):
            self.basic_stats_btn.configure(text="BSP")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.basic_stats_icon_img = ImageTk.PhotoImage(icon)
            self.basic_stats_btn.configure(image=self.basic_stats_icon_img, text="")
        except Exception:
            self.basic_stats_btn.configure(text="BSP")

    def _load_gmm_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "GMM.png")
        if not os.path.isfile(icon_path):
            self.gmm_btn.configure(text="GMM")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.gmm_icon_img = ImageTk.PhotoImage(icon)
            self.gmm_btn.configure(image=self.gmm_icon_img, text="")
        except Exception:
            self.gmm_btn.configure(text="GMM")

    def _load_bootstrap_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "Bootstrap.png")
        if not os.path.isfile(icon_path):
            self.bootstrap_btn.configure(text="BOOT")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.bootstrap_icon_img = ImageTk.PhotoImage(icon)
            self.bootstrap_btn.configure(image=self.bootstrap_icon_img, text="")
        except Exception:
            self.bootstrap_btn.configure(text="BOOT")

    def _load_dbscan_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, ".DBSCAN.png")
        if not os.path.isfile(icon_path):
            icon_path = os.path.join(script_dir, "DBSCAN.png")
        if not os.path.isfile(icon_path):
            self.dbscan_btn.configure(text="DBS")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.dbscan_icon_img = ImageTk.PhotoImage(icon)
            self.dbscan_btn.configure(image=self.dbscan_icon_img, text="")
        except Exception:
            self.dbscan_btn.configure(text="DBS")

    def _load_threshold_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "Threshold.png")
        if not os.path.isfile(icon_path):
            self.threshold_btn.configure(text="THR")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.threshold_icon_img = ImageTk.PhotoImage(icon)
            self.threshold_btn.configure(image=self.threshold_icon_img, text="")
        except Exception:
            self.threshold_btn.configure(text="THR")

    def _load_histogram_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "Histogram.png")
        if not os.path.isfile(icon_path):
            self.hist_btn.configure(text="HIST")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.hist_icon_img = ImageTk.PhotoImage(icon)
            self.hist_btn.configure(image=self.hist_icon_img, text="")
        except Exception:
            self.hist_btn.configure(text="HIST")

    def _load_nnd_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "NND.png")
        if not os.path.isfile(icon_path):
            self.nnd_btn.configure(text="NND")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.nnd_icon_img = ImageTk.PhotoImage(icon)
            self.nnd_btn.configure(image=self.nnd_icon_img, text="")
        except Exception:
            self.nnd_btn.configure(text="NND")

    def _load_ttest_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "ttest.png")
        if not os.path.isfile(icon_path):
            self.ttest_btn.configure(text="TTEST")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.ttest_icon_img = ImageTk.PhotoImage(icon)
            self.ttest_btn.configure(image=self.ttest_icon_img, text="")
        except Exception:
            self.ttest_btn.configure(text="TTEST")

    def _load_false_color_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "Colormap.png")
        if not os.path.isfile(icon_path):
            self.false_color_btn.configure(text="FC")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.false_color_icon_img = ImageTk.PhotoImage(icon)
            self.false_color_btn.configure(image=self.false_color_icon_img, text="")
        except Exception:
            self.false_color_btn.configure(text="FC")

    def _load_violin_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "Boxplot.png")
        if not os.path.isfile(icon_path):
            self.violin_btn.configure(text="VB")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.violin_icon_img = ImageTk.PhotoImage(icon)
            self.violin_btn.configure(image=self.violin_icon_img, text="")
        except Exception:
            self.violin_btn.configure(text="VB")

    def _load_bivariate_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "Bivariate.png")
        if not os.path.isfile(icon_path):
            self.bivariate_btn.configure(text="BIV")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.bivariate_icon_img = ImageTk.PhotoImage(icon)
            self.bivariate_btn.configure(image=self.bivariate_icon_img, text="")
        except Exception:
            self.bivariate_btn.configure(text="BIV")

    def _load_bivariate_gmm_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "Bivariate2.png")
        if not os.path.isfile(icon_path):
            self.bivariate_gmm_btn.configure(text="BIV+")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.bivariate_gmm_icon_img = ImageTk.PhotoImage(icon)
            self.bivariate_gmm_btn.configure(image=self.bivariate_gmm_icon_img, text="")
        except Exception:
            self.bivariate_gmm_btn.configure(text="BIV+")

    def _load_overlay_icon(self) -> None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "Overlay.png")
        if not os.path.isfile(icon_path):
            self.overlay_btn.configure(text="OVR")
            return
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((56, 56), Image.Resampling.LANCZOS)
            self.overlay_icon_img = ImageTk.PhotoImage(icon)
            self.overlay_btn.configure(image=self.overlay_icon_img, text="")
        except Exception:
            self.overlay_btn.configure(text="OVR")

    def _attach_tooltip(self, widget: tk.Widget, text: str) -> None:
        tooltip = {"window": None}

        def show(event) -> None:
            if tooltip["window"] is not None:
                return
            tw = tk.Toplevel(widget)
            tw.wm_overrideredirect(True)
            tw.wm_geometry(f"+{event.x_root + 12}+{event.y_root + 12}")
            label = tk.Label(
                tw,
                text=text,
                background="#111111",
                foreground="white",
                relief="solid",
                borderwidth=1,
                padx=6,
                pady=4,
            )
            label.pack()
            tooltip["window"] = tw

        def hide(_event=None) -> None:
            if tooltip["window"] is not None:
                tooltip["window"].destroy()
                tooltip["window"] = None

        widget.bind("<Enter>", show)
        widget.bind("<Leave>", hide)
        widget.bind("<ButtonPress>", hide)

    def _get_help_icon(self) -> Optional[ImageTk.PhotoImage]:
        if self.help_icon_img is not None:
            return self.help_icon_img
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "button_help.ico")
        if not os.path.isfile(icon_path):
            return None
        try:
            icon = Image.open(icon_path).convert("RGBA")
            icon = icon.resize((18, 18), Image.Resampling.LANCZOS)
            self.help_icon_img = ImageTk.PhotoImage(icon)
            return self.help_icon_img
        except Exception:
            return None

    def open_basic_statistical_parameters(self) -> None:
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return

        if self.basic_stats_window is not None and self.basic_stats_window.winfo_exists():
            self.basic_stats_window.lift()
            self._refresh_basic_stats_window()
            return

        self.basic_stats_window = tk.Toplevel(self.root)
        self.basic_stats_window.title("Basic Statistical Parameters")
        self.basic_stats_window.geometry("450x340+980+180")

        outer = ttk.Frame(self.basic_stats_window, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            outer,
            text="Basic Statistical Parameters",
            font=("Segoe UI", 12, "bold"),
        ).pack(anchor="w", pady=(0, 8))

        controls = ttk.Frame(outer)
        controls.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(controls, text="Property:").pack(side=tk.LEFT)
        self.basic_stats_property_var = tk.StringVar(value="length")
        prop_combo = ttk.Combobox(
            controls,
            textvariable=self.basic_stats_property_var,
            state="readonly",
            values=("length", "width", "area", "feret diameter", "circularity", "eccentricity"),
            width=12,
        )
        prop_combo.pack(side=tk.LEFT, padx=(8, 0))
        prop_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_basic_stats_window())
        ttk.Button(controls, text="Save Data", command=self.on_save_basic_stats_data).pack(side=tk.RIGHT)

        self.basic_stats_basis_var = tk.StringVar(value="")
        ttk.Label(
            outer,
            textvariable=self.basic_stats_basis_var,
            font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(0, 10))

        stats_grid = ttk.Frame(outer)
        stats_grid.pack(fill=tk.X)

        self.basic_stats_value_vars = {}
        stat_names = ("mean", "median", "std", "variance", "CV")
        for i, name in enumerate(stat_names):
            self.basic_stats_value_vars[name] = tk.StringVar(value="N/A")
            ttk.Label(stats_grid, text=name, width=12).grid(row=i, column=0, sticky="w", pady=2)
            ttk.Label(stats_grid, textvariable=self.basic_stats_value_vars[name], width=24).grid(
                row=i,
                column=1,
                sticky="e",
                pady=2,
            )

        self.basic_stats_window.protocol("WM_DELETE_WINDOW", self._close_basic_stats_window)
        self._refresh_basic_stats_window()

    def _close_basic_stats_window(self) -> None:
        if self.basic_stats_window is not None and self.basic_stats_window.winfo_exists():
            self.basic_stats_window.destroy()
        self.basic_stats_window = None
        self.basic_stats_property_var = None
        self.basic_stats_basis_var = None
        self.basic_stats_value_vars = {}

    def _selected_metric_values(self, prop: str) -> Tuple[np.ndarray, str, str]:
        nm_per_px = self._safe_um_per_px_for_refresh() * 1000.0
        values: List[float] = []
        for p in self._selected_particles():
            length_px, width_px = self._center_length_width_px(p)
            if prop == "width":
                values.append(width_px * nm_per_px)
            elif prop == "area":
                values.append(float(p.area_px) * (nm_per_px ** 2))
            elif prop == "feret diameter":
                values.append(self._feret_diameter_px(p) * nm_per_px)
            elif prop == "circularity":
                values.append(self._circularity_eccentricity(p)[0])
            elif prop == "eccentricity":
                values.append(self._circularity_eccentricity(p)[1])
            else:
                values.append(length_px * nm_per_px)

        arr = np.asarray(values, dtype=float)
        if prop == "width":
            return arr, "nm", "width (nm)"
        if prop == "area":
            return arr, "nm²", "area (nm²)"
        if prop == "feret diameter":
            return arr, "nm", "feret diameter (nm)"
        if prop == "circularity":
            return arr, "", "circularity"
        if prop == "eccentricity":
            return arr, "", "eccentricity"
        return arr, "nm", "length (nm)"

    def _variance_unit(self, base_unit: str) -> str:
        if base_unit == "":
            return ""
        if base_unit == "nm":
            return "nm²"
        if base_unit == "nm^2":
            return "nm^4"
        if base_unit == "nm²":
            return "nm⁴"
        return f"{base_unit}^2"

    def _refresh_basic_stats_window(self) -> None:
        if self.basic_stats_window is None or not self.basic_stats_window.winfo_exists():
            return
        if self.basic_stats_property_var is None or self.basic_stats_basis_var is None:
            return
        if not self.basic_stats_value_vars:
            return

        prop = self.basic_stats_property_var.get().strip().lower()
        values, unit, basis = self._selected_metric_values(prop)
        self.basic_stats_basis_var.set(f"Based on selected particle {basis}")

        if values.size == 0:
            for name in ("mean", "median", "std", "variance", "CV"):
                self.basic_stats_value_vars[name].set("N/A")
            return

        mean_val = float(np.mean(values))
        median_val = float(np.median(values))
        std_val = float(np.std(values))
        var_val = float(np.var(values))
        cv_val = float((std_val / mean_val) * 100.0) if mean_val != 0.0 else float("nan")

        var_unit = self._variance_unit(unit)
        def fmt(val: float, u: str) -> str:
            return f"{val:.4f} {u}".strip() if u else f"{val:.4f}"

        self.basic_stats_value_vars["mean"].set(fmt(mean_val, unit))
        self.basic_stats_value_vars["median"].set(fmt(median_val, unit))
        self.basic_stats_value_vars["std"].set(fmt(std_val, unit))
        self.basic_stats_value_vars["variance"].set(fmt(var_val, var_unit))
        if np.isfinite(cv_val):
            self.basic_stats_value_vars["CV"].set(f"{cv_val:.2f} %")
        else:
            self.basic_stats_value_vars["CV"].set("N/A")

    def on_save_basic_stats_data(self) -> None:
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return

        prop = "length"
        if self.basic_stats_property_var is not None:
            prop = self.basic_stats_property_var.get().strip().lower() or "length"

        values, unit, _basis = self._selected_metric_values(prop)
        if values.size == 0:
            messagebox.showerror("Error", "Select particles")
            return

        mean_val = float(np.mean(values))
        median_val = float(np.median(values))
        std_val = float(np.std(values))
        var_val = float(np.var(values))
        cv_val = float((std_val / mean_val) * 100.0) if mean_val != 0.0 else float("nan")
        var_unit = self._variance_unit(unit)

        def header(name: str, u: str) -> str:
            return f"{name} ({u})" if u else name

        headers = (
            header("Mean", unit),
            header("Median", unit),
            header("Std", unit),
            header("Variance", var_unit),
            "CV (%)",
        )
        row = (
            f"{mean_val:.6f}",
            f"{median_val:.6f}",
            f"{std_val:.6f}",
            f"{var_val:.6f}",
            f"{cv_val:.4f}" if np.isfinite(cv_val) else "N/A",
        )
        self._save_rows_to_csv_or_xlsx(
            headers=headers,
            rows=[row],
            dialog_title="Save Basic Statistical Parameters",
            initialfile="basic_statistical_parameters.csv",
            sheet_name="Basic Stats",
        )

    def open_particle_population_analysis_gmm(self) -> None:
        if not GMM_AVAILABLE:
            messagebox.showerror("GMM Error", f"GMM dependencies are missing.\n{GMM_IMPORT_ERROR}")
            return
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return

        if self.gmm_window is not None and self.gmm_window.winfo_exists():
            self.gmm_window.lift()
            self._refresh_gmm_window()
            return

        self.gmm_window = tk.Toplevel(self.root)
        self.gmm_window.title("Particle Population Analysis (GMM)")
        self.gmm_window.geometry("1080x470+920+170")

        outer = ttk.Frame(self.gmm_window, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(outer)
        header.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(
            header,
            text="Particle Population Analysis (GMM)",
            font=("Segoe UI", 12, "bold"),
        ).pack(side=tk.LEFT, anchor="w")
        help_icon = self._get_help_icon()
        self.gmm_help_btn = tk.Button(
            header,
            image=help_icon,
            text="" if help_icon is not None else "?",
            relief="flat",
            bd=0,
            command=self._toggle_gmm_help,
            cursor="hand2",
        )
        self.gmm_help_btn.pack(side=tk.LEFT, padx=(6, 0))
        if help_icon is None:
            self.gmm_help_btn.configure(font=("Segoe UI", 9, "bold"))

        controls = ttk.Frame(outer)
        controls.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(controls, text="Property:").pack(side=tk.LEFT)
        self.gmm_property_var = tk.StringVar(value="length")
        combo = ttk.Combobox(
            controls,
            textvariable=self.gmm_property_var,
            state="readonly",
            values=("length", "width", "area", "feret diameter", "circularity", "eccentricity"),
            width=12,
        )
        combo.pack(side=tk.LEFT, padx=(8, 0))
        combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_gmm_window())
        ttk.Button(controls, text="Save Data", command=self.on_save_gmm_data).pack(side=tk.RIGHT)

        self.gmm_basis_var = tk.StringVar(value="")
        ttk.Label(outer, textvariable=self.gmm_basis_var, font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 8))

        table_frame = ttk.Frame(outer)
        table_frame.pack(fill=tk.BOTH, expand=True)

        self.gmm_tree = ttk.Treeview(
            table_frame,
            columns=("subpop", "n", "mean", "median", "std", "variance", "cv", "skewness", "kurtosis"),
            show="headings",
        )
        self.gmm_tree.heading("subpop", text="Subpop")
        self.gmm_tree.heading("n", text="N")
        self.gmm_tree.heading("mean", text="Mean")
        self.gmm_tree.heading("median", text="Median")
        self.gmm_tree.heading("std", text="Std")
        self.gmm_tree.heading("variance", text="Variance")
        self.gmm_tree.heading("cv", text="CV (%)")
        self.gmm_tree.heading("skewness", text="Skewness")
        self.gmm_tree.heading("kurtosis", text="Kurtosis")

        self.gmm_tree.column("subpop", width=80, anchor="center")
        self.gmm_tree.column("n", width=60, anchor="center")
        self.gmm_tree.column("mean", width=120, anchor="e")
        self.gmm_tree.column("median", width=120, anchor="e")
        self.gmm_tree.column("std", width=120, anchor="e")
        self.gmm_tree.column("variance", width=130, anchor="e")
        self.gmm_tree.column("cv", width=90, anchor="e")
        self.gmm_tree.column("skewness", width=110, anchor="e")
        self.gmm_tree.column("kurtosis", width=110, anchor="e")
        self.gmm_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        yscroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.gmm_tree.yview)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.gmm_tree.configure(yscrollcommand=yscroll.set)

        self.gmm_window.protocol("WM_DELETE_WINDOW", self._close_gmm_window)
        self._refresh_gmm_window()

    def _close_gmm_window(self) -> None:
        if self.gmm_window is not None and self.gmm_window.winfo_exists():
            self.gmm_window.destroy()
        self.gmm_window = None
        self.gmm_property_var = None
        self.gmm_basis_var = None
        self.gmm_tree = None
        self.gmm_help_btn = None
        if self.gmm_help_window is not None and self.gmm_help_window.winfo_exists():
            self.gmm_help_window.destroy()
        self.gmm_help_window = None
        self.gmm_help_tracking = False

    def _toggle_gmm_help(self) -> None:
        if self.gmm_window is None or not self.gmm_window.winfo_exists():
            return
        if self.gmm_help_window is not None and self.gmm_help_window.winfo_exists():
            self.gmm_help_window.destroy()
            self.gmm_help_window = None
            self.gmm_help_tracking = False
            return

        self.gmm_help_window = tk.Toplevel(self.gmm_window)
        self.gmm_help_window.overrideredirect(True)
        self.gmm_help_window.attributes("-topmost", True)

        bubble = tk.Frame(self.gmm_help_window, bg="#fff7e6", bd=1, relief="solid")
        bubble.pack(fill=tk.BOTH, expand=True)

        text_1 = (
            "Skewness — measures asymmetry of the distribution. A positive skew (right tail) "
            "means a minority of larger particles or aggregates pull the tail to the right; "
            "a negative skew (left tail) means more very small particles; a value near zero "
            "indicates a roughly symmetric size distribution."
        )
        text_2 = (
            "Kurtosis — quantifies the \"peakedness\" and weight of the tails relative to a normal distribution. "
            "High kurtosis (leptokurtic) implies a sharp peak plus heavier tails (many particles near the median "
            "and more extreme outliers, e.g. aggregates), while low kurtosis (platykurtic) indicates a flatter "
            "distribution with sizes more evenly spread."
        )
        p1 = tk.Text(
            bubble,
            bg="#fff7e6",
            fg="#333333",
            wrap="word",
            width=42,
            height=5,
            relief="flat",
            highlightthickness=0,
        )
        p1.tag_configure("title", font=("Segoe UI", 9, "bold"))
        p1.tag_configure("body", font=("Segoe UI", 9))
        p1.insert("1.0", "Skewness —", "title")
        p1.insert("end", " measures asymmetry of the distribution. A positive skew (right tail) ", "body")
        p1.insert("end", "means a minority of larger particles or aggregates pull the tail to the right; ", "body")
        p1.insert("end", "a negative skew (left tail) means more very small particles; a value near zero ", "body")
        p1.insert("end", "indicates a roughly symmetric size distribution.", "body")
        p1.configure(state="disabled")
        p1.pack(anchor="w", padx=8, pady=(8, 4))

        p2 = tk.Text(
            bubble,
            bg="#fff7e6",
            fg="#333333",
            wrap="word",
            width=42,
            height=6,
            relief="flat",
            highlightthickness=0,
        )
        p2.tag_configure("title", font=("Segoe UI", 9, "bold"))
        p2.tag_configure("body", font=("Segoe UI", 9))
        p2.insert("1.0", "Kurtosis —", "title")
        p2.insert("end", " quantifies the \"peakedness\" and weight of the tails relative to a normal distribution. ", "body")
        p2.insert("end", "High kurtosis (leptokurtic) implies a sharp peak plus heavier tails (many particles near the median ", "body")
        p2.insert("end", "and more extreme outliers, e.g. aggregates), while low kurtosis (platykurtic) indicates a flatter ", "body")
        p2.insert("end", "distribution with sizes more evenly spread.", "body")
        p2.configure(state="disabled")
        p2.pack(anchor="w", padx=8, pady=(0, 8))

        self.gmm_help_tracking = True
        if self.gmm_window is not None:
            self.gmm_window.bind("<Configure>", self._on_gmm_window_configure)
        self._position_gmm_help()

    def _on_gmm_window_configure(self, _event=None) -> None:
        if self.gmm_help_tracking:
            self._position_gmm_help()

    def _position_gmm_help(self) -> None:
        if self.gmm_help_window is None or not self.gmm_help_window.winfo_exists():
            return
        if self.gmm_window is None or not self.gmm_window.winfo_exists():
            return
        self.gmm_help_window.update_idletasks()
        bubble_w = self.gmm_help_window.winfo_width()
        bubble_h = self.gmm_help_window.winfo_height()
        win_x = self.gmm_window.winfo_rootx()
        win_y = self.gmm_window.winfo_rooty()
        x = win_x + 20
        y = win_y - bubble_h - 8
        if y < 0:
            y = win_y + 30
        self.gmm_help_window.geometry(f"{bubble_w}x{bubble_h}+{x}+{y}")

    def _fit_best_gmm(self, values: np.ndarray) -> Tuple[Optional[GaussianMixture], Optional[np.ndarray]]:
        x = values.reshape(-1, 1).astype(float)
        n_samples = x.shape[0]
        if n_samples == 0:
            return None, None

        max_components = min(6, n_samples)
        best_model: Optional[GaussianMixture] = None
        best_bic = float("inf")

        for n_comp in range(1, max_components + 1):
            try:
                gmm = GaussianMixture(
                    n_components=n_comp,
                    covariance_type="full",
                    random_state=0,
                    n_init=5,
                )
                gmm.fit(x)
                bic = float(gmm.bic(x))
                if bic < best_bic:
                    best_bic = bic
                    best_model = gmm
            except Exception:
                continue

        if best_model is None:
            return None, None
        labels = best_model.predict(x)
        return best_model, labels

    def _refresh_gmm_window(self) -> None:
        if self.gmm_window is None or not self.gmm_window.winfo_exists():
            return
        if self.gmm_tree is None or self.gmm_property_var is None or self.gmm_basis_var is None:
            return

        prop = self.gmm_property_var.get().strip().lower()
        values, unit, basis = self._selected_metric_values(prop)
        self.gmm_basis_var.set(f"Based on selected particle {basis}")

        for row in self.gmm_tree.get_children():
            self.gmm_tree.delete(row)

        if values.size == 0:
            return

        var_unit = self._variance_unit(unit)
        def unit_heading(name: str, u: str) -> str:
            return f"{name} ({u})" if u else name

        self.gmm_tree.heading("mean", text=unit_heading("Mean", unit))
        self.gmm_tree.heading("median", text=unit_heading("Median", unit))
        self.gmm_tree.heading("std", text=unit_heading("Std", unit))
        self.gmm_tree.heading("variance", text=unit_heading("Variance", var_unit))

        model, labels = self._fit_best_gmm(values)
        if model is None or labels is None:
            self.gmm_tree.insert(
                "",
                tk.END,
                values=("N/A", "-", "-", "-", "-", "-", "-", "-", "-"),
            )
            return

        means = np.asarray(model.means_, dtype=float).reshape(-1)
        order = np.argsort(means)

        for rank, comp_idx in enumerate(order, start=1):
            sub_values = values[labels == comp_idx]
            if sub_values.size == 0:
                continue

            mean_gmm = float(means[comp_idx])
            cov = model.covariances_[comp_idx]
            if np.ndim(cov) == 2:
                var_gmm = float(cov[0, 0])
            elif np.ndim(cov) == 1:
                var_gmm = float(cov[0])
            else:
                var_gmm = float(cov)

            median_val = float(np.median(sub_values))
            std_val = float(np.std(sub_values))
            cv_val = float((std_val / mean_gmm) * 100.0) if mean_gmm != 0.0 else float("nan")
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", RuntimeWarning)
                skew_val = float(skew(sub_values, bias=False)) if sub_values.size >= 3 else float("nan")
                kurt_val = float(kurtosis(sub_values, fisher=True, bias=False)) if sub_values.size >= 4 else float("nan")

            def fmt(val: float, dec: int = 4) -> str:
                return f"{val:.{dec}f}" if np.isfinite(val) else "N/A"

            self.gmm_tree.insert(
                "",
                tk.END,
                values=(
                    f"#{rank}",
                    int(sub_values.size),
                    fmt(mean_gmm, 4),
                    fmt(median_val, 4),
                    fmt(std_val, 4),
                    fmt(var_gmm, 4),
                    fmt(cv_val, 2),
                    fmt(skew_val, 4),
                    fmt(kurt_val, 4),
                ),
            )

    def on_save_gmm_data(self) -> None:
        if self.gmm_tree is None:
            messagebox.showerror("Error", "Open GMM window first.")
            return

        item_ids = list(self.gmm_tree.get_children())
        if not item_ids:
            messagebox.showerror("Error", "Select particles")
            return

        columns = ("subpop", "n", "mean", "median", "std", "variance", "cv", "skewness", "kurtosis")
        headers = tuple(str(self.gmm_tree.heading(col, "text")) for col in columns)

        rows = []
        for item_id in item_ids:
            vals = self.gmm_tree.item(item_id, "values")
            if vals:
                rows.append(tuple(vals[: len(columns)]))

        if not rows:
            messagebox.showerror("Error", "Select particles")
            return
        headers = tuple(str(self.gmm_tree.heading(col, "text")) for col in columns)

        self._save_rows_to_csv_or_xlsx(
            headers=headers,
            rows=rows,
            dialog_title="Save GMM Data",
            initialfile="gmm_data.csv",
            sheet_name="GMM",
        )

    def open_dbscan_clustering(self) -> None:
        if not DBSCAN_AVAILABLE:
            messagebox.showerror("DBSCAN Error", f"DBSCAN dependencies are missing.\n{DBSCAN_IMPORT_ERROR}")
            return
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return

        if self.dbscan_window is not None and self.dbscan_window.winfo_exists():
            self.dbscan_window.lift()
            self._refresh_dbscan_window()
            return

        self.dbscan_window = tk.Toplevel(self.root)
        self.dbscan_window.title("DBSCAN Clustering")
        self.dbscan_window.geometry("1180x640+860+150")

        outer = ttk.Frame(self.dbscan_window, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(outer)
        header.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(header, text="DBSCAN Clustering", font=("Segoe UI", 12, "bold")).pack(anchor="w", side=tk.LEFT)
        help_icon = self._get_help_icon()
        self.dbscan_help_btn = tk.Button(
            header,
            image=help_icon,
            text="" if help_icon is not None else "?",
            relief="flat",
            bd=0,
            command=self._toggle_dbscan_help,
            cursor="hand2",
        )
        self.dbscan_help_btn.pack(side=tk.LEFT, padx=(6, 0))
        if help_icon is None:
            self.dbscan_help_btn.configure(font=("Segoe UI", 9, "bold"))

        controls = ttk.Frame(outer)
        controls.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(controls, text="Epsilon (nm):").pack(side=tk.LEFT)
        self.dbscan_eps_var = tk.StringVar(value="120.0")
        eps_entry = ttk.Entry(controls, textvariable=self.dbscan_eps_var, width=10)
        eps_entry.pack(side=tk.LEFT, padx=(8, 14))

        ttk.Label(controls, text="Min samples:").pack(side=tk.LEFT)
        self.dbscan_min_samples_var = tk.StringVar(value="3")
        min_spin = ttk.Spinbox(
            controls,
            from_=2,
            to=50,
            increment=1,
            textvariable=self.dbscan_min_samples_var,
            width=8,
            command=self._refresh_dbscan_window,
        )
        min_spin.pack(side=tk.LEFT, padx=(8, 14))

        ttk.Button(controls, text="Run DBSCAN", command=self._refresh_dbscan_window).pack(side=tk.LEFT)
        ttk.Button(controls, text="Save Data", command=self.on_save_dbscan_data).pack(side=tk.RIGHT)

        eps_entry.bind("<Return>", lambda _evt: self._refresh_dbscan_window())
        eps_entry.bind("<FocusOut>", lambda _evt: self._refresh_dbscan_window())
        min_spin.bind("<Return>", lambda _evt: self._refresh_dbscan_window())
        min_spin.bind("<FocusOut>", lambda _evt: self._refresh_dbscan_window())

        self.dbscan_basis_var = tk.StringVar(value="")
        ttk.Label(outer, textvariable=self.dbscan_basis_var, font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 8))

        summary_frame = ttk.LabelFrame(outer, text="Aggregate Statistics", padding=10)
        summary_frame.pack(fill=tk.X, pady=(0, 8))

        self.dbscan_summary_vars = {}
        summary_specs = [
            ("particles_analyzed", "Particles analyzed"),
            ("clusters_found", "Clusters found"),
            ("particles_in_clusters", "Particles in clusters"),
            ("noise_particles", "Noise particles"),
            ("aggregation_index", "Aggregation index (%)"),
            ("average_cluster_size", "Average cluster size (particles)"),
            ("median_cluster_size", "Median cluster size (particles)"),
            ("mean_cluster_density", "Mean cluster density"),
        ]
        for idx, (key, label) in enumerate(summary_specs):
            row = idx // 2
            col = (idx % 2) * 2
            self.dbscan_summary_vars[key] = tk.StringVar(value="N/A")
            ttk.Label(summary_frame, text=label, width=34).grid(row=row, column=col, sticky="w", pady=2, padx=(0, 6))
            ttk.Label(summary_frame, textvariable=self.dbscan_summary_vars[key], width=18).grid(
                row=row,
                column=col + 1,
                sticky="e",
                pady=2,
            )

        table_frame = ttk.Frame(outer)
        table_frame.pack(fill=tk.BOTH, expand=True)

        self.dbscan_tree = ttk.Treeview(
            table_frame,
            columns=(
                "cluster_id",
                "n_particles",
                "total_area",
                "hull_area",
                "density",
                "mean_length",
                "mean_width",
                "mean_area",
                "mean_nnd",
            ),
            show="headings",
        )
        self.dbscan_tree.heading("cluster_id", text="Cluster")
        self.dbscan_tree.heading("n_particles", text="Particles (N)")
        self.dbscan_tree.heading("total_area", text="Total area (nm²)")
        self.dbscan_tree.heading("hull_area", text="Convex hull area (nm²)")
        self.dbscan_tree.heading("density", text="Density (area/hull)")
        self.dbscan_tree.heading("mean_length", text="Mean length (nm)")
        self.dbscan_tree.heading("mean_width", text="Mean width (nm)")
        self.dbscan_tree.heading("mean_area", text="Mean area (nm²)")
        self.dbscan_tree.heading("mean_nnd", text="Mean NND (nm)")

        self.dbscan_tree.column("cluster_id", width=90, anchor="center")
        self.dbscan_tree.column("n_particles", width=110, anchor="e")
        self.dbscan_tree.column("total_area", width=130, anchor="e")
        self.dbscan_tree.column("hull_area", width=150, anchor="e")
        self.dbscan_tree.column("density", width=120, anchor="e")
        self.dbscan_tree.column("mean_length", width=120, anchor="e")
        self.dbscan_tree.column("mean_width", width=120, anchor="e")
        self.dbscan_tree.column("mean_area", width=120, anchor="e")
        self.dbscan_tree.column("mean_nnd", width=110, anchor="e")
        self.dbscan_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        yscroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.dbscan_tree.yview)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.dbscan_tree.configure(yscrollcommand=yscroll.set)

        self.dbscan_window.protocol("WM_DELETE_WINDOW", self._close_dbscan_window)
        self._refresh_dbscan_window()

    def _close_dbscan_window(self) -> None:
        if self.dbscan_window is not None and self.dbscan_window.winfo_exists():
            self.dbscan_window.destroy()
        self.dbscan_window = None
        self.dbscan_eps_var = None
        self.dbscan_min_samples_var = None
        self.dbscan_basis_var = None
        self.dbscan_summary_vars = {}
        self.dbscan_tree = None
        self.dbscan_export_headers = ()
        self.dbscan_export_rows = []
        self.dbscan_summary_snapshot = {}
        self.dbscan_help_btn = None
        if self.dbscan_help_window is not None and self.dbscan_help_window.winfo_exists():
            self.dbscan_help_window.destroy()
        self.dbscan_help_window = None
        self.dbscan_help_tracking = False

    def _toggle_dbscan_help(self) -> None:
        if self.dbscan_window is None or not self.dbscan_window.winfo_exists():
            return
        if self.dbscan_help_window is not None and self.dbscan_help_window.winfo_exists():
            self.dbscan_help_window.destroy()
            self.dbscan_help_window = None
            self.dbscan_help_tracking = False
            return

        self.dbscan_help_window = tk.Toplevel(self.dbscan_window)
        self.dbscan_help_window.overrideredirect(True)
        self.dbscan_help_window.attributes("-topmost", True)

        bubble = tk.Frame(self.dbscan_help_window, bg="#fff7e6", bd=1, relief="solid")
        bubble.pack(fill=tk.BOTH, expand=True)

        p1 = tk.Text(
            bubble,
            bg="#fff7e6",
            fg="#333333",
            wrap="word",
            width=50,
            height=5,
            relief="flat",
            highlightthickness=0,
        )
        p1.tag_configure("title", font=("Segoe UI", 9, "bold"))
        p1.tag_configure("body", font=("Segoe UI", 9))
        p1.insert("1.0", "Epsilon (nm) —", "title")
        p1.insert(
            "end",
            " the maximum distance between particles to be considered neighbors. Smaller values detect only very tight clusters, "
            "while larger values merge particles into broader clusters and may group nearby structures together.",
            "body",
        )
        p1.configure(state="disabled")
        p1.pack(anchor="w", padx=8, pady=(8, 4))

        p2 = tk.Text(
            bubble,
            bg="#fff7e6",
            fg="#333333",
            wrap="word",
            width=50,
            height=4,
            relief="flat",
            highlightthickness=0,
        )
        p2.tag_configure("title", font=("Segoe UI", 9, "bold"))
        p2.tag_configure("body", font=("Segoe UI", 9))
        p2.insert("1.0", "Min samples —", "title")
        p2.insert(
            "end",
            " the minimum number of neighboring particles required to form a cluster. Higher values make the clustering more strict "
            "(fewer, denser clusters), while lower values allow smaller or looser clusters to be detected.",
            "body",
        )
        p2.configure(state="disabled")
        p2.pack(anchor="w", padx=8, pady=(0, 4))

        p3 = tk.Text(
            bubble,
            bg="#fff7e6",
            fg="#333333",
            wrap="word",
            width=50,
            height=5,
            relief="flat",
            highlightthickness=0,
        )
        p3.tag_configure("title", font=("Segoe UI", 9, "bold"))
        p3.tag_configure("body", font=("Segoe UI", 9))
        p3.insert("1.0", "Convex hull area (nm²) —", "title")
        p3.insert(
            "end",
            " the area of the smallest convex polygon enclosing all particles in a cluster. It provides a measure of the "
            "cluster’s spatial extent and spread, useful for comparing cluster sizes and compactness.",
            "body",
        )
        p3.configure(state="disabled")
        p3.pack(anchor="w", padx=8, pady=(0, 8))

        self.dbscan_help_tracking = True
        if self.dbscan_window is not None:
            self.dbscan_window.bind("<Configure>", self._on_dbscan_window_configure)
        self._position_dbscan_help()

    def _on_dbscan_window_configure(self, _event=None) -> None:
        if self.dbscan_help_tracking:
            self._position_dbscan_help()

    def _position_dbscan_help(self) -> None:
        if self.dbscan_help_window is None or not self.dbscan_help_window.winfo_exists():
            return
        if self.dbscan_window is None or not self.dbscan_window.winfo_exists():
            return
        self.dbscan_help_window.update_idletasks()
        bubble_w = self.dbscan_help_window.winfo_width()
        bubble_h = self.dbscan_help_window.winfo_height()
        win_x = self.dbscan_window.winfo_rootx()
        win_y = self.dbscan_window.winfo_rooty()
        x = win_x + 20
        y = win_y - bubble_h - 8
        if y < 0:
            y = win_y + 30
        self.dbscan_help_window.geometry(f"{bubble_w}x{bubble_h}+{x}+{y}")

    def _parse_dbscan_eps_nm(self) -> float:
        if self.dbscan_eps_var is None:
            return 120.0
        try:
            eps_nm = float(self.dbscan_eps_var.get().strip())
        except Exception:
            eps_nm = 120.0
        eps_nm = max(1e-6, eps_nm)
        self.dbscan_eps_var.set(f"{eps_nm:.4f}")
        return eps_nm

    def _parse_dbscan_min_samples(self) -> int:
        if self.dbscan_min_samples_var is None:
            return 3
        try:
            min_samples = int(self.dbscan_min_samples_var.get().strip())
        except Exception:
            min_samples = 3
        min_samples = max(2, min_samples)
        self.dbscan_min_samples_var.set(str(min_samples))
        return min_samples

    def _convex_hull_area_nm2(self, points_nm: np.ndarray) -> float:
        if points_nm.size == 0:
            return 0.0
        n = int(points_nm.shape[0])
        if n < 3:
            if n < 2:
                return 0.0
            span = np.ptp(points_nm, axis=0)
            return float(max(0.0, span[0] * span[1]))
        try:
            hull = ConvexHull(points_nm)
            return float(max(0.0, hull.volume))
        except (QhullError, ValueError):
            span = np.ptp(points_nm, axis=0)
            return float(max(0.0, span[0] * span[1]))

    def _mean_nearest_neighbor_nm(self, points_nm: np.ndarray) -> float:
        if points_nm.shape[0] < 2:
            return float("nan")
        diff = points_nm[:, np.newaxis, :] - points_nm[np.newaxis, :, :]
        dist = np.sqrt(np.sum(diff * diff, axis=2))
        np.fill_diagonal(dist, np.inf)
        nearest = np.min(dist, axis=1)
        return float(np.mean(nearest))

    def _refresh_dbscan_window(self) -> None:
        if self.dbscan_window is None or not self.dbscan_window.winfo_exists():
            return
        if self.dbscan_tree is None or self.dbscan_basis_var is None:
            return
        if self.dbscan_eps_var is None or self.dbscan_min_samples_var is None:
            return

        self.dbscan_basis_var.set("Based on selected particle centroids (nm)")
        for row in self.dbscan_tree.get_children():
            self.dbscan_tree.delete(row)

        self.dbscan_export_headers = (
            "Cluster",
            "Particles (N)",
            "Total area (nm²)",
            "Convex hull area (nm²)",
            "Density (area/hull)",
            "Mean length (nm)",
            "Mean width (nm)",
            "Mean area (nm²)",
            "Mean NND (nm)",
        )
        self.dbscan_export_rows = []

        selected = self._selected_particles()
        if not selected:
            for key in self.dbscan_summary_vars:
                self.dbscan_summary_vars[key].set("N/A")
            self.dbscan_summary_snapshot = {}
            return

        nm_per_px = self._safe_um_per_px_for_refresh() * 1000.0
        eps_nm = self._parse_dbscan_eps_nm()
        min_samples = self._parse_dbscan_min_samples()

        coords_nm = np.asarray(
            [
                [
                    float(self._particle_centroid_global(p)[0]) * nm_per_px,
                    float(self._particle_centroid_global(p)[1]) * nm_per_px,
                ]
                for p in selected
            ],
            dtype=float,
        )
        area_nm2 = np.asarray([float(p.area_px) * (nm_per_px ** 2) for p in selected], dtype=float)

        lengths_nm_list: List[float] = []
        widths_nm_list: List[float] = []
        for p in selected:
            length_px, width_px = self._center_length_width_px(p)
            lengths_nm_list.append(float(length_px) * nm_per_px)
            widths_nm_list.append(float(width_px) * nm_per_px)
        length_nm = np.asarray(lengths_nm_list, dtype=float)
        width_nm = np.asarray(widths_nm_list, dtype=float)

        labels = DBSCAN(eps=eps_nm, min_samples=min_samples).fit_predict(coords_nm)
        cluster_ids = sorted(int(c) for c in set(labels.tolist()) if int(c) >= 0)

        n_total = int(len(selected))
        n_in_clusters = int(np.sum(labels >= 0))
        n_noise = int(np.sum(labels < 0))
        agg_index = (100.0 * n_in_clusters / n_total) if n_total > 0 else float("nan")

        cluster_sizes: List[int] = []
        cluster_densities: List[float] = []

        def fmt(val: float, decimals: int = 4) -> str:
            return f"{val:.{decimals}f}" if np.isfinite(val) else "N/A"

        for rank, cid in enumerate(cluster_ids, start=1):
            idx = np.flatnonzero(labels == cid)
            if idx.size == 0:
                continue

            n_particles = int(idx.size)
            cluster_sizes.append(n_particles)
            c_coords = coords_nm[idx]
            c_area_total = float(np.sum(area_nm2[idx]))
            c_hull_area = self._convex_hull_area_nm2(c_coords)
            c_density = float(c_area_total / c_hull_area) if c_hull_area > 0 else float("nan")
            if np.isfinite(c_density):
                cluster_densities.append(c_density)

            c_mean_length = float(np.mean(length_nm[idx]))
            c_mean_width = float(np.mean(width_nm[idx]))
            c_mean_area = float(np.mean(area_nm2[idx]))
            c_mean_nnd = self._mean_nearest_neighbor_nm(c_coords)

            row = (
                f"C{rank}",
                str(n_particles),
                fmt(c_area_total, 4),
                fmt(c_hull_area, 4),
                fmt(c_density, 4),
                fmt(c_mean_length, 4),
                fmt(c_mean_width, 4),
                fmt(c_mean_area, 4),
                fmt(c_mean_nnd, 4),
            )
            self.dbscan_tree.insert("", tk.END, values=row)
            self.dbscan_export_rows.append(row)

        avg_cluster_size = float(np.mean(cluster_sizes)) if cluster_sizes else float("nan")
        med_cluster_size = float(np.median(cluster_sizes)) if cluster_sizes else float("nan")
        mean_density = float(np.mean(cluster_densities)) if cluster_densities else float("nan")

        summary_values = {
            "particles_analyzed": f"{n_total}",
            "clusters_found": f"{len(cluster_ids)}",
            "particles_in_clusters": f"{n_in_clusters}",
            "noise_particles": f"{n_noise}",
            "aggregation_index": fmt(agg_index, 2),
            "average_cluster_size": fmt(avg_cluster_size, 2),
            "median_cluster_size": fmt(med_cluster_size, 2),
            "mean_cluster_density": fmt(mean_density, 4),
        }
        for key, var in self.dbscan_summary_vars.items():
            var.set(summary_values.get(key, "N/A"))

        self.dbscan_summary_snapshot = {
            "Particles analyzed": summary_values["particles_analyzed"],
            "Clusters found": summary_values["clusters_found"],
            "Particles in clusters": summary_values["particles_in_clusters"],
            "Noise particles": summary_values["noise_particles"],
            "Aggregation index (%)": summary_values["aggregation_index"],
            "Average cluster size (particles)": summary_values["average_cluster_size"],
            "Median cluster size (particles)": summary_values["median_cluster_size"],
            "Mean cluster density": summary_values["mean_cluster_density"],
            "Epsilon (nm)": f"{eps_nm:.4f}",
            "Min samples": f"{min_samples}",
        }

    def on_save_dbscan_data(self) -> None:
        if not self.dbscan_summary_snapshot:
            messagebox.showerror("Error", "Run DBSCAN first.")
            return

        save_path = filedialog.asksaveasfilename(
            title="Save DBSCAN Data",
            defaultextension=".csv",
            initialfile="dbscan_clustering.csv",
            filetypes=[
                ("CSV", "*.csv"),
                ("Excel workbook", "*.xlsx"),
            ],
        )
        if not save_path:
            return

        ext = os.path.splitext(save_path)[1].lower()
        if ext not in (".csv", ".xlsx"):
            save_path = f"{save_path}.csv"
            ext = ".csv"

        summary_headers = tuple(self.dbscan_summary_snapshot.keys())
        summary_values = tuple(self.dbscan_summary_snapshot.values())
        cluster_headers = self.dbscan_export_headers
        cluster_rows = self.dbscan_export_rows

        try:
            if ext == ".csv":
                with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(("DBSCAN Clustering Summary",))
                    writer.writerow(summary_headers)
                    writer.writerow(summary_values)
                    writer.writerow(())
                    writer.writerow(("Cluster Statistics",))
                    writer.writerow(cluster_headers)
                    writer.writerows(cluster_rows)
            else:
                if not XLSX_AVAILABLE:
                    messagebox.showerror(
                        "XLSX unavailable",
                        "Saving as .xlsx requires openpyxl.\n\nInstall with:\npip install openpyxl",
                    )
                    return
                wb = Workbook()
                ws = wb.active
                ws.title = "DBSCAN"
                ws.append(["DBSCAN Clustering Summary"])
                ws.append(list(summary_headers))
                ws.append(list(summary_values))
                ws.append([])
                ws.append(["Cluster Statistics"])
                ws.append(list(cluster_headers))
                for row in cluster_rows:
                    ws.append(list(row))
                wb.save(save_path)
        except Exception as exc:
            messagebox.showerror("Save error", f"Could not save DBSCAN file:\n{exc}")
            return

        self.status_var.set(f"DBSCAN data saved: {os.path.basename(save_path)}")

    def open_bootstrapping(self) -> None:
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return

        if self.bootstrap_window is not None and self.bootstrap_window.winfo_exists():
            self.bootstrap_window.lift()
            self._refresh_bootstrap_window()
            return

        self.bootstrap_window = tk.Toplevel(self.root)
        self.bootstrap_window.title("Bootstrapping")
        self.bootstrap_window.geometry("980x650+860+160")

        outer = ttk.Frame(self.bootstrap_window, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(outer)
        header.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(header, text="Bootstrapping", font=("Segoe UI", 12, "bold")).pack(side=tk.LEFT, anchor="w")
        help_icon = self._get_help_icon()
        self.bootstrap_help_btn = tk.Button(
            header,
            image=help_icon,
            text="" if help_icon is not None else "?",
            relief="flat",
            bd=0,
            command=self._toggle_bootstrap_help,
            cursor="hand2",
        )
        self.bootstrap_help_btn.pack(side=tk.LEFT, padx=(6, 0))
        if help_icon is None:
            self.bootstrap_help_btn.configure(font=("Segoe UI", 9, "bold"))

        controls = ttk.Frame(outer)
        controls.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(controls, text="Property:").pack(side=tk.LEFT)
        self.bootstrap_property_var = tk.StringVar(value="length")
        prop_combo = ttk.Combobox(
            controls,
            textvariable=self.bootstrap_property_var,
            state="readonly",
            values=("length", "width", "area", "feret diameter", "circularity", "eccentricity"),
            width=12,
        )
        prop_combo.pack(side=tk.LEFT, padx=(8, 14))
        prop_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_bootstrap_window())

        ttk.Label(controls, text="Bootstrap samples:").pack(side=tk.LEFT)
        self.bootstrap_n_var = tk.StringVar(value="500")
        n_spin = ttk.Spinbox(
            controls,
            from_=100,
            to=20000,
            increment=100,
            textvariable=self.bootstrap_n_var,
            width=10,
            command=self._refresh_bootstrap_window,
        )
        n_spin.pack(side=tk.LEFT, padx=(8, 0))
        n_spin.bind("<Return>", lambda _evt: self._refresh_bootstrap_window())
        n_spin.bind("<FocusOut>", lambda _evt: self._refresh_bootstrap_window())

        self.bootstrap_basis_var = tk.StringVar(value="")
        ttk.Label(outer, textvariable=self.bootstrap_basis_var, font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 10))

        stats_frame = ttk.Frame(outer)
        stats_frame.pack(fill=tk.X, pady=(0, 10))

        self.bootstrap_stat_vars = {}
        stat_names = (
            "Bootstrap N",
            "Mean",
            "Median",
            "Percentile 2.5",
            "Percentile 97.5",
        )
        for i, name in enumerate(stat_names):
            key = name.lower().replace(" ", "_").replace(".", "_")
            self.bootstrap_stat_vars[key] = tk.StringVar(value="N/A")
            ttk.Label(stats_frame, text=name, width=16).grid(row=i, column=0, sticky="w", pady=2)
            ttk.Label(stats_frame, textvariable=self.bootstrap_stat_vars[key], width=28).grid(
                row=i,
                column=1,
                sticky="e",
                pady=2,
            )

        self.bootstrap_fig = Figure(figsize=(8.8, 4.2), dpi=100)
        self.bootstrap_ax = self.bootstrap_fig.add_subplot(111)
        self.bootstrap_canvas = FigureCanvasTkAgg(self.bootstrap_fig, master=outer)
        self.bootstrap_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self.bootstrap_window.protocol("WM_DELETE_WINDOW", self._close_bootstrap_window)
        self._refresh_bootstrap_window()

    def _close_bootstrap_window(self) -> None:
        if self.bootstrap_window is not None and self.bootstrap_window.winfo_exists():
            self.bootstrap_window.destroy()
        self.bootstrap_window = None
        self.bootstrap_property_var = None
        self.bootstrap_n_var = None
        self.bootstrap_basis_var = None
        self.bootstrap_stat_vars = {}
        self.bootstrap_fig = None
        self.bootstrap_ax = None
        self.bootstrap_canvas = None
        self.bootstrap_help_btn = None
        if self.bootstrap_help_window is not None and self.bootstrap_help_window.winfo_exists():
            self.bootstrap_help_window.destroy()
        self.bootstrap_help_window = None
        self.bootstrap_help_tracking = False

    def _toggle_bootstrap_help(self) -> None:
        if self.bootstrap_window is None or not self.bootstrap_window.winfo_exists():
            return
        if self.bootstrap_help_window is not None and self.bootstrap_help_window.winfo_exists():
            self.bootstrap_help_window.destroy()
            self.bootstrap_help_window = None
            self.bootstrap_help_tracking = False
            return

        self.bootstrap_help_window = tk.Toplevel(self.bootstrap_window)
        self.bootstrap_help_window.overrideredirect(True)
        self.bootstrap_help_window.attributes("-topmost", True)

        bubble = tk.Frame(self.bootstrap_help_window, bg="#fff7e6", bd=1, relief="solid")
        bubble.pack(fill=tk.BOTH, expand=True)

        p1 = tk.Text(
            bubble,
            bg="#fff7e6",
            fg="#333333",
            wrap="word",
            width=46,
            height=5,
            relief="flat",
            highlightthickness=0,
        )
        p1.tag_configure("title", font=("Segoe UI", 9, "bold"))
        p1.tag_configure("body", font=("Segoe UI", 9))
        p1.insert("1.0", "Bootstrap samples —", "title")
        p1.insert(
            "end",
            " the number of times the dataset is resampled (with replacement) to estimate statistics. "
            "Increasing this value improves the stability and precision of the results (smoother confidence intervals), "
            "but also increases computation time.",
            "body",
        )
        p1.configure(state="disabled")
        p1.pack(anchor="w", padx=8, pady=(8, 4))

        p2 = tk.Text(
            bubble,
            bg="#fff7e6",
            fg="#333333",
            wrap="word",
            width=46,
            height=5,
            relief="flat",
            highlightthickness=0,
        )
        p2.tag_configure("title", font=("Segoe UI", 9, "bold"))
        p2.tag_configure("body", font=("Segoe UI", 9))
        p2.insert("1.0", "2.5 and 97.5 percentiles —", "title")
        p2.insert(
            "end",
            " these define the bounds of a 95% confidence interval for the estimated parameter. "
            "They indicate the range within which the true value is likely to lie with high confidence, "
            "based on the variability of the resampled data.",
            "body",
        )
        p2.configure(state="disabled")
        p2.pack(anchor="w", padx=8, pady=(0, 8))

        self.bootstrap_help_tracking = True
        if self.bootstrap_window is not None:
            self.bootstrap_window.bind("<Configure>", self._on_bootstrap_window_configure)
        self._position_bootstrap_help()

    def _on_bootstrap_window_configure(self, _event=None) -> None:
        if self.bootstrap_help_tracking:
            self._position_bootstrap_help()

    def _position_bootstrap_help(self) -> None:
        if self.bootstrap_help_window is None or not self.bootstrap_help_window.winfo_exists():
            return
        if self.bootstrap_window is None or not self.bootstrap_window.winfo_exists():
            return
        self.bootstrap_help_window.update_idletasks()
        bubble_w = self.bootstrap_help_window.winfo_width()
        bubble_h = self.bootstrap_help_window.winfo_height()
        win_x = self.bootstrap_window.winfo_rootx()
        win_y = self.bootstrap_window.winfo_rooty()
        x = win_x + 20
        y = win_y - bubble_h - 8
        if y < 0:
            y = win_y + 30
        self.bootstrap_help_window.geometry(f"{bubble_w}x{bubble_h}+{x}+{y}")

    def _parse_bootstrap_n(self) -> int:
        if self.bootstrap_n_var is None:
            return 500
        try:
            n_boot = int(self.bootstrap_n_var.get().strip())
        except Exception:
            n_boot = 500
        return max(50, n_boot)

    def _bootstrap_mean_distribution(self, values: np.ndarray, n_boot: int) -> np.ndarray:
        n = values.size
        if n == 0 or n_boot <= 0:
            return np.asarray([], dtype=float)
        rng = np.random.default_rng()
        idx = rng.integers(0, n, size=(n_boot, n))
        sample_means = values[idx].mean(axis=1)
        return np.asarray(sample_means, dtype=float)

    def _refresh_bootstrap_window(self) -> None:
        if self.bootstrap_window is None or not self.bootstrap_window.winfo_exists():
            return
        if self.bootstrap_property_var is None or self.bootstrap_basis_var is None:
            return
        if self.bootstrap_ax is None or self.bootstrap_canvas is None:
            return
        if not self.bootstrap_stat_vars:
            return

        prop = self.bootstrap_property_var.get().strip().lower()
        values, unit, basis = self._selected_metric_values(prop)
        self.bootstrap_basis_var.set(f"Based on selected particle {basis}")
        n_boot = self._parse_bootstrap_n()
        if self.bootstrap_n_var is not None:
            self.bootstrap_n_var.set(str(n_boot))

        means_dist = self._bootstrap_mean_distribution(values, n_boot)

        def set_stat(key: str, value: str) -> None:
            if key in self.bootstrap_stat_vars:
                self.bootstrap_stat_vars[key].set(value)

        if means_dist.size == 0:
            set_stat("bootstrap_n", str(n_boot))
            set_stat("mean", "N/A")
            set_stat("median", "N/A")
            set_stat("percentile_2_5", "N/A")
            set_stat("percentile_97_5", "N/A")
            self.bootstrap_ax.clear()
            self.bootstrap_ax.text(0.5, 0.5, "Select particles", ha="center", va="center")
            self.bootstrap_canvas.draw_idle()
            return

        mean_val = float(np.mean(means_dist))
        median_val = float(np.median(means_dist))
        p2_5 = float(np.percentile(means_dist, 2.5))
        p97_5 = float(np.percentile(means_dist, 97.5))

        def fmt(val: float) -> str:
            return f"{val:.4f} {unit}".strip() if unit else f"{val:.4f}"

        set_stat("bootstrap_n", str(n_boot))
        set_stat("mean", fmt(mean_val))
        set_stat("median", fmt(median_val))
        set_stat("percentile_2_5", fmt(p2_5))
        set_stat("percentile_97_5", fmt(p97_5))

        self.bootstrap_ax.clear()
        bins = max(10, min(50, int(np.sqrt(means_dist.size) * 2)))
        self.bootstrap_ax.hist(means_dist, bins=bins, color="#2E86C1", edgecolor="white", linewidth=0.6, alpha=0.9)
        self.bootstrap_ax.set_title(f"Bootstrap Mean Distribution ({basis})")
        self.bootstrap_ax.set_xlabel(f"Bootstrap mean ({unit})" if unit else "Bootstrap mean")
        self.bootstrap_ax.set_ylabel("Count")
        self.bootstrap_ax.grid(alpha=0.25, linestyle="--")
        self.bootstrap_canvas.draw_idle()

    def open_threshold_tool(self) -> None:
        if not self.particles:
            messagebox.showerror("Error", "Run AI segmentation first.")
            return

        if self.threshold_window is not None and self.threshold_window.winfo_exists():
            self.threshold_window.lift()
            return

        self.threshold_window = tk.Toplevel(self.root)
        self.threshold_window.title("Threshold")
        self.threshold_window.geometry("760x520+980+190")

        outer = ttk.Frame(self.threshold_window, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        ttk.Label(outer, text="Threshold", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))

        controls = ttk.Frame(outer)
        controls.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(controls, text="Property:").pack(side=tk.LEFT)
        self.threshold_property_var = tk.StringVar(value="length")
        prop_combo = ttk.Combobox(
            controls,
            textvariable=self.threshold_property_var,
            state="readonly",
            values=("length", "width", "area", "feret diameter", "circularity", "eccentricity"),
            width=12,
        )
        prop_combo.pack(side=tk.LEFT, padx=(8, 0))
        prop_combo.bind(
            "<<ComboboxSelected>>",
            lambda _evt: self._refresh_threshold_window(reset_range=True, apply_after=True),
        )

        self.threshold_basis_var = tk.StringVar(value="")
        ttk.Label(outer, textvariable=self.threshold_basis_var, font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 8))

        self.threshold_fig = Figure(figsize=(7.2, 2.5), dpi=100)
        self.threshold_ax = self.threshold_fig.add_subplot(111)
        self.threshold_canvas = FigureCanvasTkAgg(self.threshold_fig, master=outer)
        self.threshold_canvas.get_tk_widget().pack(fill=tk.X, pady=(0, 10))

        mapping = ttk.LabelFrame(outer, text="Color Mapping", padding=10)
        mapping.pack(fill=tk.X, pady=(0, 4))

        row1 = ttk.Frame(mapping)
        row1.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(row1, text="Start:").pack(side=tk.LEFT)
        self.threshold_min_var = tk.StringVar(value="")
        min_entry = ttk.Entry(row1, textvariable=self.threshold_min_var, width=14)
        min_entry.pack(side=tk.LEFT, padx=(8, 6))
        self.threshold_unit_var = tk.StringVar(value="nm")
        ttk.Label(row1, textvariable=self.threshold_unit_var).pack(side=tk.LEFT)

        row2 = ttk.Frame(mapping)
        row2.pack(fill=tk.X)
        ttk.Label(row2, text="End:").pack(side=tk.LEFT)
        self.threshold_max_var = tk.StringVar(value="")
        max_entry = ttk.Entry(row2, textvariable=self.threshold_max_var, width=14)
        max_entry.pack(side=tk.LEFT, padx=(12, 6))
        ttk.Label(row2, textvariable=self.threshold_unit_var).pack(side=tk.LEFT)

        for entry in (min_entry, max_entry):
            entry.bind("<Return>", lambda _evt: self._on_threshold_entries_changed())
            entry.bind("<FocusOut>", lambda _evt: self._on_threshold_entries_changed())

        self.threshold_window.protocol("WM_DELETE_WINDOW", self._close_threshold_window)
        self._refresh_threshold_window(reset_range=True, apply_after=True)

    def _close_threshold_window(self) -> None:
        if self.threshold_window is not None and self.threshold_window.winfo_exists():
            self.threshold_window.destroy()
        self.threshold_window = None
        self.threshold_property_var = None
        self.threshold_min_var = None
        self.threshold_max_var = None
        self.threshold_unit_var = None
        self.threshold_basis_var = None
        self.threshold_fig = None
        self.threshold_ax = None
        self.threshold_canvas = None
        self.threshold_span = None
        self.threshold_values = None
        self.threshold_ids = None

    def _all_metric_values(self, prop: str) -> Tuple[np.ndarray, np.ndarray, str, str]:
        nm_per_px = self._safe_um_per_px_for_refresh() * 1000.0
        values: List[float] = []
        ids: List[int] = []

        if prop == "width":
            unit = "nm"
            basis = "width (nm)"
        elif prop == "area":
            unit = "nm^2"
            basis = "area (nm^2)"
        elif prop == "feret diameter":
            unit = "nm"
            basis = "feret diameter (nm)"
        elif prop == "circularity":
            unit = ""
            basis = "circularity"
        elif prop == "eccentricity":
            unit = ""
            basis = "eccentricity"
        else:
            unit = "nm"
            basis = "length (nm)"

        for p in self.particles:
            length_px, width_px = self._center_length_width_px(p)
            if prop == "width":
                values.append(width_px * nm_per_px)
            elif prop == "area":
                values.append(float(p.area_px) * (nm_per_px ** 2))
            elif prop == "feret diameter":
                values.append(self._feret_diameter_px(p) * nm_per_px)
            elif prop == "circularity":
                values.append(self._circularity_eccentricity(p)[0])
            elif prop == "eccentricity":
                values.append(self._circularity_eccentricity(p)[1])
            else:
                values.append(length_px * nm_per_px)
            ids.append(int(p.mask_id))
        return np.asarray(values, dtype=float), np.asarray(ids, dtype=int), unit, basis

    def _refresh_threshold_window(self, reset_range: bool = True, apply_after: bool = True) -> None:
        if self.threshold_window is None or not self.threshold_window.winfo_exists():
            return
        if self.threshold_property_var is None:
            return
        if self.threshold_min_var is None or self.threshold_max_var is None:
            return
        if self.threshold_ax is None or self.threshold_canvas is None:
            return
        if self.threshold_unit_var is None or self.threshold_basis_var is None:
            return

        prop = self.threshold_property_var.get().strip().lower()
        values, ids, unit, basis = self._all_metric_values(prop)
        self.threshold_values = values
        self.threshold_ids = ids
        self.threshold_unit_var.set(unit)
        self.threshold_basis_var.set(f"Based on all particle {basis}")

        self.threshold_ax.clear()
        self.threshold_span = None

        if values.size == 0:
            self.threshold_min_var.set("")
            self.threshold_max_var.set("")
            self.threshold_ax.text(0.5, 0.5, "Run AI segmentation first", ha="center", va="center")
            self.threshold_ax.set_axis_off()
            self.threshold_canvas.draw_idle()
            return

        bins = max(10, min(60, int(np.sqrt(values.size) * 2)))
        self.threshold_ax.hist(
            values,
            bins=bins,
            color="#d5d5d5",
            edgecolor="white",
            linewidth=0.6,
        )
        self.threshold_ax.grid(alpha=0.25, linestyle="--")
        self.threshold_ax.set_ylabel("Count")
        self.threshold_ax.set_xlabel(f"{prop.title()} ({unit})" if unit else prop.title())

        full_min = float(np.min(values))
        full_max = float(np.max(values))
        if full_max <= full_min:
            full_min -= 0.5
            full_max += 0.5

        if reset_range:
            lo = full_min
            hi = full_max
        else:
            try:
                lo = float(self.threshold_min_var.get().strip())
            except Exception:
                lo = full_min
            try:
                hi = float(self.threshold_max_var.get().strip())
            except Exception:
                hi = full_max
            lo = max(full_min, min(full_max, lo))
            hi = max(full_min, min(full_max, hi))
            if lo > hi:
                lo, hi = hi, lo

        self.threshold_min_var.set(f"{lo:.4f}")
        self.threshold_max_var.set(f"{hi:.4f}")

        try:
            self.threshold_span = SpanSelector(
                self.threshold_ax,
                self._on_threshold_span_selected,
                "horizontal",
                useblit=True,
                interactive=True,
                drag_from_anywhere=True,
                props={"facecolor": "#c97bb0", "alpha": 0.45},
                handle_props={"color": "#7a2a6a", "alpha": 0.9},
                minspan=0.0,
            )
        except TypeError:
            try:
                self.threshold_span = SpanSelector(
                    self.threshold_ax,
                    self._on_threshold_span_selected,
                    "horizontal",
                    useblit=True,
                    interactive=True,
                    rectprops={"facecolor": "#c97bb0", "alpha": 0.45},
                    minspan=0.0,
                )
            except TypeError:
                self.threshold_span = SpanSelector(
                    self.threshold_ax,
                    self._on_threshold_span_selected,
                    "horizontal",
                    useblit=True,
                    minspan=0.0,
                )
        try:
            self.threshold_span.extents = (lo, hi)
        except Exception:
            pass

        self.threshold_canvas.draw_idle()

        if apply_after:
            self._apply_threshold_selection(lo, hi)

    def _on_threshold_entries_changed(self) -> None:
        if self.threshold_values is None or self.threshold_values.size == 0:
            return
        if self.threshold_min_var is None or self.threshold_max_var is None:
            return

        full_min = float(np.min(self.threshold_values))
        full_max = float(np.max(self.threshold_values))
        if full_max <= full_min:
            full_min -= 0.5
            full_max += 0.5

        try:
            lo = float(self.threshold_min_var.get().strip())
        except Exception:
            lo = full_min
        try:
            hi = float(self.threshold_max_var.get().strip())
        except Exception:
            hi = full_max

        lo = max(full_min, min(full_max, lo))
        hi = max(full_min, min(full_max, hi))
        if lo > hi:
            lo, hi = hi, lo

        self.threshold_min_var.set(f"{lo:.4f}")
        self.threshold_max_var.set(f"{hi:.4f}")
        if self.threshold_span is not None:
            try:
                self.threshold_span.extents = (lo, hi)
            except Exception:
                pass
        if self.threshold_canvas is not None:
            self.threshold_canvas.draw_idle()
        self._apply_threshold_selection(lo, hi)

    def _on_threshold_span_selected(self, vmin: float, vmax: float) -> None:
        if self.threshold_values is None or self.threshold_values.size == 0:
            return
        if self.threshold_min_var is None or self.threshold_max_var is None:
            return

        full_min = float(np.min(self.threshold_values))
        full_max = float(np.max(self.threshold_values))
        lo = max(full_min, min(full_max, float(min(vmin, vmax))))
        hi = max(full_min, min(full_max, float(max(vmin, vmax))))
        if lo > hi:
            lo, hi = hi, lo

        self.threshold_min_var.set(f"{lo:.4f}")
        self.threshold_max_var.set(f"{hi:.4f}")
        self._apply_threshold_selection(lo, hi)

    def _apply_threshold_selection(self, lo: float, hi: float) -> None:
        if self.threshold_values is None or self.threshold_ids is None:
            return

        selected_mask = (self.threshold_values >= float(lo)) & (self.threshold_values <= float(hi))
        new_selected_ids = {int(pid) for pid, keep in zip(self.threshold_ids, selected_mask) if keep}

        self.selected_ids = new_selected_ids
        self.highlighted_ids.clear()
        self.render_image(keep_view=True)
        self.refresh_table()

        prop = self.threshold_property_var.get().strip().lower() if self.threshold_property_var is not None else "length"
        unit = self.threshold_unit_var.get() if self.threshold_unit_var is not None else ""
        self.status_var.set(
            f"Threshold {prop}: {len(new_selected_ids)} particles in [{lo:.4f}, {hi:.4f}] {unit}"
        )

    def open_nearest_neighbor_analysis(self) -> None:
        if not NND_AVAILABLE:
            messagebox.showerror("NND Error", f"NND dependencies are missing.\n{NND_IMPORT_ERROR}")
            return
        if len(self.selected_ids) < 2:
            messagebox.showerror("Error", "Select at least 2 particles")
            return

        if self.nnd_summary_window is None or not self.nnd_summary_window.winfo_exists():
            self.nnd_summary_window = tk.Toplevel(self.root)
            self.nnd_summary_window.title("Nearest-Neighbor Analysis")
            self.nnd_summary_window.geometry("700x520+860+160")

            outer = ttk.Frame(self.nnd_summary_window, padding=12)
            outer.pack(fill=tk.BOTH, expand=True)

            header = ttk.Frame(outer)
            header.pack(fill=tk.X, pady=(0, 8))
            ttk.Label(header, text="Nearest-Neighbor Analysis", font=("Segoe UI", 12, "bold")).pack(anchor="w", side=tk.LEFT)
            help_icon = self._get_help_icon()
            self.nnd_help_btn = tk.Button(
                header,
                image=help_icon,
                text="" if help_icon is not None else "?",
                relief="flat",
                bd=0,
                command=self._toggle_nnd_help,
                cursor="hand2",
            )
            self.nnd_help_btn.pack(side=tk.LEFT, padx=(6, 0))
            if help_icon is None:
                self.nnd_help_btn.configure(font=("Segoe UI", 9, "bold"))

            controls = ttk.Frame(outer)
            controls.pack(fill=tk.X, pady=(0, 8))

            ttk.Label(controls, text="Monte-Carlo B:").pack(side=tk.LEFT)
            self.nnd_mc_var = tk.StringVar(value="300")
            b_spin = ttk.Spinbox(
                controls,
                from_=50,
                to=5000,
                increment=50,
                textvariable=self.nnd_mc_var,
                width=9,
                command=self._refresh_nnd_outputs,
            )
            b_spin.pack(side=tk.LEFT, padx=(8, 12))
            b_spin.bind("<Return>", lambda _evt: self._refresh_nnd_outputs())
            b_spin.bind("<FocusOut>", lambda _evt: self._refresh_nnd_outputs())

            ttk.Button(controls, text="Run Analysis", command=self._refresh_nnd_outputs).pack(side=tk.LEFT)
            ttk.Button(controls, text="Save Data", command=self.on_save_nnd_data).pack(side=tk.RIGHT)

            self.nnd_basis_var = tk.StringVar(value="")
            ttk.Label(outer, textvariable=self.nnd_basis_var, font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 8))

            summary_frame = ttk.LabelFrame(outer, text="Summary Table", padding=10)
            summary_frame.pack(fill=tk.BOTH, expand=True)

            self.nnd_summary_vars = {}
            summary_specs = [
                ("n", "N"),
                ("a_nm2", "A (nm²)"),
                ("lambda", "lambda (points/nm²)"),
                ("mean", "Mean NND (nm)"),
                ("median", "Median NND (nm)"),
                ("std", "Std NND (nm)"),
                ("d10", "D10 (nm)"),
                ("d50", "D50 (nm)"),
                ("d90", "D90 (nm)"),
                ("mean_csr", "mean_CSR (nm)"),
                ("var_csr", "var_CSR (nm²)"),
                ("clark_evans_r", "Clark-Evans R"),
                ("z", "z-statistic"),
                ("p_theoretical", "p (theoretical)"),
                ("p_simulated", "p (simulated)"),
            ]
            for idx, (key, label) in enumerate(summary_specs):
                row = idx // 2
                col = (idx % 2) * 2
                self.nnd_summary_vars[key] = tk.StringVar(value="N/A")
                ttk.Label(summary_frame, text=label, width=24).grid(row=row, column=col, sticky="w", pady=2, padx=(0, 6))
                ttk.Label(summary_frame, textvariable=self.nnd_summary_vars[key], width=20).grid(
                    row=row,
                    column=col + 1,
                    sticky="e",
                    pady=2,
                )

            self.nnd_summary_window.protocol("WM_DELETE_WINDOW", self._close_nnd_summary_window)

        if self.nnd_plot_window is None or not self.nnd_plot_window.winfo_exists():
            self.nnd_plot_window = tk.Toplevel(self.root)
            self.nnd_plot_window.title("Nearest-Neighbor Analysis - Plots")
            self.nnd_plot_window.geometry("1220x760+780+120")

            outer_plot = ttk.Frame(self.nnd_plot_window, padding=10)
            outer_plot.pack(fill=tk.BOTH, expand=True)

            map_frame = ttk.LabelFrame(outer_plot, text="Centroid NND Map", padding=8)
            map_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
            header = ttk.Frame(map_frame)
            header.pack(fill=tk.X, pady=(0, 4))
            self.nnd_map_kde_btn = ttk.Button(header, text="KDE fitting", command=self._toggle_nnd_map_kde)
            self.nnd_map_kde_btn.pack(side=tk.LEFT)
            self._attach_tooltip(self.nnd_map_kde_btn, "probability density function")
            ttk.Button(header, text="Save Histogram", command=self.on_save_nnd_map_hist_image).pack(side=tk.RIGHT)
            ttk.Button(header, text="Save Map", command=self.on_save_nnd_map_image).pack(side=tk.RIGHT, padx=(0, 6))
            self.nnd_map_fig = Figure(figsize=(10.2, 3.2), dpi=100)
            self.nnd_map_ax = self.nnd_map_fig.add_subplot(111)
            self.nnd_map_canvas = FigureCanvasTkAgg(self.nnd_map_fig, master=map_frame)
            self.nnd_map_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            if self.nnd_map_hover_cid is None:
                self.nnd_map_hover_cid = self.nnd_map_canvas.mpl_connect(
                    "motion_notify_event",
                    self._on_nnd_map_kde_hover,
                )

            hist_frame = ttk.LabelFrame(outer_plot, text="NND Histogram / KDE / CSR", padding=8)
            hist_frame.pack(fill=tk.BOTH, expand=True)
            ttk.Button(hist_frame, text="Save Image", command=self.on_save_nnd_hist_image).pack(anchor="e", pady=(0, 4))
            self.nnd_hist_fig = Figure(figsize=(8.6, 3.2), dpi=100)
            self.nnd_hist_ax = self.nnd_hist_fig.add_subplot(111)
            self.nnd_hist_canvas = FigureCanvasTkAgg(self.nnd_hist_fig, master=hist_frame)
            self.nnd_hist_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            self.nnd_plot_window.protocol("WM_DELETE_WINDOW", self._close_nnd_plot_window)

        self._refresh_nnd_outputs()

    def _close_nnd_summary_window(self) -> None:
        if self.nnd_summary_window is not None and self.nnd_summary_window.winfo_exists():
            self.nnd_summary_window.destroy()
        self.nnd_summary_window = None
        self.nnd_basis_var = None
        self.nnd_mc_var = None
        self.nnd_summary_vars = {}
        self.nnd_help_btn = None
        if self.nnd_help_window is not None and self.nnd_help_window.winfo_exists():
            self.nnd_help_window.destroy()
        self.nnd_help_window = None
        self.nnd_help_tracking = False

    def _toggle_nnd_help(self) -> None:
        if self.nnd_summary_window is None or not self.nnd_summary_window.winfo_exists():
            return
        if self.nnd_help_window is not None and self.nnd_help_window.winfo_exists():
            self.nnd_help_window.destroy()
            self.nnd_help_window = None
            self.nnd_help_tracking = False
            return

        self.nnd_help_window = tk.Toplevel(self.nnd_summary_window)
        self.nnd_help_window.overrideredirect(True)
        self.nnd_help_window.attributes("-topmost", True)

        bubble = tk.Frame(self.nnd_help_window, bg="#fff7e6", bd=1, relief="solid")
        bubble.pack(fill=tk.BOTH, expand=True)

        def add_paragraph(title: str, body: str, last: bool = False) -> None:
            text = tk.Text(
                bubble,
                bg="#fff7e6",
                fg="#333333",
                wrap="word",
                width=60,
                height=3,
                relief="flat",
                highlightthickness=0,
            )
            text.tag_configure("title", font=("Segoe UI", 9, "bold"))
            text.tag_configure("body", font=("Segoe UI", 9))
            text.insert("1.0", title, "title")
            text.insert("end", f" {body}", "body")
            text.configure(state="disabled")
            text.pack(anchor="w", padx=8, pady=(8, 4) if not last else (0, 8))

        add_paragraph(
            "NND (Nearest Neighbor Distance) —",
            "is a measured property: it quantifies the distance from each particle to its closest neighbor, describing the actual spatial arrangement in your image.",
        )
        add_paragraph(
            "CSR (Complete Spatial Randomness) —",
            "is a reference model: it represents how particles would be distributed if placed completely at random, and is used as a baseline to determine whether the observed NND indicates clustering or ordering.",
        )
        add_paragraph(
            "Monte-Carlo B (default 300) —",
            "number of random CSR simulations used to build the null distribution; more simulations give more stable simulated p-values but take more time.",
        )
        add_paragraph(
            "D10 —",
            "the 10th percentile of the nearest-neighbor distances: 10% of particles have a nearest neighbor closer than this distance (D10 is useful to detect very tight packing and D90 to characterise the long-distance tail).",
        )
        add_paragraph(
            "var_CSR —",
            "theoretical variance of nearest-neighbor distance under Complete Spatial Randomness; it describes how much the NND would fluctuate by chance in a random pattern.",
        )
        add_paragraph(
            "z-statistic —",
            "standardized score comparing the observed mean NND to the CSR expected mean (difference divided by the standard error); large |z| means the observed pattern differs substantially from CSR.",
        )
        add_paragraph(
            "p (simulated) —",
            "empirical p-value from the Monte-Carlo runs: fraction of CSR simulations as extreme (or more) than the observed statistic; small values indicate non-randomness.",
        )
        add_paragraph(
            "Clark–Evans R —",
            "ratio of observed mean NND to the CSR expected mean (R≈1 random; R<1 clustering; R>1 regular spacing).",
        )
        add_paragraph(
            "p (theoretical) —",
            "p-value computed from the z-statistic using the normal approximation (quick test against CSR; compare with the simulated p for robustness).",
            last=True,
        )

        self.nnd_help_tracking = True
        if self.nnd_summary_window is not None:
            self.nnd_summary_window.bind("<Configure>", self._on_nnd_window_configure)
        self._position_nnd_help()

    def _on_nnd_window_configure(self, _event=None) -> None:
        if self.nnd_help_tracking:
            self._position_nnd_help()

    def _position_nnd_help(self) -> None:
        if self.nnd_help_window is None or not self.nnd_help_window.winfo_exists():
            return
        if self.nnd_summary_window is None or not self.nnd_summary_window.winfo_exists():
            return
        self.nnd_help_window.update_idletasks()
        bubble_w = self.nnd_help_window.winfo_width()
        bubble_h = self.nnd_help_window.winfo_height()
        win_x = self.nnd_summary_window.winfo_rootx()
        win_y = self.nnd_summary_window.winfo_rooty()
        x = win_x + 20
        y = win_y - bubble_h - 8
        if y < 0:
            y = win_y + 30
        self.nnd_help_window.geometry(f"{bubble_w}x{bubble_h}+{x}+{y}")

    def _close_nnd_plot_window(self) -> None:
        if self.nnd_plot_window is not None and self.nnd_plot_window.winfo_exists():
            self.nnd_plot_window.destroy()
        self.nnd_plot_window = None
        if self.nnd_map_canvas is not None and self.nnd_map_hover_cid is not None:
            try:
                self.nnd_map_canvas.mpl_disconnect(self.nnd_map_hover_cid)
            except Exception:
                pass
        self.nnd_map_hover_cid = None
        self.nnd_map_fig = None
        self.nnd_map_ax = None
        self.nnd_map_hist_ax = None
        self.nnd_map_canvas = None
        self.nnd_map_colorbar = None
        self.nnd_map_hist_values = None
        self.nnd_map_hist_bins = None
        self.nnd_map_kde_line = None
        self.nnd_map_kde_annotation = None
        self.nnd_map_kde_enabled = False
        self.nnd_map_kde_x = None
        self.nnd_map_kde_y = None
        self.nnd_map_kde_btn = None
        self.nnd_hist_fig = None
        self.nnd_hist_ax = None
        self.nnd_hist_canvas = None

    def _parse_nnd_monte_carlo(self) -> int:
        if self.nnd_mc_var is None:
            return 300
        try:
            b = int(self.nnd_mc_var.get().strip())
        except Exception:
            b = 300
        b = max(20, b)
        self.nnd_mc_var.set(str(b))
        return b

    def _compute_nnd_results(self) -> Dict[str, object]:
        selected = self._selected_particles()
        if len(selected) < 2:
            raise ValueError("Select at least 2 particles")

        nm_per_px = self._safe_um_per_px_for_refresh() * 1000.0
        ids = np.asarray([int(p.mask_id) for p in selected], dtype=int)
        centroids_px = np.asarray(
            [[float(c[0]), float(c[1])] for c in (self._particle_centroid_global(p) for p in selected)],
            dtype=float,
        )
        centroids_nm = centroids_px * nm_per_px

        tree = cKDTree(centroids_nm)
        distances, nn_idx = tree.query(centroids_nm, k=2)
        nnd_nm = np.asarray(distances[:, 1], dtype=float)
        nearest_ids = ids[np.asarray(nn_idx[:, 1], dtype=int)]

        n = int(nnd_nm.size)
        mean_obs = float(np.mean(nnd_nm))
        median_obs = float(np.median(nnd_nm))
        std_obs = float(np.std(nnd_nm))
        d10 = float(np.percentile(nnd_nm, 10))
        d50 = float(np.percentile(nnd_nm, 50))
        d90 = float(np.percentile(nnd_nm, 90))

        active_images = self._active_images()
        if active_images:
            areas_px = np.asarray([img.shape[0] * img.shape[1] for img in active_images], dtype=float)
            total_area_px = float(np.sum(areas_px))
            area_nm2 = total_area_px * (nm_per_px ** 2) if total_area_px > 0 else float("nan")
            if self.image_rgb is not None:
                h, w = self.image_rgb.shape[:2]
                width_nm = float(w) * nm_per_px
                height_nm = float(h) * nm_per_px
            else:
                span = np.ptp(centroids_nm, axis=0)
                width_nm = float(max(1e-6, span[0]))
                height_nm = float(max(1e-6, span[1]))
        else:
            span = np.ptp(centroids_nm, axis=0)
            width_nm = float(max(1e-6, span[0]))
            height_nm = float(max(1e-6, span[1]))
            area_nm2 = width_nm * height_nm

        lam = float(n / area_nm2) if area_nm2 > 0 else float("nan")
        mean_csr = float(1.0 / (2.0 * math.sqrt(lam))) if lam > 0 else float("nan")
        var_csr = float((4.0 - math.pi) / (4.0 * math.pi * lam)) if lam > 0 else float("nan")
        r_index = float(mean_obs / mean_csr) if mean_csr > 0 else float("nan")
        se_mean = float(0.26136 / math.sqrt(n * lam)) if (n > 1 and lam > 0) else float("nan")
        z_value = float((mean_obs - mean_csr) / se_mean) if se_mean > 0 else float("nan")
        p_theoretical = float(math.erfc(abs(z_value) / math.sqrt(2.0))) if np.isfinite(z_value) else float("nan")

        b = self._parse_nnd_monte_carlo()
        rng = np.random.default_rng()
        areas_px = None
        probs = None
        if active_images:
            areas_px = np.asarray([img.shape[0] * img.shape[1] for img in active_images], dtype=float)
            total_area = float(np.sum(areas_px))
            if total_area > 0:
                probs = areas_px / total_area

        def sample_csr_points_px(n_points: int) -> np.ndarray:
            if not active_images or probs is None:
                return np.column_stack(
                    (
                        rng.uniform(0.0, width_nm / nm_per_px, size=n_points),
                        rng.uniform(0.0, height_nm / nm_per_px, size=n_points),
                    )
                )
            choices = rng.choice(len(active_images), size=n_points, p=probs)
            pts = np.zeros((n_points, 2), dtype=float)
            for idx in range(len(active_images)):
                mask = choices == idx
                if not np.any(mask):
                    continue
                img = active_images[idx]
                h, w = img.shape[:2]
                off_x, off_y = self.image_offsets.get(idx, (0, 0))
                pts[mask, 0] = rng.uniform(0.0, float(w), size=int(np.sum(mask))) + float(off_x)
                pts[mask, 1] = rng.uniform(0.0, float(h), size=int(np.sum(mask))) + float(off_y)
            return pts

        sim_means = np.zeros(b, dtype=float)
        x_max = float(max(np.percentile(nnd_nm, 99.5), mean_obs * 3.0, 1e-6))
        bin_edges = np.linspace(0.0, x_max, 31, dtype=float)
        sim_hist = np.zeros((b, len(bin_edges) - 1), dtype=float)
        for i in range(b):
            sim_pts_px = sample_csr_points_px(n)
            sim_pts_nm = sim_pts_px * nm_per_px
            sim_tree = cKDTree(sim_pts_nm)
            sim_d, _ = sim_tree.query(sim_pts_nm, k=2)
            sim_nnd = np.asarray(sim_d[:, 1], dtype=float)
            sim_means[i] = float(np.mean(sim_nnd))
            sim_hist[i, :] = np.histogram(sim_nnd, bins=bin_edges, density=True)[0]

        center = float(np.mean(sim_means))
        p_simulated = float((1 + np.sum(np.abs(sim_means - center) >= abs(mean_obs - center))) / (b + 1))
        envelope_low = np.percentile(sim_hist, 2.5, axis=0)
        envelope_high = np.percentile(sim_hist, 97.5, axis=0)
        bin_centers = 0.5 * (bin_edges[:-1] + bin_edges[1:])

        x_grid = np.linspace(0.0, x_max, 350, dtype=float)
        kde_vals = None
        if nnd_nm.size >= 3:
            try:
                kde_vals = gaussian_kde(nnd_nm)(x_grid)
            except Exception:
                kde_vals = None
        csr_pdf = 2.0 * math.pi * lam * x_grid * np.exp(-math.pi * lam * (x_grid ** 2)) if lam > 0 else np.zeros_like(x_grid)

        return {
            "ids": ids,
            "nearest_ids": nearest_ids,
            "centroids_px": centroids_px,
            "nnd_nm": nnd_nm,
            "bin_edges": bin_edges,
            "bin_centers": bin_centers,
            "envelope_low": envelope_low,
            "envelope_high": envelope_high,
            "x_grid": x_grid,
            "kde_vals": kde_vals,
            "csr_pdf": csr_pdf,
            "summary": {
                "n": n,
                "a_nm2": area_nm2,
                "lambda": lam,
                "mean": mean_obs,
                "median": median_obs,
                "std": std_obs,
                "d10": d10,
                "d50": d50,
                "d90": d90,
                "mean_csr": mean_csr,
                "var_csr": var_csr,
                "clark_evans_r": r_index,
                "z": z_value,
                "p_theoretical": p_theoretical,
                "p_simulated": p_simulated,
                "b": b,
            },
        }

    def _refresh_nnd_outputs(self) -> None:
        if self.nnd_summary_window is None and self.nnd_plot_window is None:
            return
        if len(self.selected_ids) < 2:
            messagebox.showerror("Error", "Select at least 2 particles")
            return

        try:
            results = self._compute_nnd_results()
        except Exception as exc:
            messagebox.showerror("NND Error", str(exc))
            return
        self.nnd_results = results

        summary = results["summary"]
        if self.nnd_basis_var is not None:
            self.nnd_basis_var.set("NND computed with KD-tree on selected centroids; CSR compared on full image area.")

        def fmt(value: float, decimals: int = 6) -> str:
            return f"{value:.{decimals}f}" if np.isfinite(value) else "N/A"

        for key, var in self.nnd_summary_vars.items():
            value = summary.get(key, float("nan"))
            if key == "n":
                var.set(str(int(value)))
            elif key in ("p_theoretical", "p_simulated"):
                var.set(fmt(float(value), 6))
            elif key in ("mean", "median", "std", "d10", "d50", "d90", "mean_csr"):
                var.set(fmt(float(value), 4))
            elif key in ("a_nm2", "var_csr"):
                var.set(fmt(float(value), 4))
            elif key in ("lambda",):
                var.set(fmt(float(value), 8))
            elif key in ("clark_evans_r", "z"):
                var.set(fmt(float(value), 5))
            else:
                var.set(fmt(float(value), 5))

        if self.nnd_map_fig is not None and self.nnd_map_canvas is not None:
            self.nnd_map_fig.clear()
            gs = self.nnd_map_fig.add_gridspec(1, 2, width_ratios=[1.0, 1.6], wspace=0.25)
            self.nnd_map_hist_ax = self.nnd_map_fig.add_subplot(gs[0, 0])
            self.nnd_map_ax = self.nnd_map_fig.add_subplot(gs[0, 1])

            centroids_px = results["centroids_px"]
            nnd_nm = results["nnd_nm"]
            nearest_ids = results["nearest_ids"]
            ids = results["ids"]

            self.nnd_map_hist_values = np.asarray(nnd_nm, dtype=float)
            bins = max(8, min(30, int(np.sqrt(self.nnd_map_hist_values.size) * 2)))
            self.nnd_map_hist_bins = bins
            self.nnd_map_hist_ax.hist(
                self.nnd_map_hist_values,
                bins=bins,
                color="#7fb3d5",
                edgecolor="white",
                linewidth=0.6,
                alpha=0.9,
            )
            self.nnd_map_hist_ax.set_title("NND Histogram")
            self.nnd_map_hist_ax.set_xlabel("NND (nm)")
            self.nnd_map_hist_ax.set_ylabel("Count")
            self.nnd_map_hist_ax.grid(alpha=0.25, linestyle="--")

            self.nnd_map_kde_line = None
            self.nnd_map_kde_x = None
            self.nnd_map_kde_y = None
            if self.nnd_map_kde_annotation is not None:
                try:
                    self.nnd_map_kde_annotation.set_visible(False)
                except Exception:
                    pass

            ax = self.nnd_map_ax
            if self.image_rgb is not None:
                ax.imshow(self.image_rgb, interpolation="nearest")
            sc = ax.scatter(
                centroids_px[:, 0],
                centroids_px[:, 1],
                c=nnd_nm,
                cmap="viridis",
                s=36,
                edgecolors="white",
                linewidths=0.45,
            )
            id_to_idx = {int(pid): i for i, pid in enumerate(ids.tolist())}
            for i, nn_pid in enumerate(nearest_ids.tolist()):
                j = id_to_idx.get(int(nn_pid), None)
                if j is None:
                    continue
                ax.plot(
                    [centroids_px[i, 0], centroids_px[j, 0]],
                    [centroids_px[i, 1], centroids_px[j, 1]],
                    color="#00c2ff",
                    alpha=0.2,
                    linewidth=0.8,
                )
            ax.set_title("Centroid map colored by nearest-neighbor distance")
            ax.set_xlabel("x (px)")
            ax.set_ylabel("y (px)")
            if self.image_rgb is not None:
                h, w = self.image_rgb.shape[:2]
                ax.set_xlim(0, w - 1)
                ax.set_ylim(h - 1, 0)
            self.nnd_map_colorbar = self.nnd_map_fig.colorbar(sc, ax=ax, label="NND (nm)")
            if self.nnd_map_kde_enabled:
                self._render_nnd_map_kde()
            self.nnd_map_fig.tight_layout()
            self.nnd_map_canvas.draw_idle()

        if self.nnd_hist_fig is not None and self.nnd_hist_canvas is not None:
            self.nnd_hist_fig.clear()
            self.nnd_hist_ax = self.nnd_hist_fig.add_subplot(111)
            ax = self.nnd_hist_ax
            nnd_nm = results["nnd_nm"]
            bin_edges = results["bin_edges"]
            x_grid = results["x_grid"]
            kde_vals = results["kde_vals"]
            csr_pdf = results["csr_pdf"]
            bin_centers = results["bin_centers"]
            env_low = results["envelope_low"]
            env_high = results["envelope_high"]

            ax.hist(nnd_nm, bins=bin_edges, density=True, color="#1f77b4", alpha=0.5, edgecolor="white", label="Observed histogram")
            if kde_vals is not None:
                ax.plot(x_grid, kde_vals, color="#003f88", linewidth=2.0, label="Observed KDE")
            ax.plot(x_grid, csr_pdf, color="#e63946", linewidth=2.0, label="CSR theoretical pdf")
            ax.fill_between(bin_centers, env_low, env_high, color="#f4a261", alpha=0.25, label="CSR MC envelope (95%)")
            ax.set_title("Nearest-neighbor distribution vs CSR")
            ax.set_xlabel("Nearest-neighbor distance (nm)")
            ax.set_ylabel("Density")
            ax.grid(alpha=0.25, linestyle="--")
            ax.legend(loc="best")
            self.nnd_hist_fig.tight_layout()
            self.nnd_hist_canvas.draw_idle()

    def _toggle_nnd_map_kde(self) -> None:
        if self.nnd_map_hist_ax is None or self.nnd_map_canvas is None:
            return
        if not NND_AVAILABLE:
            messagebox.showerror("KDE fitting", f"NND dependencies are missing.\n{NND_IMPORT_ERROR}")
            return
        self.nnd_map_kde_enabled = not self.nnd_map_kde_enabled
        self._render_nnd_map_kde()

    def _render_nnd_map_kde(self) -> None:
        if self.nnd_map_hist_ax is None or self.nnd_map_canvas is None:
            return
        if self.nnd_map_kde_line is not None:
            try:
                self.nnd_map_kde_line.remove()
            except Exception:
                pass
            self.nnd_map_kde_line = None
        if self.nnd_map_kde_annotation is not None:
            try:
                self.nnd_map_kde_annotation.set_visible(False)
            except Exception:
                pass

        if not self.nnd_map_kde_enabled:
            self.nnd_map_kde_x = None
            self.nnd_map_kde_y = None
            self.nnd_map_canvas.draw_idle()
            return

        values = self.nnd_map_hist_values
        if values is None or values.size < 3:
            messagebox.showerror("KDE fitting", "Select at least 3 particles to fit a KDE.")
            self.nnd_map_kde_enabled = False
            return

        x_min = float(np.min(values))
        x_max = float(np.max(values))
        if not np.isfinite(x_min) or not np.isfinite(x_max) or x_max <= x_min:
            messagebox.showerror("KDE fitting", "KDE fitting requires a range of values.")
            self.nnd_map_kde_enabled = False
            return

        try:
            kde = gaussian_kde(values)
        except Exception as exc:
            messagebox.showerror("KDE fitting", f"Could not fit KDE:\n{exc}")
            self.nnd_map_kde_enabled = False
            return

        x_grid = np.linspace(x_min, x_max, 300)
        y_vals = kde(x_grid)
        bins = self.nnd_map_hist_bins or max(8, min(30, int(np.sqrt(values.size) * 2)))
        bin_width = (x_max - x_min) / float(max(1, bins))
        y_scaled = y_vals * values.size * bin_width

        line, = self.nnd_map_hist_ax.plot(x_grid, y_scaled, color="#e53935", linewidth=2.0, label="KDE")
        self.nnd_map_kde_line = line
        self.nnd_map_kde_x = x_grid
        self.nnd_map_kde_y = y_scaled
        try:
            self.nnd_map_hist_ax.legend(loc="best", fontsize=8)
        except Exception:
            pass
        self.nnd_map_canvas.draw_idle()

    def _on_nnd_map_kde_hover(self, event) -> None:
        if (
            self.nnd_map_hist_ax is None
            or self.nnd_map_canvas is None
            or self.nnd_map_kde_line is None
            or self.nnd_map_kde_x is None
            or self.nnd_map_kde_y is None
        ):
            return
        if event.inaxes != self.nnd_map_hist_ax or event.xdata is None or event.ydata is None:
            if self.nnd_map_kde_annotation is not None:
                self.nnd_map_kde_annotation.set_visible(False)
                self.nnd_map_canvas.draw_idle()
            return

        x_vals = self.nnd_map_kde_x
        y_vals = self.nnd_map_kde_y
        idx = int(np.searchsorted(x_vals, float(event.xdata)))
        idx = max(0, min(idx, len(x_vals) - 1))
        x = float(x_vals[idx])
        y = float(y_vals[idx])
        max_y = float(np.max(y_vals)) if y_vals.size > 0 else 0.0
        tol = max_y * 0.05 if max_y > 0 else 0.01
        if abs(float(event.ydata) - y) > tol:
            if self.nnd_map_kde_annotation is not None:
                self.nnd_map_kde_annotation.set_visible(False)
                self.nnd_map_canvas.draw_idle()
            return

        if self.nnd_map_kde_annotation is None:
            self.nnd_map_kde_annotation = self.nnd_map_hist_ax.annotate(
                "",
                xy=(x, y),
                xytext=(8, 8),
                textcoords="offset points",
                bbox=dict(boxstyle="round,pad=0.2", facecolor="#fff5f5", edgecolor="#e53935"),
                fontsize=8,
                color="#b71c1c",
            )
        self.nnd_map_kde_annotation.xy = (x, y)
        self.nnd_map_kde_annotation.set_text(f"NND: {x:.2f} nm")
        self.nnd_map_kde_annotation.set_visible(True)
        self.nnd_map_canvas.draw_idle()

        self.status_var.set("Nearest-Neighbor Analysis updated.")

    def on_save_nnd_data(self) -> None:
        if self.nnd_results is None:
            messagebox.showerror("Error", "Run analysis first.")
            return
        summary = self.nnd_results.get("summary", {})
        headers = (
            "N",
            "A (nm²)",
            "lambda (points/nm²)",
            "mean (nm)",
            "median (nm)",
            "std (nm)",
            "D10 (nm)",
            "D50 (nm)",
            "D90 (nm)",
            "mean_CSR (nm)",
            "var_CSR (nm²)",
            "R (Clark-Evans)",
            "z",
            "p_theoretical",
            "p_simulated",
            "B",
        )
        row = (
            f"{int(summary.get('n', 0))}",
            f"{float(summary.get('a_nm2', float('nan'))):.6f}",
            f"{float(summary.get('lambda', float('nan'))):.10f}",
            f"{float(summary.get('mean', float('nan'))):.6f}",
            f"{float(summary.get('median', float('nan'))):.6f}",
            f"{float(summary.get('std', float('nan'))):.6f}",
            f"{float(summary.get('d10', float('nan'))):.6f}",
            f"{float(summary.get('d50', float('nan'))):.6f}",
            f"{float(summary.get('d90', float('nan'))):.6f}",
            f"{float(summary.get('mean_csr', float('nan'))):.6f}",
            f"{float(summary.get('var_csr', float('nan'))):.6f}",
            f"{float(summary.get('clark_evans_r', float('nan'))):.6f}",
            f"{float(summary.get('z', float('nan'))):.6f}",
            f"{float(summary.get('p_theoretical', float('nan'))):.8f}",
            f"{float(summary.get('p_simulated', float('nan'))):.8f}",
            f"{int(summary.get('b', 0))}",
        )
        self._save_rows_to_csv_or_xlsx(
            headers=headers,
            rows=[row],
            dialog_title="Save NND Data",
            initialfile="nearest_neighbor_analysis.csv",
            sheet_name="NND",
        )

    def _save_figure_image(self, fig: Optional[Figure], title: str, initialfile: str) -> None:
        if fig is None:
            messagebox.showerror("Error", "No figure available.")
            return
        save_path = filedialog.asksaveasfilename(
            title=title,
            defaultextension=".png",
            initialfile=initialfile,
            filetypes=[
                ("PNG", "*.png"),
                ("JPEG", "*.jpg"),
                ("TIFF", "*.tiff"),
            ],
        )
        if not save_path:
            return
        ext = os.path.splitext(save_path)[1].lower()
        if ext not in (".png", ".jpg", ".jpeg", ".tif", ".tiff"):
            save_path = f"{save_path}.png"
        try:
            fig.savefig(save_path, dpi=300, bbox_inches="tight")
        except Exception as exc:
            messagebox.showerror("Save error", f"Could not save image:\n{exc}")
            return
        self.status_var.set(f"Image saved: {os.path.basename(save_path)}")

    def _save_axes_image(self, fig: Optional[Figure], axes: List, title: str, initialfile: str) -> None:
        if fig is None or not axes:
            messagebox.showerror("Error", "No figure available.")
            return
        save_path = filedialog.asksaveasfilename(
            title=title,
            defaultextension=".png",
            initialfile=initialfile,
            filetypes=[
                ("PNG", "*.png"),
                ("JPEG", "*.jpg"),
                ("TIFF", "*.tiff"),
            ],
        )
        if not save_path:
            return
        ext = os.path.splitext(save_path)[1].lower()
        if ext not in (".png", ".jpg", ".jpeg", ".tif", ".tiff"):
            save_path = f"{save_path}.png"
        try:
            fig.canvas.draw()
            renderer = fig.canvas.get_renderer()
            bboxes = [ax.get_tightbbox(renderer) for ax in axes if ax is not None]
            if not bboxes:
                raise RuntimeError("No axes to save.")
            bbox = Bbox.union(bboxes).transformed(fig.dpi_scale_trans.inverted())
            fig.savefig(save_path, dpi=300, bbox_inches=bbox)
        except Exception as exc:
            messagebox.showerror("Save error", f"Could not save image:\n{exc}")
            return
        self.status_var.set(f"Image saved: {os.path.basename(save_path)}")

    def on_save_nnd_map_image(self) -> None:
        axes = []
        if self.nnd_map_ax is not None:
            axes.append(self.nnd_map_ax)
        if self.nnd_map_colorbar is not None and getattr(self.nnd_map_colorbar, "ax", None) is not None:
            axes.append(self.nnd_map_colorbar.ax)
        self._save_axes_image(self.nnd_map_fig, axes, "Save NND centroid map", "nnd_centroid_map.png")

    def on_save_nnd_map_hist_image(self) -> None:
        axes = [self.nnd_map_hist_ax] if self.nnd_map_hist_ax is not None else []
        self._save_axes_image(self.nnd_map_fig, axes, "Save NND histogram", "nnd_histogram.png")

    def on_save_nnd_hist_image(self) -> None:
        self._save_figure_image(self.nnd_hist_fig, "Save NND distribution plot", "nnd_distribution.png")

    def open_compare_particle_groups_ttest(self) -> None:
        if not GMM_AVAILABLE:
            messagebox.showerror("t-test Error", f"GMM dependencies are missing.\n{GMM_IMPORT_ERROR}")
            return
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return

        if self.ttest_window is not None and self.ttest_window.winfo_exists():
            self.ttest_window.lift()
            self._refresh_ttest_window()
            return

        self.ttest_window = tk.Toplevel(self.root)
        self.ttest_window.title("Compare Particle Groups (t-test)")
        self.ttest_window.geometry("860x520+900+170")

        outer = ttk.Frame(self.ttest_window, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(outer)
        header.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(header, text="Compare Particle Groups (t-test)", font=("Segoe UI", 12, "bold")).pack(anchor="w", side=tk.LEFT)
        help_icon = self._get_help_icon()
        self.ttest_help_btn = tk.Button(
            header,
            image=help_icon,
            text="" if help_icon is not None else "?",
            relief="flat",
            bd=0,
            command=self._toggle_ttest_help,
            cursor="hand2",
        )
        self.ttest_help_btn.pack(side=tk.LEFT, padx=(6, 0))
        if help_icon is None:
            self.ttest_help_btn.configure(font=("Segoe UI", 9, "bold"))

        controls = ttk.Frame(outer)
        controls.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(controls, text="Property:").pack(side=tk.LEFT)
        self.ttest_property_var = tk.StringVar(value="length")
        prop_combo = ttk.Combobox(
            controls,
            textvariable=self.ttest_property_var,
            state="readonly",
            values=("length", "width", "area", "feret diameter", "circularity", "eccentricity"),
            width=12,
        )
        prop_combo.pack(side=tk.LEFT, padx=(8, 14))
        prop_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_ttest_window())

        ttk.Label(controls, text="Test type:").pack(side=tk.LEFT)
        self.ttest_type_var = tk.StringVar(value="Welch t-test")
        test_combo = ttk.Combobox(
            controls,
            textvariable=self.ttest_type_var,
            state="readonly",
            values=("Welch t-test", "Student t-test"),
            width=14,
        )
        test_combo.pack(side=tk.LEFT, padx=(8, 14))
        test_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_ttest_window())

        ttk.Button(controls, text="Save Data", command=self.on_save_ttest_data).pack(side=tk.RIGHT)

        self.ttest_basis_var = tk.StringVar(value="")
        ttk.Label(outer, textvariable=self.ttest_basis_var, font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 8))

        table_frame = ttk.Frame(outer)
        table_frame.pack(fill=tk.BOTH, expand=True)

        self.ttest_tree = ttk.Treeview(
            table_frame,
            columns=("pair", "test", "t_stat", "df", "p_value", "ci_low", "ci_high"),
            show="headings",
        )
        self.ttest_tree.heading("pair", text="Group Pair")
        self.ttest_tree.heading("test", text="Test")
        self.ttest_tree.heading("t_stat", text="t-statistic")
        self.ttest_tree.heading("df", text="df")
        self.ttest_tree.heading("p_value", text="p-value")
        self.ttest_tree.heading("ci_low", text="CI low")
        self.ttest_tree.heading("ci_high", text="CI high")
        self.ttest_tree.column("pair", width=140, anchor="center")
        self.ttest_tree.column("test", width=120, anchor="center")
        self.ttest_tree.column("t_stat", width=120, anchor="e")
        self.ttest_tree.column("df", width=90, anchor="e")
        self.ttest_tree.column("p_value", width=110, anchor="e")
        self.ttest_tree.column("ci_low", width=120, anchor="e")
        self.ttest_tree.column("ci_high", width=120, anchor="e")
        self.ttest_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        yscroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.ttest_tree.yview)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.ttest_tree.configure(yscrollcommand=yscroll.set)

        self.ttest_window.protocol("WM_DELETE_WINDOW", self._close_ttest_window)
        self._refresh_ttest_window()

    def _close_ttest_window(self) -> None:
        if self.ttest_window is not None and self.ttest_window.winfo_exists():
            self.ttest_window.destroy()
        self.ttest_window = None
        self.ttest_property_var = None
        self.ttest_type_var = None
        self.ttest_basis_var = None
        self.ttest_tree = None
        self.ttest_export_headers = ()
        self.ttest_export_rows = []
        self.ttest_help_btn = None
        if self.ttest_help_window is not None and self.ttest_help_window.winfo_exists():
            self.ttest_help_window.destroy()
        self.ttest_help_window = None
        self.ttest_help_tracking = False

    def _toggle_ttest_help(self) -> None:
        if self.ttest_window is None or not self.ttest_window.winfo_exists():
            return
        if self.ttest_help_window is not None and self.ttest_help_window.winfo_exists():
            self.ttest_help_window.destroy()
            self.ttest_help_window = None
            self.ttest_help_tracking = False
            return

        self.ttest_help_window = tk.Toplevel(self.ttest_window)
        self.ttest_help_window.overrideredirect(True)
        self.ttest_help_window.attributes("-topmost", True)

        bubble = tk.Frame(self.ttest_help_window, bg="#fff7e6", bd=1, relief="solid")
        bubble.pack(fill=tk.BOTH, expand=True)

        def add_paragraph(title: str, body: str, last: bool = False) -> None:
            text = tk.Text(
                bubble,
                bg="#fff7e6",
                fg="#333333",
                wrap="word",
                width=58,
                height=3,
                relief="flat",
                highlightthickness=0,
            )
            text.tag_configure("title", font=("Segoe UI", 9, "bold"))
            text.tag_configure("body", font=("Segoe UI", 9))
            text.insert("1.0", title, "title")
            text.insert("end", f" {body}", "body")
            text.configure(state="disabled")
            text.pack(anchor="w", padx=8, pady=(8, 4) if not last else (0, 8))

        add_paragraph(
            "Welch t-test —",
            "compares the means of two groups without assuming equal variances; more robust when the groups have different spreads or sample sizes.",
        )
        add_paragraph(
            "Student t-test —",
            "compares the means assuming both groups have equal variances; appropriate when this assumption is reasonably satisfied.",
        )
        add_paragraph(
            "t-statistic —",
            "measures the difference between group means relative to variability; larger absolute values indicate a stronger difference between groups.",
        )
        add_paragraph(
            "df (degrees of freedom) —",
            "reflects the effective sample size used in the test; affects how the t-statistic is translated into a p-value.",
        )
        add_paragraph(
            "p-value —",
            "probability of observing a difference as large as this (or larger) if the groups truly have the same mean; small values indicate a statistically significant difference.",
        )
        add_paragraph(
            "CI low —",
            "lower bound of the confidence interval for the difference in means; indicates the minimum plausible difference between groups.",
        )
        add_paragraph(
            "CI high —",
            "upper bound of the confidence interval for the difference in means; indicates the maximum plausible difference between groups.",
            last=True,
        )

        self.ttest_help_tracking = True
        if self.ttest_window is not None:
            self.ttest_window.bind("<Configure>", self._on_ttest_window_configure)
        self._position_ttest_help()

    def _on_ttest_window_configure(self, _event=None) -> None:
        if self.ttest_help_tracking:
            self._position_ttest_help()

    def _position_ttest_help(self) -> None:
        if self.ttest_help_window is None or not self.ttest_help_window.winfo_exists():
            return
        if self.ttest_window is None or not self.ttest_window.winfo_exists():
            return
        self.ttest_help_window.update_idletasks()
        bubble_w = self.ttest_help_window.winfo_width()
        bubble_h = self.ttest_help_window.winfo_height()
        win_x = self.ttest_window.winfo_rootx()
        win_y = self.ttest_window.winfo_rooty()
        x = win_x + 20
        y = win_y - bubble_h - 8
        if y < 0:
            y = win_y + 30
        self.ttest_help_window.geometry(f"{bubble_w}x{bubble_h}+{x}+{y}")

    def _gmm_groups_for_property(self, prop: str) -> Tuple[List[np.ndarray], List[str], str, str]:
        values, unit, basis = self._selected_metric_values(prop)
        if values.size == 0:
            return [], [], unit, basis

        model, labels = self._fit_best_gmm(values)
        if model is None or labels is None:
            return [], [], unit, basis

        means = np.asarray(model.means_, dtype=float).reshape(-1)
        order = np.argsort(means)
        groups: List[np.ndarray] = []
        labels_out: List[str] = []
        for rank, comp_idx in enumerate(order, start=1):
            group = values[labels == comp_idx]
            if group.size == 0:
                continue
            groups.append(np.asarray(group, dtype=float))
            labels_out.append(f"G{rank}")
        return groups, labels_out, unit, basis

    def _two_sample_ttest(
        self,
        x: np.ndarray,
        y: np.ndarray,
        test_type: str,
        alpha: float = 0.05,
    ) -> Tuple[float, float, float, float, float]:
        n1 = int(x.size)
        n2 = int(y.size)
        if n1 < 2 or n2 < 2:
            return float("nan"), float("nan"), float("nan"), float("nan"), float("nan")

        mean1 = float(np.mean(x))
        mean2 = float(np.mean(y))
        var1 = float(np.var(x, ddof=1))
        var2 = float(np.var(y, ddof=1))
        diff = mean1 - mean2

        if test_type.lower().startswith("student"):
            df = float(n1 + n2 - 2)
            if df <= 0:
                return float("nan"), float("nan"), float("nan"), float("nan"), float("nan")
            pooled = ((n1 - 1) * var1 + (n2 - 1) * var2) / df
            se = math.sqrt(pooled * (1.0 / n1 + 1.0 / n2)) if pooled >= 0 else float("nan")
        else:
            se = math.sqrt(var1 / n1 + var2 / n2) if n1 > 0 and n2 > 0 else float("nan")
            denom = (var1 / n1) ** 2 / (n1 - 1) + (var2 / n2) ** 2 / (n2 - 1)
            df = (var1 / n1 + var2 / n2) ** 2 / denom if denom > 0 else float("nan")

        if not np.isfinite(se) or se <= 0:
            return float("nan"), float("nan"), float("nan"), float("nan"), float("nan")

        t_stat = diff / se
        p_value = 2.0 * (1.0 - float(student_t.cdf(abs(t_stat), df))) if np.isfinite(df) else float("nan")
        t_crit = float(student_t.ppf(1.0 - alpha / 2.0, df)) if np.isfinite(df) else float("nan")
        ci_low = diff - t_crit * se if np.isfinite(t_crit) else float("nan")
        ci_high = diff + t_crit * se if np.isfinite(t_crit) else float("nan")
        return t_stat, df, p_value, ci_low, ci_high

    def _refresh_ttest_window(self) -> None:
        if self.ttest_window is None or not self.ttest_window.winfo_exists():
            return
        if self.ttest_tree is None or self.ttest_property_var is None or self.ttest_type_var is None:
            return

        prop = self.ttest_property_var.get().strip().lower()
        test_type = self.ttest_type_var.get().strip()
        groups, labels_out, unit, basis = self._gmm_groups_for_property(prop)

        if self.ttest_basis_var is not None:
            self.ttest_basis_var.set(f"GMM subpopulations on selected particle {basis}")

        for row in self.ttest_tree.get_children():
            self.ttest_tree.delete(row)

        self.ttest_export_headers = ("Group Pair", "Test", "t-statistic", "df", "p-value", "CI low", "CI high")
        self.ttest_export_rows = []

        if len(groups) < 2:
            self.status_var.set("t-test: GMM found fewer than 2 groups.")
            return

        def fmt(val: float, dec: int = 5) -> str:
            return f"{val:.{dec}f}" if np.isfinite(val) else "N/A"

        for i in range(len(groups)):
            for j in range(i + 1, len(groups)):
                t_stat, df, p_val, ci_low, ci_high = self._two_sample_ttest(groups[i], groups[j], test_type)
                pair_label = f"{labels_out[i]} vs {labels_out[j]}"
                ci_low_str = fmt(ci_low, 5)
                ci_high_str = fmt(ci_high, 5)
                if unit:
                    ci_low_str = f"{ci_low_str} {unit}"
                    ci_high_str = f"{ci_high_str} {unit}"
                row = (
                    pair_label,
                    test_type,
                    fmt(t_stat, 5),
                    fmt(df, 3),
                    fmt(p_val, 6),
                    ci_low_str,
                    ci_high_str,
                )
                self.ttest_tree.insert("", tk.END, values=row)
                self.ttest_export_rows.append(row)

    def on_save_ttest_data(self) -> None:
        if not self.ttest_export_rows:
            messagebox.showerror("Error", "Run the t-test first.")
            return
        self._save_rows_to_csv_or_xlsx(
            headers=self.ttest_export_headers,
            rows=self.ttest_export_rows,
            dialog_title="Save t-test Data",
            initialfile="ttest_results.csv",
            sheet_name="t-test",
        )

    def open_false_color_map(self) -> None:
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return
        if self.image_rgb is None:
            messagebox.showerror("Error", "Import an image first.")
            return

        if self.false_color_window is not None and self.false_color_window.winfo_exists():
            self.false_color_window.lift()
            return

        self.false_color_window = tk.Toplevel(self.root)
        self.false_color_window.title("False color map")
        self.false_color_window.geometry("1180x760+820+120")

        outer = ttk.Frame(self.false_color_window, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        left_frame = ttk.Frame(outer)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        right_frame = ttk.Frame(outer, width=260)
        right_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(10, 0))

        ttk.Label(right_frame, text="False color map", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))

        ttk.Label(right_frame, text="Value:").pack(anchor="w")
        self.false_color_property_var = tk.StringVar(value="Voronoi local density")
        prop_combo = ttk.Combobox(
            right_frame,
            textvariable=self.false_color_property_var,
            state="readonly",
            values=(
                "length", "width", "area", "feret diameter", "circularity", "eccentricity", "perimeter",
                "Nearest Neighbor Distance",
                "Voronoi local density",
            ),
            width=24,
        )
        prop_combo.pack(anchor="w", pady=(4, 10))

        ttk.Label(right_frame, text="Gradient:").pack(anchor="w")
        self.false_color_gradient_var = tk.StringVar(value="Halcyon")
        ttk.Label(right_frame, textvariable=self.false_color_gradient_var, foreground="#555555").pack(anchor="w")
        ttk.Button(right_frame, text="Color Gradients", command=self._open_false_color_gradient_picker).pack(
            anchor="w",
            pady=(4, 12),
        )

        ttk.Button(right_frame, text="Generate False Color Map", command=self._render_false_color_map).pack(
            anchor="w",
            pady=(2, 8),
            fill=tk.X,
        )
        ttk.Button(right_frame, text="Save False Color Map", command=self.on_save_false_color_map).pack(
            anchor="w",
            pady=(2, 8),
            fill=tk.X,
        )

        self.false_color_fig = Figure(figsize=(8.6, 6.0), dpi=100)
        self.false_color_ax = self.false_color_fig.add_subplot(111)
        self.false_color_canvas = FigureCanvasTkAgg(self.false_color_fig, master=left_frame)
        self.false_color_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self.false_color_window.protocol("WM_DELETE_WINDOW", self._close_false_color_window)

    def _close_false_color_window(self) -> None:
        if self.false_color_window is not None and self.false_color_window.winfo_exists():
            self.false_color_window.destroy()
        self.false_color_window = None
        self.false_color_property_var = None
        self.false_color_gradient_var = None
        self.false_color_fig = None
        self.false_color_ax = None
        self.false_color_canvas = None
        self.false_color_colorbar = None
        self.false_color_gradient_icons = []

    def _false_color_colormap_options(self) -> Dict[str, str]:
        return {
            "Caribbean": "viridis",
            "Cold": "cool",
            "DFit": "cividis",
            "Gray": "gray",
            "Gray-inverted": "gray_r",
            "Halcyon": "ocean",
            "Maple": "inferno",
            "Sky": "summer",
            "Spectral": "Spectral",
            "Spectral-white": "Spectral_r",
            "Green-Stripes-4": "Greens",
            "Green-Violet": "PRGn",
            "Green-Yellow": "YlGn",
            "Green-White": "Greens",
            "Lines": "tab20",
            "MetroPro": "cubehelix",
            "NT-MDT": "magma",
            "Neon": "hsv",
            "Olive": "YlGn",
            "Pink": "pink",
            "Plum": "PuRd",
            "Pm3d": "gnuplot2",
            "RGB-Blue": "Blues",
            "RGB-Green": "Greens",
            "RGB-Red": "Reds",
            "Rainbow1": "rainbow",
            "Rainbow2": "turbo",
            "Red": "Reds",
        }

    def _make_gradient_swatch(self, cmap_name: str, width: int = 120, height: int = 14) -> ImageTk.PhotoImage:
        cmap = matplotlib.colormaps.get_cmap(cmap_name)
        gradient = np.linspace(0.0, 1.0, width, dtype=float)
        rgba = cmap(gradient)
        rgb = np.clip(rgba[:, :3] * 255.0, 0, 255).astype(np.uint8)
        img = np.tile(rgb[np.newaxis, :, :], (height, 1, 1))
        pil_img = Image.fromarray(img, mode="RGB")
        return ImageTk.PhotoImage(pil_img)

    def _open_false_color_gradient_picker(self) -> None:
        options = self._false_color_colormap_options()
        picker = tk.Toplevel(self.false_color_window)
        picker.title("Choose Color Gradient")
        picker.geometry("360x460+900+200")

        frame = ttk.Frame(picker, padding=8)
        frame.pack(fill=tk.BOTH, expand=True)

        tree = ttk.Treeview(frame, columns=("name",), show="tree", height=16)
        tree.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        self.false_color_gradient_icons = []
        for name, cmap_name in options.items():
            swatch = self._make_gradient_swatch(cmap_name, width=120, height=14)
            self.false_color_gradient_icons.append(swatch)
            tree.insert("", tk.END, text=name, image=swatch)

        def apply_selection() -> None:
            sel = tree.selection()
            if not sel:
                return
            chosen = tree.item(sel[0], "text")
            if self.false_color_gradient_var is not None:
                self.false_color_gradient_var.set(chosen)
            picker.destroy()

        tree.bind("<Double-Button-1>", lambda _evt: apply_selection())
        ttk.Button(frame, text="Select", command=apply_selection).pack(anchor="e")

    def _perimeter_px(self, mask: np.ndarray) -> float:
        if mask.size == 0:
            return 0.0
        mask_bool = np.asarray(mask, dtype=bool)
        if not np.any(mask_bool):
            return 0.0

        # Prefer Crofton perimeter when scikit-image is available because it is
        # substantially less biased than raw border-pixel counting.
        if SKIMAGE_AVAILABLE:
            try:
                return float(perimeter_crofton(mask_bool, directions=4))
            except Exception:
                pass

        if CV2_AVAILABLE:
            try:
                mask_u8 = (mask_bool.astype(np.uint8) * 255)
                contours, _hier = cv2.findContours(mask_u8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
                if contours:
                    return float(sum(cv2.arcLength(cnt, True) for cnt in contours))
            except Exception:
                pass

        pad = np.pad(mask, 1, mode="constant", constant_values=False)
        center = pad[1:-1, 1:-1]
        up = pad[:-2, 1:-1]
        down = pad[2:, 1:-1]
        left = pad[1:-1, :-2]
        right = pad[1:-1, 2:]
        edge = center & (~up | ~down | ~left | ~right)
        return float(np.sum(edge))

    def _ellipse_perimeter_from_area_ratio(self, area_px2: float, ratio: float) -> float:
        ratio = max(1.0e-6, min(1.0, float(ratio)))
        a = math.sqrt(max(area_px2, 0.0) / (math.pi * ratio))
        b = ratio * a
        if a <= 0.0 or b <= 0.0:
            return 0.0
        h = ((a - b) / (a + b)) ** 2
        return float(
            math.pi * (a + b) * (1.0 + (3.0 * h) / (10.0 + math.sqrt(max(1.0e-12, 4.0 - 3.0 * h))))
        )

    def _crofton_equivalent_eccentricity(self, area_px2: float, perimeter_px: float) -> float:
        if area_px2 <= 0.0 or perimeter_px <= 0.0:
            return 0.0

        # Circle is the minimum-perimeter shape for a given area.
        circle_perimeter = 2.0 * math.sqrt(math.pi * area_px2)
        if perimeter_px <= circle_perimeter:
            return 0.0

        lo = 1.0e-6
        hi = 1.0
        for _ in range(64):
            mid = 0.5 * (lo + hi)
            p_mid = self._ellipse_perimeter_from_area_ratio(area_px2, mid)
            # For fixed area, the perimeter decreases as b/a approaches 1.
            if p_mid > perimeter_px:
                lo = mid
            else:
                hi = mid
        ratio = 0.5 * (lo + hi)
        return float(math.sqrt(max(0.0, 1.0 - ratio * ratio)))

    def _circularity_eccentricity(self, particle: ParticleMask) -> Tuple[float, float]:
        seg = particle.segmentation
        area = float(max(0, particle.area_px))
        per = float(self._perimeter_px(seg))
        if per <= 0.0 or area <= 0.0:
            circularity = 0.0
        else:
            circularity = float((4.0 * math.pi * area) / (per * per))
            circularity = max(0.0, min(1.0, circularity))

        eccentricity = self._crofton_equivalent_eccentricity(area, per)
        return circularity, eccentricity

    def _feret_diameter_px(self, particle: ParticleMask) -> float:
        cached = getattr(particle, "feret_px", None)
        if cached is not None:
            return float(cached)

        seg = particle.segmentation
        if seg is None or seg.size == 0:
            particle.feret_px = 0.0
            return 0.0

        pad = np.pad(seg, 1, mode="constant", constant_values=False)
        center = pad[1:-1, 1:-1]
        up = pad[:-2, 1:-1]
        down = pad[2:, 1:-1]
        left = pad[1:-1, :-2]
        right = pad[1:-1, 2:]
        edge = center & (~up | ~down | ~left | ~right)
        ys, xs = np.nonzero(edge)
        if xs.size < 2:
            particle.feret_px = 0.0
            return 0.0

        pts = np.column_stack((xs.astype(float), ys.astype(float)))
        max_points = 2000
        if pts.shape[0] > max_points:
            idx = np.linspace(0, pts.shape[0] - 1, max_points).astype(int)
            pts = pts[idx]

        diff = pts[:, None, :] - pts[None, :, :]
        dist2 = np.sum(diff * diff, axis=2)
        feret = float(math.sqrt(np.max(dist2))) if dist2.size > 0 else 0.0
        particle.feret_px = feret
        return feret

    def _polygon_area(self, pts: np.ndarray) -> float:
        if pts.shape[0] < 3:
            return 0.0
        x = pts[:, 0]
        y = pts[:, 1]
        return float(0.5 * abs(np.dot(x, np.roll(y, -1)) - np.dot(y, np.roll(x, -1))))

    def _compute_false_color_values(self, prop: str) -> Tuple[np.ndarray, str]:
        selected = self._selected_particles()
        if not selected:
            raise ValueError("Select particles")

        nm_per_px = self._safe_um_per_px_for_refresh() * 1000.0
        if prop == "length":
            values = [self._center_length_width_px(p)[0] * nm_per_px for p in selected]
            return np.asarray(values, dtype=float), "nm"
        if prop == "width":
            values = [self._center_length_width_px(p)[1] * nm_per_px for p in selected]
            return np.asarray(values, dtype=float), "nm"
        if prop == "area":
            values = [float(p.area_px) * (nm_per_px ** 2) for p in selected]
            return np.asarray(values, dtype=float), "nm^2"
        if prop == "feret diameter":
            values = [self._feret_diameter_px(p) * nm_per_px for p in selected]
            return np.asarray(values, dtype=float), "nm"
        if prop == "feret diameter":
            values = [self._feret_diameter_px(p) * nm_per_px for p in selected]
            return np.asarray(values, dtype=float), "nm"
        if prop == "circularity":
            values = [self._circularity_eccentricity(p)[0] for p in selected]
            return np.asarray(values, dtype=float), ""
        if prop == "eccentricity":
            values = [self._circularity_eccentricity(p)[1] for p in selected]
            return np.asarray(values, dtype=float), ""
        if prop == "circularity":
            values = [self._circularity_eccentricity(p)[0] for p in selected]
            return np.asarray(values, dtype=float), ""
        if prop == "eccentricity":
            values = [self._circularity_eccentricity(p)[1] for p in selected]
            return np.asarray(values, dtype=float), ""
        if prop == "perimeter":
            values = [self._perimeter_px(p.segmentation) * nm_per_px for p in selected]
            return np.asarray(values, dtype=float), "nm"
        if prop == "Nearest Neighbor Distance":
            if not NND_AVAILABLE:
                raise ValueError(f"NND dependencies are missing.\n{NND_IMPORT_ERROR}")
            if len(selected) < 2:
                raise ValueError("Select at least 2 particles")
            centroids_nm = np.asarray(
                [
                    [
                        float(self._particle_centroid_global(p)[0]) * nm_per_px,
                        float(self._particle_centroid_global(p)[1]) * nm_per_px,
                    ]
                    for p in selected
                ],
                dtype=float,
            )
            tree = cKDTree(centroids_nm)
            distances, _ = tree.query(centroids_nm, k=2)
            return np.asarray(distances[:, 1], dtype=float), "nm"
        if prop == "Voronoi local density":
            if not VORONOI_AVAILABLE:
                raise ValueError(f"Voronoi dependencies are missing.\n{VORONOI_IMPORT_ERROR}")
            centroids_nm = np.asarray(
                [
                    [
                        float(self._particle_centroid_global(p)[0]) * nm_per_px,
                        float(self._particle_centroid_global(p)[1]) * nm_per_px,
                    ]
                    for p in selected
                ],
                dtype=float,
            )
            vor = Voronoi(centroids_nm)
            densities: List[float] = []
            for i, region_idx in enumerate(vor.point_region):
                region = vor.regions[region_idx]
                if not region or -1 in region:
                    densities.append(float("nan"))
                    continue
                poly = vor.vertices[region]
                area = self._polygon_area(poly)
                densities.append((1.0 / area) if area > 0 else float("nan"))
            arr = np.asarray(densities, dtype=float)
            finite = np.isfinite(arr)
            if not np.any(finite):
                raise ValueError("Voronoi density could not be computed.")
            min_val = float(np.nanmin(arr[finite]))
            arr[~finite] = min_val
            return arr, "1/nm²"

        raise ValueError("Unsupported property")

    def _render_false_color_map(self) -> None:
        if self.false_color_fig is None or self.false_color_ax is None or self.false_color_canvas is None:
            return
        if self.image_rgb is None:
            messagebox.showerror("Error", "Import an image first.")
            return
        if self.false_color_property_var is None or self.false_color_gradient_var is None:
            return

        prop = self.false_color_property_var.get().strip()
        try:
            values, unit = self._compute_false_color_values(prop)
        except Exception as exc:
            messagebox.showerror("False color map", str(exc))
            return

        finite = np.isfinite(values)
        if not np.any(finite):
            messagebox.showerror("False color map", "No valid values for selected property.")
            return
        vmin = float(np.min(values[finite]))
        vmax = float(np.max(values[finite]))
        if vmax <= vmin:
            vmax = vmin + 1.0

        gradient_name = self.false_color_gradient_var.get().strip()
        cmap_name = self._false_color_colormap_options().get(gradient_name, "viridis")
        cmap = matplotlib.colormaps.get_cmap(cmap_name)
        norm = (values - vmin) / (vmax - vmin)
        norm = np.clip(norm, 0.0, 1.0)

        h, w = self.image_rgb.shape[:2]
        overlay = np.zeros((h, w, 4), dtype=np.float32)
        selected = self._selected_particles()
        for idx, p in enumerate(selected):
            color = cmap(norm[idx])
            self._apply_mask_to_canvas(
                overlay,
                p,
                np.array([color[0], color[1], color[2], 0.65], dtype=np.float32),
            )

        self.false_color_ax.clear()
        self.false_color_ax.imshow(self.image_rgb, interpolation="nearest")
        self.false_color_ax.imshow(overlay, interpolation="nearest")
        self.false_color_ax.set_title(f"False color map: {prop}")
        self.false_color_ax.axis("off")

        if self.false_color_colorbar is not None:
            try:
                self.false_color_colorbar.remove()
            except Exception:
                pass
            self.false_color_colorbar = None

        sm = matplotlib.cm.ScalarMappable(cmap=cmap, norm=matplotlib.colors.Normalize(vmin=vmin, vmax=vmax))
        sm.set_array([])
        self.false_color_colorbar = self.false_color_fig.colorbar(sm, ax=self.false_color_ax, fraction=0.046, pad=0.02)
        label = f"{prop} ({unit})" if unit else prop
        self.false_color_colorbar.set_label(label)

        self.false_color_fig.tight_layout()
        self.false_color_canvas.draw_idle()
        self.status_var.set("False color map generated.")

    def on_save_false_color_map(self) -> None:
        self._save_figure_image(self.false_color_fig, "Save False Color Map", "false_color_map.png")

    def open_violin_box_plots(self) -> None:
        if not GMM_AVAILABLE:
            messagebox.showerror("Violin/Box Error", f"GMM dependencies are missing.\n{GMM_IMPORT_ERROR}")
            return
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return

        if self.violin_window is not None and self.violin_window.winfo_exists():
            self.violin_window.lift()
            self._refresh_violin_plot()
            return

        self.violin_window = tk.Toplevel(self.root)
        self.violin_window.title("Violin + Box Plots")
        self.violin_window.geometry("980x720+880+140")

        outer = ttk.Frame(self.violin_window, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        controls = ttk.Frame(outer)
        controls.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(controls, text="Property:").pack(side=tk.LEFT)
        self.violin_property_var = tk.StringVar(value="length")
        prop_combo = ttk.Combobox(
            controls,
            textvariable=self.violin_property_var,
            state="readonly",
            values=("length", "width", "area", "feret diameter", "circularity", "eccentricity"),
            width=12,
        )
        prop_combo.pack(side=tk.LEFT, padx=(8, 12))
        prop_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_violin_plot())
        help_icon = self._get_help_icon()
        self.violin_help_btn = tk.Button(
            controls,
            image=help_icon,
            text="" if help_icon is not None else "?",
            relief="flat",
            bd=0,
            command=self._toggle_violin_help,
            cursor="hand2",
        )
        self.violin_help_btn.pack(side=tk.LEFT, padx=(0, 12))
        if help_icon is None:
            self.violin_help_btn.configure(font=("Segoe UI", 9, "bold"))

        ttk.Button(controls, text="Save Graph", command=self.on_save_violin_plot).pack(side=tk.RIGHT)

        self.violin_fig = Figure(figsize=(8.8, 6.4), dpi=100)
        self.violin_ax = self.violin_fig.add_subplot(111)
        self.violin_canvas = FigureCanvasTkAgg(self.violin_fig, master=outer)
        self.violin_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self.violin_window.protocol("WM_DELETE_WINDOW", self._close_violin_window)
        self._refresh_violin_plot()

    def _close_violin_window(self) -> None:
        if self.violin_window is not None and self.violin_window.winfo_exists():
            self.violin_window.destroy()
        self.violin_window = None
        self.violin_property_var = None
        self.violin_fig = None
        self.violin_ax = None
        self.violin_canvas = None
        self.violin_help_btn = None
        if self.violin_help_window is not None and self.violin_help_window.winfo_exists():
            self.violin_help_window.destroy()
        self.violin_help_window = None
        self.violin_help_tracking = False

    def _toggle_violin_help(self) -> None:
        if self.violin_window is None or not self.violin_window.winfo_exists():
            return
        if self.violin_help_window is not None and self.violin_help_window.winfo_exists():
            self.violin_help_window.destroy()
            self.violin_help_window = None
            self.violin_help_tracking = False
            return

        self.violin_help_window = tk.Toplevel(self.violin_window)
        self.violin_help_window.overrideredirect(True)
        self.violin_help_window.attributes("-topmost", True)

        bubble = tk.Frame(self.violin_help_window, bg="#fff7e6", bd=1, relief="solid")
        bubble.pack(fill=tk.BOTH, expand=True)

        def add_paragraph(title: str, body: str, last: bool = False) -> None:
            text = tk.Text(
                bubble,
                bg="#fff7e6",
                fg="#333333",
                wrap="word",
                width=56,
                height=4,
                relief="flat",
                highlightthickness=0,
            )
            text.tag_configure("title", font=("Segoe UI", 9, "bold"))
            text.tag_configure("body", font=("Segoe UI", 9))
            text.insert("1.0", title, "title")
            text.insert("end", f" {body}", "body")
            text.configure(state="disabled")
            text.pack(anchor="w", padx=8, pady=(8, 4) if not last else (0, 8))

        add_paragraph(
            "Box plot —",
            "a compact summary of a distribution showing the median, interquartile range (IQR) and spread of the data. "
            "The box represents the middle 50% (Q1–Q3), the line is the median, and the whiskers/outliers indicate "
            "variability and extreme values.",
        )
        add_paragraph(
            "Violin plot —",
            "combines a box plot with a smoothed density shape, showing not only summary statistics but also the full distribution "
            "of the data. The width of the violin reflects how frequent values are, making it easier to see multimodal distributions "
            "or skewness.",
            last=True,
        )

        self.violin_help_tracking = True
        if self.violin_window is not None:
            self.violin_window.bind("<Configure>", self._on_violin_window_configure)
        self._position_violin_help()

    def _on_violin_window_configure(self, _event=None) -> None:
        if self.violin_help_tracking:
            self._position_violin_help()

    def _position_violin_help(self) -> None:
        if self.violin_help_window is None or not self.violin_help_window.winfo_exists():
            return
        if self.violin_window is None or not self.violin_window.winfo_exists():
            return
        self.violin_help_window.update_idletasks()
        bubble_w = self.violin_help_window.winfo_width()
        bubble_h = self.violin_help_window.winfo_height()
        win_x = self.violin_window.winfo_rootx()
        win_y = self.violin_window.winfo_rooty()
        x = win_x + 20
        y = win_y - bubble_h - 8
        if y < 0:
            y = win_y + 30
        self.violin_help_window.geometry(f"{bubble_w}x{bubble_h}+{x}+{y}")

    def _refresh_violin_plot(self) -> None:
        if self.violin_window is None or not self.violin_window.winfo_exists():
            return
        if self.violin_ax is None or self.violin_canvas is None or self.violin_property_var is None:
            return

        prop = self.violin_property_var.get().strip().lower()
        groups, labels_out, unit, _basis = self._gmm_groups_for_property(prop)
        self.violin_ax.clear()

        if not groups:
            self.violin_ax.text(0.5, 0.5, "Select particles", ha="center", va="center")
            self.violin_canvas.draw_idle()
            return

        positions = np.arange(1, len(groups) + 1, dtype=float)
        self.violin_ax.violinplot(groups, positions=positions, showmeans=False, showmedians=True, showextrema=False)
        self.violin_ax.boxplot(
            groups,
            positions=positions,
            widths=0.18,
            patch_artist=True,
            boxprops=dict(facecolor="#cfe2f3", edgecolor="#1f77b4"),
            medianprops=dict(color="#d62728"),
            whiskerprops=dict(color="#1f77b4"),
            capprops=dict(color="#1f77b4"),
        )

        self.violin_ax.set_xticks(positions)
        self.violin_ax.set_xticklabels(labels_out)
        self.violin_ax.set_title(f"Violin + Box Plot by GMM groups ({prop})")
        ylabel = f"{prop} ({unit})" if unit else prop
        self.violin_ax.set_ylabel(ylabel)
        self.violin_ax.grid(alpha=0.25, linestyle="--", axis="y")
        self.violin_fig.tight_layout()
        self.violin_canvas.draw_idle()

    def on_save_violin_plot(self) -> None:
        self._save_figure_image(self.violin_fig, "Save Violin + Box Plot", "violin_box_plot.png")

    def open_bivariate_analysis(self) -> None:
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return

        if self.bivariate_window is not None and self.bivariate_window.winfo_exists():
            self.bivariate_window.lift()
            self._refresh_bivariate_plot()
            return

        self.bivariate_window = tk.Toplevel(self.root)
        self.bivariate_window.title("Bivariate Analysis")
        self.bivariate_window.geometry("980x720+900+150")

        outer = ttk.Frame(self.bivariate_window, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        controls = ttk.Frame(outer)
        controls.pack(fill=tk.X, pady=(0, 8))

        options = ("length", "width", "area", "feret diameter", "circularity", "eccentricity", "aspect ratio", "perimeter", "NND")

        ttk.Label(controls, text="X axis:").pack(side=tk.LEFT)
        self.bivariate_x_var = tk.StringVar(value="length")
        x_combo = ttk.Combobox(
            controls,
            textvariable=self.bivariate_x_var,
            state="readonly",
            values=options,
            width=14,
        )
        x_combo.pack(side=tk.LEFT, padx=(8, 14))
        x_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_bivariate_plot())

        ttk.Label(controls, text="Y axis:").pack(side=tk.LEFT)
        self.bivariate_y_var = tk.StringVar(value="aspect ratio")
        y_combo = ttk.Combobox(
            controls,
            textvariable=self.bivariate_y_var,
            state="readonly",
            values=options,
            width=14,
        )
        y_combo.pack(side=tk.LEFT, padx=(8, 14))
        y_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_bivariate_plot())

        ttk.Button(controls, text="Data Fitting", command=lambda: self._open_fit_dialog("bivariate")).pack(
            side=tk.LEFT,
            padx=(8, 10),
        )
        ttk.Button(controls, text="Save Graph", command=self.on_save_bivariate_plot).pack(side=tk.RIGHT)

        self.bivariate_fig = Figure(figsize=(8.8, 6.4), dpi=100)
        self.bivariate_ax = self.bivariate_fig.add_subplot(111)
        self.bivariate_canvas = FigureCanvasTkAgg(self.bivariate_fig, master=outer)
        self.bivariate_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self.bivariate_window.protocol("WM_DELETE_WINDOW", self._close_bivariate_window)
        self._refresh_bivariate_plot()

    def _close_bivariate_window(self) -> None:
        if self.bivariate_window is not None and self.bivariate_window.winfo_exists():
            self.bivariate_window.destroy()
        if self.bivariate_canvas is not None and self.bivariate_pick_cid is not None:
            try:
                self.bivariate_canvas.mpl_disconnect(self.bivariate_pick_cid)
            except Exception:
                pass
        self.bivariate_window = None
        self.bivariate_x_var = None
        self.bivariate_y_var = None
        self.bivariate_fig = None
        self.bivariate_ax = None
        self.bivariate_canvas = None
        self.bivariate_ids = None
        self.bivariate_scatter = None
        self.bivariate_pick_cid = None
        self.bivariate_x_data = None
        self.bivariate_y_data = None
        self.bivariate_fit_line = None
        self.bivariate_fit_text = None
        if self.bivariate_fit_info_window is not None and self.bivariate_fit_info_window.winfo_exists():
            self.bivariate_fit_info_window.destroy()
        self.bivariate_fit_info_window = None

    def _bivariate_values(self, prop: str) -> Tuple[np.ndarray, str]:
        selected = self._selected_particles()
        return self._bivariate_values_for_particles(prop, selected)

    def _bivariate_values_for_particles(self, prop: str, selected: List[ParticleMask]) -> Tuple[np.ndarray, str]:
        if not selected:
            raise ValueError("Select particles")

        nm_per_px = self._safe_um_per_px_for_refresh() * 1000.0
        if prop == "length":
            values = [self._center_length_width_px(p)[0] * nm_per_px for p in selected]
            return np.asarray(values, dtype=float), "nm"
        if prop == "width":
            values = [self._center_length_width_px(p)[1] * nm_per_px for p in selected]
            return np.asarray(values, dtype=float), "nm"
        if prop == "area":
            values = [float(p.area_px) * (nm_per_px ** 2) for p in selected]
            return np.asarray(values, dtype=float), "nm^2"
        if prop == "feret diameter":
            values = [self._feret_diameter_px(p) * nm_per_px for p in selected]
            return np.asarray(values, dtype=float), "nm"
        if prop == "circularity":
            values = [self._circularity_eccentricity(p)[0] for p in selected]
            return np.asarray(values, dtype=float), ""
        if prop == "eccentricity":
            values = [self._circularity_eccentricity(p)[1] for p in selected]
            return np.asarray(values, dtype=float), ""
        if prop == "perimeter":
            values = [self._perimeter_px(p.segmentation) * nm_per_px for p in selected]
            return np.asarray(values, dtype=float), "nm"
        if prop == "aspect ratio":
            ratios = []
            for p in selected:
                length_px, width_px = self._center_length_width_px(p)
                ratios.append(float(length_px / width_px) if width_px > 0 else float("nan"))
            arr = np.asarray(ratios, dtype=float)
            arr[~np.isfinite(arr)] = 0.0
            return arr, ""
        if prop == "NND":
            if not NND_AVAILABLE:
                raise ValueError(f"NND dependencies are missing.\n{NND_IMPORT_ERROR}")
            if len(selected) < 2:
                raise ValueError("Select at least 2 particles")
            centroids_nm = np.asarray(
                [
                    [
                        float(self._particle_centroid_global(p)[0]) * nm_per_px,
                        float(self._particle_centroid_global(p)[1]) * nm_per_px,
                    ]
                    for p in selected
                ],
                dtype=float,
            )
            tree = cKDTree(centroids_nm)
            distances, _ = tree.query(centroids_nm, k=2)
            return np.asarray(distances[:, 1], dtype=float), "nm"

        raise ValueError("Unsupported property")

    def _refresh_bivariate_plot(self) -> None:
        if self.bivariate_window is None or not self.bivariate_window.winfo_exists():
            return
        if self.bivariate_ax is None or self.bivariate_canvas is None:
            return
        if self.bivariate_x_var is None or self.bivariate_y_var is None:
            return

        x_prop = self.bivariate_x_var.get().strip()
        y_prop = self.bivariate_y_var.get().strip()
        selected = self._selected_particles()
        if not selected:
            self.bivariate_ax.clear()
            self.bivariate_ax.text(0.5, 0.5, "Select particles", ha="center", va="center")
            self.bivariate_canvas.draw_idle()
            return

        try:
            x_vals, x_unit = self._bivariate_values_for_particles(x_prop, selected)
            y_vals, y_unit = self._bivariate_values_for_particles(y_prop, selected)
        except Exception as exc:
            messagebox.showerror("Bivariate Analysis", str(exc))
            return

        self.highlighted_ids &= self.selected_ids
        ids = np.asarray([int(p.mask_id) for p in selected], dtype=int)
        self.bivariate_ids = ids

        self.bivariate_ax.clear()
        colors = ["#e53935" if int(pid) in self.highlighted_ids else "#1f77b4" for pid in ids.tolist()]
        self.bivariate_scatter = self.bivariate_ax.scatter(
            x_vals,
            y_vals,
            s=32,
            c=colors,
            alpha=0.85,
            edgecolors="white",
            linewidths=0.5,
            picker=True,
            pickradius=6,
        )
        xlabel = f"{x_prop} ({x_unit})" if x_unit else x_prop
        ylabel = f"{y_prop} ({y_unit})" if y_unit else y_prop
        self.bivariate_ax.set_xlabel(xlabel)
        self.bivariate_ax.set_ylabel(ylabel)
        self.bivariate_ax.set_title("Bivariate Analysis")
        self.bivariate_ax.grid(alpha=0.25, linestyle="--")
        self.bivariate_fig.tight_layout()
        self.bivariate_canvas.draw_idle()

        self.bivariate_x_data = np.asarray(x_vals, dtype=float)
        self.bivariate_y_data = np.asarray(y_vals, dtype=float)

        if self.bivariate_pick_cid is None:
            self.bivariate_pick_cid = self.bivariate_canvas.mpl_connect("pick_event", self._on_bivariate_pick)

    def on_save_bivariate_plot(self) -> None:
        self._save_figure_image(self.bivariate_fig, "Save Bivariate Plot", "bivariate_plot.png")

    def open_bivariate_gmm_analysis(self) -> None:
        if not GMM_AVAILABLE:
            messagebox.showerror("Bivariate + GMM Error", f"GMM dependencies are missing.\n{GMM_IMPORT_ERROR}")
            return
        if not self.selected_ids:
            messagebox.showerror("Error", "Select particles")
            return

        if self.bivariate_gmm_window is not None and self.bivariate_gmm_window.winfo_exists():
            self.bivariate_gmm_window.lift()
            self._refresh_bivariate_gmm_plot()
            return

        self.bivariate_gmm_window = tk.Toplevel(self.root)
        self.bivariate_gmm_window.title("Bivariate Analysis + GMM")
        self.bivariate_gmm_window.geometry("980x720+930+150")

        outer = ttk.Frame(self.bivariate_gmm_window, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        controls = ttk.Frame(outer)
        controls.pack(fill=tk.X, pady=(0, 8))

        options = ("length", "width", "area", "feret diameter", "circularity", "eccentricity", "aspect ratio", "perimeter", "NND")

        ttk.Label(controls, text="X axis:").pack(side=tk.LEFT)
        self.bivariate_gmm_x_var = tk.StringVar(value="length")
        x_combo = ttk.Combobox(
            controls,
            textvariable=self.bivariate_gmm_x_var,
            state="readonly",
            values=options,
            width=14,
        )
        x_combo.pack(side=tk.LEFT, padx=(8, 14))
        x_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_bivariate_gmm_plot())

        ttk.Label(controls, text="Y axis:").pack(side=tk.LEFT)
        self.bivariate_gmm_y_var = tk.StringVar(value="aspect ratio")
        y_combo = ttk.Combobox(
            controls,
            textvariable=self.bivariate_gmm_y_var,
            state="readonly",
            values=options,
            width=14,
        )
        y_combo.pack(side=tk.LEFT, padx=(8, 14))
        y_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_bivariate_gmm_plot())

        ttk.Label(controls, text="Point distinction:").pack(side=tk.LEFT)
        self.bivariate_gmm_distinction_var = tk.StringVar(value="shape")
        d_combo = ttk.Combobox(
            controls,
            textvariable=self.bivariate_gmm_distinction_var,
            state="readonly",
            values=("shape", "color"),
            width=10,
        )
        d_combo.pack(side=tk.LEFT, padx=(8, 14))
        d_combo.bind("<<ComboboxSelected>>", lambda _evt: self._refresh_bivariate_gmm_plot())

        ttk.Button(controls, text="Data Fitting", command=lambda: self._open_fit_dialog("bivariate_gmm")).pack(
            side=tk.LEFT,
            padx=(8, 10),
        )
        ttk.Button(controls, text="Save Graph", command=self.on_save_bivariate_gmm_plot).pack(side=tk.RIGHT)

        self.bivariate_gmm_fig = Figure(figsize=(8.8, 6.4), dpi=100)
        self.bivariate_gmm_ax = self.bivariate_gmm_fig.add_subplot(111)
        self.bivariate_gmm_canvas = FigureCanvasTkAgg(self.bivariate_gmm_fig, master=outer)
        self.bivariate_gmm_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self.bivariate_gmm_window.protocol("WM_DELETE_WINDOW", self._close_bivariate_gmm_window)
        self._refresh_bivariate_gmm_plot()

    def _close_bivariate_gmm_window(self) -> None:
        if self.bivariate_gmm_window is not None and self.bivariate_gmm_window.winfo_exists():
            self.bivariate_gmm_window.destroy()
        if self.bivariate_gmm_canvas is not None and self.bivariate_gmm_pick_cid is not None:
            try:
                self.bivariate_gmm_canvas.mpl_disconnect(self.bivariate_gmm_pick_cid)
            except Exception:
                pass
        self.bivariate_gmm_window = None
        self.bivariate_gmm_x_var = None
        self.bivariate_gmm_y_var = None
        self.bivariate_gmm_distinction_var = None
        self.bivariate_gmm_fig = None
        self.bivariate_gmm_ax = None
        self.bivariate_gmm_canvas = None
        self.bivariate_gmm_scatter_groups = []
        self.bivariate_gmm_pick_cid = None
        self.bivariate_gmm_x_data = None
        self.bivariate_gmm_y_data = None
        self.bivariate_gmm_fit_line = None
        self.bivariate_gmm_fit_text = None
        if self.bivariate_gmm_fit_info_window is not None and self.bivariate_gmm_fit_info_window.winfo_exists():
            self.bivariate_gmm_fit_info_window.destroy()
        self.bivariate_gmm_fit_info_window = None

    def open_overlay_tool(self) -> None:
        if self.image_rgb is None:
            messagebox.showerror("Error", "Import an image first.")
            return

        if self.overlay_prompt_window is not None and self.overlay_prompt_window.winfo_exists():
            self.overlay_prompt_window.lift()
            return

        self.overlay_prompt_window = tk.Toplevel(self.root)
        self.overlay_prompt_window.title("Overlay")
        self.overlay_prompt_window.geometry("320x160+940+260")

        outer = ttk.Frame(self.overlay_prompt_window, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        ttk.Label(outer, text="open another image", font=("Segoe UI", 11)).pack(anchor="center", pady=(6, 14))

        btn_row = ttk.Frame(outer)
        btn_row.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(btn_row, text="Ok", command=self._overlay_pick_image).pack(side=tk.LEFT, expand=True, padx=6)
        ttk.Button(btn_row, text="Cancel", command=self._close_overlay_prompt).pack(side=tk.LEFT, expand=True, padx=6)
        ttk.Button(btn_row, text="Remove Overlay", command=self._remove_overlay).pack(
            side=tk.LEFT,
            expand=True,
            padx=6,
        )

        self.overlay_prompt_window.protocol("WM_DELETE_WINDOW", self._close_overlay_prompt)

    def _close_overlay_prompt(self) -> None:
        if self.overlay_prompt_window is not None and self.overlay_prompt_window.winfo_exists():
            self.overlay_prompt_window.destroy()
        self.overlay_prompt_window = None

    def _overlay_pick_image(self) -> None:
        self._close_overlay_prompt()
        path = filedialog.askopenfilename(
            title="Select overlay image",
            filetypes=[
                ("Image files", "*.jpg *.jpeg *.png *.tif *.tiff"),
                ("JPEG", "*.jpg *.jpeg"),
                ("PNG", "*.png"),
                ("TIFF", "*.tif *.tiff"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return

        try:
            overlay_img = Image.open(path).convert("RGB")
        except Exception as exc:
            messagebox.showerror("Overlay error", f"Could not load overlay image:\n{exc}")
            return

        self._open_overlay_opacity_dialog(overlay_img)

    def _open_overlay_opacity_dialog(self, overlay_img: Image.Image) -> None:
        if self.overlay_opacity_window is not None and self.overlay_opacity_window.winfo_exists():
            self.overlay_opacity_window.destroy()

        self.overlay_opacity_window = tk.Toplevel(self.root)
        self.overlay_opacity_window.title("Overlay Opacity")
        self.overlay_opacity_window.geometry("260x170+960+300")

        outer = ttk.Frame(self.overlay_opacity_window, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        ttk.Label(outer, text="Opacity", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 6))

        opacity_var = tk.StringVar(value="100")
        opacity_entry = ttk.Entry(outer, textvariable=opacity_var, width=10)
        opacity_entry.pack(anchor="w", pady=(0, 10))
        opacity_entry.focus_set()

        def apply_opacity() -> None:
            try:
                value = float(opacity_var.get().strip())
            except Exception:
                value = 100.0
            value = max(0.0, min(100.0, value))
            self._apply_overlay(overlay_img, value)
            if self.overlay_opacity_window is not None and self.overlay_opacity_window.winfo_exists():
                self.overlay_opacity_window.destroy()
            self.overlay_opacity_window = None

        ttk.Button(outer, text="Ok", command=apply_opacity).pack(anchor="center", pady=(6, 0))
        self.overlay_opacity_window.protocol("WM_DELETE_WINDOW", lambda: self._close_overlay_opacity_window())

    def _close_overlay_opacity_window(self) -> None:
        if self.overlay_opacity_window is not None and self.overlay_opacity_window.winfo_exists():
            self.overlay_opacity_window.destroy()
        self.overlay_opacity_window = None

    def _apply_overlay(self, overlay_img: Image.Image, opacity_percent: float) -> None:
        if self.image_rgb is None:
            messagebox.showerror("Overlay error", "Import an image first.")
            return
        if overlay_img is None:
            return

        if self.overlay_base_image is None:
            self.overlay_base_image = self.image_rgb.copy()
        base = self.overlay_base_image
        h, w = base.shape[:2]
        if overlay_img.size != (w, h):
            overlay_img = overlay_img.resize((w, h), Image.Resampling.BILINEAR)

        overlay_np = np.asarray(overlay_img, dtype=np.float32)
        base_np = base.astype(np.float32)
        alpha = max(0.0, min(1.0, opacity_percent / 100.0))
        blended = (base_np * (1.0 - alpha)) + (overlay_np * alpha)
        self.image_rgb = np.clip(blended, 0, 255).astype(np.uint8)

        self.render_image(keep_view=True)
        self.status_var.set(f"Overlay applied ({opacity_percent:.0f}%)")

    def _remove_overlay(self) -> None:
        if self.overlay_base_image is None:
            messagebox.showerror("Overlay", "No Overlay Selected")
            return
        self.image_rgb = self.overlay_base_image.copy()
        self.overlay_base_image = None
        self.render_image(keep_view=True)
        self.status_var.set("Overlay removed.")

    def _fit_best_gmm_2d(self, x_vals: np.ndarray, y_vals: np.ndarray) -> Tuple[Optional[GaussianMixture], Optional[np.ndarray]]:
        data = np.column_stack((x_vals, y_vals)).astype(float)
        n_samples = data.shape[0]
        if n_samples == 0:
            return None, None
        max_components = min(6, n_samples)
        best_model: Optional[GaussianMixture] = None
        best_bic = float("inf")
        for n_comp in range(1, max_components + 1):
            try:
                gmm = GaussianMixture(
                    n_components=n_comp,
                    covariance_type="full",
                    random_state=0,
                    n_init=5,
                )
                gmm.fit(data)
                bic = float(gmm.bic(data))
                if bic < best_bic:
                    best_bic = bic
                    best_model = gmm
            except Exception:
                continue
        if best_model is None:
            return None, None
        labels = best_model.predict(data)
        return best_model, labels

    def _refresh_bivariate_gmm_plot(self) -> None:
        if self.bivariate_gmm_window is None or not self.bivariate_gmm_window.winfo_exists():
            return
        if self.bivariate_gmm_ax is None or self.bivariate_gmm_canvas is None:
            return
        if self.bivariate_gmm_x_var is None or self.bivariate_gmm_y_var is None:
            return
        if self.bivariate_gmm_distinction_var is None:
            return

        selected = self._selected_particles()
        if not selected:
            self.bivariate_gmm_ax.clear()
            self.bivariate_gmm_ax.text(0.5, 0.5, "Select particles", ha="center", va="center")
            self.bivariate_gmm_canvas.draw_idle()
            return
        self.highlighted_ids &= self.selected_ids

        x_prop = self.bivariate_gmm_x_var.get().strip()
        y_prop = self.bivariate_gmm_y_var.get().strip()
        try:
            x_vals, x_unit = self._bivariate_values_for_particles(x_prop, selected)
            y_vals, y_unit = self._bivariate_values_for_particles(y_prop, selected)
        except Exception as exc:
            messagebox.showerror("Bivariate Analysis + GMM", str(exc))
            return

        ids = np.asarray([int(p.mask_id) for p in selected], dtype=int)
        _model, labels = self._fit_best_gmm_2d(x_vals, y_vals)
        if labels is None:
            labels = np.zeros_like(x_vals, dtype=int)

        groups = sorted(int(lbl) for lbl in set(labels.tolist()))
        if not groups:
            groups = [0]
            labels = np.zeros_like(x_vals, dtype=int)

        markers = ["o", "^", "s", "D", "P", "X", "v", ">", "<"]
        colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2", "#7f7f7f", "#17becf"]
        distinction = self.bivariate_gmm_distinction_var.get().strip().lower()

        self.bivariate_gmm_ax.clear()
        self.bivariate_gmm_scatter_groups = []

        for idx, group_id in enumerate(groups):
            mask = labels == group_id
            if distinction == "color":
                color = colors[idx % len(colors)]
                marker = "o"
            else:
                color = "#1f77b4"
                marker = markers[idx % len(markers)]
            label = f"G{idx + 1}"
            scatter = self.bivariate_gmm_ax.scatter(
                x_vals[mask],
                y_vals[mask],
                s=34,
                marker=marker,
                color=color,
                alpha=0.8,
                edgecolors="white",
                linewidths=0.5,
                label=label,
                picker=True,
                pickradius=6,
            )
            ids_group = ids[mask]
            self.bivariate_gmm_scatter_groups.append(
                {
                    "scatter": scatter,
                    "ids": ids_group,
                    "base_color": color,
                }
            )

        xlabel = f"{x_prop} ({x_unit})" if x_unit else x_prop
        ylabel = f"{y_prop} ({y_unit})" if y_unit else y_prop
        self.bivariate_gmm_ax.set_xlabel(xlabel)
        self.bivariate_gmm_ax.set_ylabel(ylabel)
        self.bivariate_gmm_ax.set_title("Bivariate Analysis + GMM")
        self.bivariate_gmm_ax.grid(alpha=0.25, linestyle="--")
        self.bivariate_gmm_ax.legend(loc="center left", bbox_to_anchor=(1.02, 0.5), frameon=False)
        self.bivariate_gmm_fig.tight_layout()
        self.bivariate_gmm_canvas.draw_idle()

        self.bivariate_gmm_x_data = np.asarray(x_vals, dtype=float)
        self.bivariate_gmm_y_data = np.asarray(y_vals, dtype=float)

        if self.bivariate_gmm_pick_cid is None:
            self.bivariate_gmm_pick_cid = self.bivariate_gmm_canvas.mpl_connect(
                "pick_event",
                self._on_bivariate_gmm_pick,
            )

    def on_save_bivariate_gmm_plot(self) -> None:
        self._save_figure_image(self.bivariate_gmm_fig, "Save Bivariate GMM Plot", "bivariate_gmm_plot.png")

    def _on_bivariate_gmm_pick(self, event) -> None:
        if not self.bivariate_gmm_scatter_groups:
            return
        hit_group = None
        for group in self.bivariate_gmm_scatter_groups:
            if event.artist == group.get("scatter"):
                hit_group = group
                break
        if hit_group is None:
            return
        if not hasattr(event, "ind") or len(event.ind) == 0:
            return
        idx = int(event.ind[0])
        ids_group = hit_group.get("ids")
        if ids_group is None or idx < 0 or idx >= len(ids_group):
            return
        pid = int(ids_group[idx])

        if pid in self.highlighted_ids:
            self.highlighted_ids.remove(pid)
            self.status_var.set(f"Bivariate GMM: unhighlighted #{pid}")
        else:
            self.highlighted_ids.add(pid)
            self.status_var.set(f"Bivariate GMM: highlighted #{pid}")

        for group in self.bivariate_gmm_scatter_groups:
            scatter = group.get("scatter")
            ids_group = group.get("ids")
            base_color = group.get("base_color")
            if scatter is None or ids_group is None:
                continue
            colors = [
                "#e53935" if int(p) in self.highlighted_ids else base_color
                for p in ids_group.tolist()
            ]
            scatter.set_facecolors(colors)

        if self.bivariate_gmm_canvas is not None:
            self.bivariate_gmm_canvas.draw_idle()
        self.render_image(keep_view=True)

    def _on_bivariate_pick(self, event) -> None:
        if self.bivariate_scatter is None or self.bivariate_ids is None:
            return
        if event.artist != self.bivariate_scatter:
            return
        if not hasattr(event, "ind") or len(event.ind) == 0:
            return
        idx = int(event.ind[0])
        if idx < 0 or idx >= len(self.bivariate_ids):
            return

        pid = int(self.bivariate_ids[idx])
        if pid in self.highlighted_ids:
            self.highlighted_ids.remove(pid)
            self.status_var.set(f"Bivariate: unhighlighted #{pid}")
        else:
            self.highlighted_ids.add(pid)
            self.status_var.set(f"Bivariate: highlighted #{pid}")

        if self.bivariate_scatter is not None:
            colors = [
                "#e53935" if int(p) in self.highlighted_ids else "#1f77b4"
                for p in self.bivariate_ids.tolist()
            ]
            self.bivariate_scatter.set_facecolors(colors)
            self.bivariate_canvas.draw_idle()

        self.render_image(keep_view=True)

    def _open_fit_dialog(self, context: str) -> None:
        if not FIT_AVAILABLE:
            messagebox.showerror("Fitting Error", f"Fitting dependencies are missing.\n{FIT_IMPORT_ERROR}")
            return

        fit_window = tk.Toplevel(self.root)
        fit_window.title("Data Fitting")
        fit_window.geometry("360x230+980+220")

        outer = ttk.Frame(fit_window, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        ttk.Label(outer, text="Fit type:").pack(anchor="w")
        fit_type_var = tk.StringVar(value="linear")
        fit_combo = ttk.Combobox(
            outer,
            textvariable=fit_type_var,
            state="readonly",
            values=("linear", "exponential", "logarithmic", "polinomic", "sigmoidal"),
            width=18,
        )
        fit_combo.pack(anchor="w", pady=(4, 8))

        degree_frame = ttk.Frame(outer)
        degree_frame.pack(anchor="w", pady=(0, 8))
        ttk.Label(degree_frame, text="Polynomial degree:").pack(side=tk.LEFT)
        degree_var = tk.StringVar(value="2")
        degree_spin = ttk.Spinbox(degree_frame, from_=2, to=8, increment=1, textvariable=degree_var, width=5)
        degree_spin.pack(side=tk.LEFT, padx=(6, 0))

        def toggle_degree(*_args) -> None:
            if fit_type_var.get().strip() == "polinomic":
                degree_frame.pack(anchor="w", pady=(0, 8))
            else:
                degree_frame.pack_forget()

        fit_combo.bind("<<ComboboxSelected>>", toggle_degree)
        toggle_degree()

        ttk.Button(outer, text="Fit", command=lambda: self._apply_fit(context, fit_type_var, degree_var)).pack(
            anchor="e",
            pady=(12, 0),
        )

    def _apply_fit(self, context: str, fit_type_var: tk.StringVar, degree_var: tk.StringVar) -> None:
        if context == "bivariate":
            x = self.bivariate_x_data
            y = self.bivariate_y_data
            ax = self.bivariate_ax
            fig = self.bivariate_fig
            canvas = self.bivariate_canvas
            line_attr = "bivariate_fit_line"
            text_attr = "bivariate_fit_text"
            info_attr = "bivariate_fit_info_window"
        else:
            x = self.bivariate_gmm_x_data
            y = self.bivariate_gmm_y_data
            ax = self.bivariate_gmm_ax
            fig = self.bivariate_gmm_fig
            canvas = self.bivariate_gmm_canvas
            line_attr = "bivariate_gmm_fit_line"
            text_attr = "bivariate_gmm_fit_text"
            info_attr = "bivariate_gmm_fit_info_window"

        if x is None or y is None or ax is None or fig is None or canvas is None:
            messagebox.showerror("Data Fitting", "Plot data is not available.")
            return

        x = np.asarray(x, dtype=float)
        y = np.asarray(y, dtype=float)
        finite = np.isfinite(x) & np.isfinite(y)
        x = x[finite]
        y = y[finite]
        if x.size < 3:
            messagebox.showerror("Data Fitting", "Not enough data points.")
            return

        fit_type = fit_type_var.get().strip()
        degree = 2
        if fit_type == "polinomic":
            try:
                degree = int(degree_var.get().strip())
            except Exception:
                degree = 2
            degree = max(1, min(8, degree))

        try:
            y_fit, eqn, r2 = self._fit_model(x, y, fit_type, degree)
        except Exception as exc:
            messagebox.showerror("Data Fitting", str(exc))
            return

        order = np.argsort(x)
        x_sorted = x[order]
        y_sorted = y_fit[order]

        existing_line = getattr(self, line_attr, None)
        if existing_line is not None:
            try:
                existing_line.remove()
            except Exception:
                pass
        line = ax.plot(x_sorted, y_sorted, color="#ff7f0e", linewidth=2.0, label="Fit")[0]
        setattr(self, line_attr, line)

        existing_text = getattr(self, text_attr, None)
        if existing_text is not None:
            try:
                existing_text.remove()
            except Exception:
                pass
        text = ax.text(
            0.02,
            0.98,
            f"{eqn}\nR² = {r2:.4f}",
            transform=ax.transAxes,
            va="top",
            ha="left",
            fontsize=9,
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.75, edgecolor="#cccccc"),
        )
        setattr(self, text_attr, text)

        canvas.draw_idle()

        info_window = getattr(self, info_attr, None)
        if info_window is not None and info_window.winfo_exists():
            info_window.destroy()
        info_window = tk.Toplevel(self.root)
        info_window.title("Fit Results")
        info_window.geometry("360x200+1020+260")
        info_frame = ttk.Frame(info_window, padding=10)
        info_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(info_frame, text="Fit equation:", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        ttk.Label(info_frame, text=eqn, wraplength=320).pack(anchor="w", pady=(2, 8))
        ttk.Label(info_frame, text=f"R² = {r2:.6f}", font=("Segoe UI", 10)).pack(anchor="w")
        setattr(self, info_attr, info_window)

    def _fit_model(self, x: np.ndarray, y: np.ndarray, fit_type: str, degree: int) -> Tuple[np.ndarray, str, float]:
        if fit_type == "linear":
            coeffs = np.polyfit(x, y, 1)
            y_fit = np.polyval(coeffs, x)
            eqn = f"y = {coeffs[0]:.6f} x + {coeffs[1]:.6f}"
        elif fit_type == "polinomic":
            coeffs = np.polyfit(x, y, degree)
            y_fit = np.polyval(coeffs, x)
            terms = []
            for i, c in enumerate(coeffs):
                power = degree - i
                if power == 0:
                    terms.append(f"{c:.6f}")
                elif power == 1:
                    terms.append(f"{c:.6f} x")
                else:
                    terms.append(f"{c:.6f} x^{power}")
            eqn = "y = " + " + ".join(terms)
        elif fit_type == "exponential":
            def exp_fn(xv, a, b, c):
                return a * np.exp(b * xv) + c

            p0 = (1.0, 0.01, np.min(y))
            params, _ = curve_fit(exp_fn, x, y, p0=p0, maxfev=5000)
            y_fit = exp_fn(x, *params)
            eqn = f"y = {params[0]:.6f} * exp({params[1]:.6f} x) + {params[2]:.6f}"
        elif fit_type == "logarithmic":
            if np.any(x <= 0):
                raise ValueError("Logarithmic fit requires x > 0.")

            def log_fn(xv, a, b):
                return a * np.log(xv) + b

            p0 = (1.0, np.mean(y))
            params, _ = curve_fit(log_fn, x, y, p0=p0, maxfev=5000)
            y_fit = log_fn(x, *params)
            eqn = f"y = {params[0]:.6f} * ln(x) + {params[1]:.6f}"
        elif fit_type == "sigmoidal":
            def sig_fn(xv, L, k, x0, b):
                return L / (1.0 + np.exp(-k * (xv - x0))) + b

            p0 = (np.max(y) - np.min(y), 0.1, np.median(x), np.min(y))
            params, _ = curve_fit(sig_fn, x, y, p0=p0, maxfev=8000)
            y_fit = sig_fn(x, *params)
            eqn = f"y = {params[0]:.6f} / (1 + exp(-{params[1]:.6f}(x - {params[2]:.6f}))) + {params[3]:.6f}"
        else:
            raise ValueError("Unsupported fit type.")

        ss_res = float(np.sum((y - y_fit) ** 2))
        ss_tot = float(np.sum((y - np.mean(y)) ** 2))
        r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")
        return y_fit, eqn, r2

    def open_color_settings(self) -> None:
        if self.color_settings_window is not None and self.color_settings_window.winfo_exists():
            self.color_settings_window.lift()
            return

        self.color_settings_window = tk.Toplevel(self.root)
        self.color_settings_window.title("Color Settings")
        self.color_settings_window.geometry("360x430+920+220")

        outer = ttk.Frame(self.color_settings_window, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        ttk.Label(outer, text="Selection Color", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))

        if self.color_wheel_img is None:
            self.color_wheel_img = self._generate_color_wheel(220)
        self.color_wheel_tk = ImageTk.PhotoImage(self.color_wheel_img)

        self.color_wheel_canvas = tk.Canvas(outer, width=220, height=220, highlightthickness=1, highlightbackground="#cccccc")
        self.color_wheel_canvas.create_image(0, 0, anchor="nw", image=self.color_wheel_tk)
        self.color_wheel_canvas.pack(anchor="w", pady=(0, 10))
        self.color_wheel_canvas.bind("<Button-1>", self._on_color_wheel_click)
        self.color_wheel_canvas.bind("<B1-Motion>", self._on_color_wheel_click)

        rgb_frame = ttk.Frame(outer)
        rgb_frame.pack(fill=tk.X, pady=(4, 8))

        self.color_r_var = tk.StringVar(value=str(self.selection_color_rgb[0]))
        self.color_g_var = tk.StringVar(value=str(self.selection_color_rgb[1]))
        self.color_b_var = tk.StringVar(value=str(self.selection_color_rgb[2]))

        def rgb_row(label: str, var: tk.StringVar, color: str) -> None:
            row = ttk.Frame(rgb_frame)
            row.pack(fill=tk.X, pady=2)
            ttk.Label(row, text=label, width=3).pack(side=tk.LEFT)
            ttk.Entry(row, textvariable=var, width=8).pack(side=tk.LEFT, padx=(4, 6))
            swatch = tk.Canvas(row, width=22, height=14, highlightthickness=1, highlightbackground="#999999")
            swatch.create_rectangle(0, 0, 22, 14, fill=color, outline="")
            swatch.pack(side=tk.LEFT)

        rgb_row("R", self.color_r_var, "#ff0000")
        rgb_row("G", self.color_g_var, "#00ff00")
        rgb_row("B", self.color_b_var, "#0000ff")

        preview_row = ttk.Frame(outer)
        preview_row.pack(anchor="w", pady=(8, 0))
        ttk.Label(preview_row, text="Preview:").pack(side=tk.LEFT)
        self.color_preview_canvas = tk.Canvas(preview_row, width=40, height=18, highlightthickness=1, highlightbackground="#999999")
        self.color_preview_canvas.pack(side=tk.LEFT, padx=(6, 0))
        self._update_color_preview()

        for var in (self.color_r_var, self.color_g_var, self.color_b_var):
            var.trace_add("write", lambda *_: self._apply_rgb_entries())

        self.color_settings_window.protocol("WM_DELETE_WINDOW", self._close_color_settings)

    def _close_color_settings(self) -> None:
        if self.color_settings_window is not None and self.color_settings_window.winfo_exists():
            self.color_settings_window.destroy()
        self.color_settings_window = None
        self.color_wheel_tk = None
        self.color_wheel_canvas = None
        self.color_r_var = None
        self.color_g_var = None
        self.color_b_var = None
        self.color_preview_canvas = None

    def _generate_color_wheel(self, size: int = 220) -> Image.Image:
        img = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        px = img.load()
        radius = size / 2.0
        cx = cy = radius - 0.5
        for y in range(size):
            for x in range(size):
                dx = x - cx
                dy = y - cy
                r = math.hypot(dx, dy)
                if r > radius:
                    continue
                sat = r / radius
                hue = (math.degrees(math.atan2(dy, dx)) + 360.0) % 360.0
                h = hue / 360.0
                r_rgb, g_rgb, b_rgb = colorsys.hsv_to_rgb(h, sat, 1.0)
                px[x, y] = (int(r_rgb * 255), int(g_rgb * 255), int(b_rgb * 255), 255)
        return img

    def _on_color_wheel_click(self, event) -> None:
        if self.color_wheel_img is None:
            return
        x = int(event.x)
        y = int(event.y)
        if x < 0 or y < 0 or x >= self.color_wheel_img.width or y >= self.color_wheel_img.height:
            return
        r, g, b, a = self.color_wheel_img.getpixel((x, y))
        if a == 0:
            return
        self._set_selection_color((int(r), int(g), int(b)))

    def _apply_rgb_entries(self) -> None:
        if self.color_r_var is None or self.color_g_var is None or self.color_b_var is None:
            return
        try:
            r = int(self.color_r_var.get())
            g = int(self.color_g_var.get())
            b = int(self.color_b_var.get())
        except Exception:
            return
        r = max(0, min(255, r))
        g = max(0, min(255, g))
        b = max(0, min(255, b))
        self._set_selection_color((r, g, b), update_entries=False)

    def _set_selection_color(self, rgb: Tuple[int, int, int], update_entries: bool = True) -> None:
        self.selection_color_rgb = rgb
        if update_entries and self.color_r_var is not None:
            self.color_r_var.set(str(rgb[0]))
            self.color_g_var.set(str(rgb[1]))
            self.color_b_var.set(str(rgb[2]))
        self._update_color_preview()
        self.render_image(keep_view=True)

    def _update_color_preview(self) -> None:
        if self.color_preview_canvas is None:
            return
        r, g, b = self.selection_color_rgb
        self.color_preview_canvas.delete("all")
        self.color_preview_canvas.create_rectangle(0, 0, 40, 18, fill=f"#{r:02x}{g:02x}{b:02x}", outline="")

    def on_select_checkpoint(self) -> None:
        path = filedialog.askopenfilename(
            title="Select SAM checkpoint (.pth)",
            filetypes=[("PyTorch checkpoint", "*.pth"), ("All files", "*.*")],
        )
        if path:
            self.checkpoint_var.set(path)

    def on_import_image(self) -> None:
        if self.sam_running:
            messagebox.showwarning("SAM running", "Wait for current SAM run to finish.")
            return
        paths = filedialog.askopenfilenames(
            title="Import image",
            filetypes=[
                ("Image files", "*.jpg *.jpeg *.png *.tif *.tiff"),
                ("JPEG", "*.jpg *.jpeg"),
                ("PNG", "*.png"),
                ("TIFF", "*.tif *.tiff"),
                ("All files", "*.*"),
            ],
        )
        if not paths:
            return
        paths = list(paths)
        if len(paths) > 4:
            messagebox.showwarning("Limit reached", "Select up to 4 images. The first 4 will be used.")
            paths = paths[:4]

        for pth in paths:
            if not pth.lower().endswith(SUPPORTED_EXT):
                messagebox.showerror("Unsupported format", "Use jpg, png, tif, or tiff.")
                return

        images: List[np.ndarray] = []
        for pth in paths:
            try:
                images.append(load_rgb_image(pth))
            except Exception as exc:
                messagebox.showerror("Load error", f"Could not load image:\n{exc}")
                return

        self.image_paths = paths
        self.image_rgbs = images
        self.particles = []
        self.particles_by_image = {}
        self.selected_ids.clear()
        self.highlighted_ids.clear()
        self.mask_pick_map = None
        self.measure_line = None
        self.measure_label = None
        self.view_mode = min(len(self.image_rgbs), 4) if self.image_rgbs else 1
        self._update_view_mode_buttons()
        self._compose_active_images()
        self.status_var.set(f"Loaded: {len(self.image_rgbs)} image(s). Run SAM.")
        self.set_mode("none")
        self.render_image()
        self.refresh_table()

    def _ask_save_image_mode(self) -> Optional[str]:
        if len(self.image_rgbs) <= 1:
            return "collage"

        dialog = tk.Toplevel(self.root)
        dialog.title("Save Image")
        dialog.geometry("360x160+920+240")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        outer = ttk.Frame(dialog, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)
        ttk.Label(outer, text="Save images separately or as a collage?", font=("Segoe UI", 10)).pack(
            anchor="center",
            pady=(6, 14),
        )

        choice: Dict[str, Optional[str]] = {"mode": None}

        def set_choice(val: str) -> None:
            choice["mode"] = val
            dialog.destroy()

        btn_row = ttk.Frame(outer)
        btn_row.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(btn_row, text="Separate", command=lambda: set_choice("separate")).pack(
            side=tk.LEFT,
            expand=True,
            padx=6,
        )
        ttk.Button(btn_row, text="Collage", command=lambda: set_choice("collage")).pack(
            side=tk.LEFT,
            expand=True,
            padx=6,
        )
        ttk.Button(btn_row, text="Cancel", command=lambda: set_choice("cancel")).pack(
            side=tk.LEFT,
            expand=True,
            padx=6,
        )

        dialog.protocol("WM_DELETE_WINDOW", lambda: set_choice("cancel"))
        self.root.wait_window(dialog)
        if choice["mode"] == "cancel":
            return None
        return choice["mode"]

    def on_save_image(self) -> None:
        if not self.image_rgbs or self.image_rgb is None:
            messagebox.showerror("Save Image", "Import an image first.")
            return

        mode = self._ask_save_image_mode()
        if mode is None:
            return

        save_path = filedialog.asksaveasfilename(
            title="Save Image",
            defaultextension=".png",
            initialfile="nanosegment_image.png",
            filetypes=[
                ("PNG", "*.png"),
                ("JPEG", "*.jpg"),
                ("TIFF", "*.tif *.tiff"),
            ],
        )
        if not save_path:
            return

        ext = os.path.splitext(save_path)[1].lower()
        if ext not in (".png", ".jpg", ".jpeg", ".tif", ".tiff"):
            save_path = f"{save_path}.png"
            ext = ".png"

        try:
            if mode == "separate" and len(self.image_rgbs) > 1:
                base = os.path.splitext(os.path.basename(save_path))[0]
                out_dir = os.path.dirname(save_path)
                for idx, img in enumerate(self.image_rgbs, start=1):
                    out_name = f"{base}_{idx}{ext}"
                    out_path = os.path.join(out_dir, out_name)
                    Image.fromarray(img).save(out_path)
                self.status_var.set(f"Saved {len(self.image_rgbs)} images to {out_dir}")
            else:
                Image.fromarray(self.image_rgb).save(save_path)
                self.status_var.set(f"Image saved: {os.path.basename(save_path)}")
        except Exception as exc:
            messagebox.showerror("Save error", f"Could not save image:\n{exc}")

    def on_autocalibration(self) -> None:
        if self.image_rgb is None:
            messagebox.showerror("Autocalibration", "Import an image first.")
            return
        if not CV2_AVAILABLE:
            messagebox.showerror("Autocalibration", f"OpenCV is required.\n{CV2_IMPORT_ERROR}")
            return
        messagebox.showinfo(
            "Autocalibration",
            "Draw a rectangle around the scale bar to continue.",
        )
        # Always start with manual rectangle selection for better scale bar localization.
        self._start_manual_scale_bar_selection()

    def _start_manual_scale_bar_selection(self) -> None:
        self.autocalibration_pending = True
        self.autocalibration_prev_mode = self.mode
        self.set_mode("scale_bar_rect")
        self.status_var.set("Autocalibration: draw a rectangle around the scale bar.")

    def _manual_scale_bar_selected(self, x_min: float, y_min: float, x_max: float, y_max: float) -> None:
        if not self.autocalibration_pending:
            return
        self.autocalibration_pending = False
        prev_mode = self.autocalibration_prev_mode or "none"
        self.set_mode(prev_mode)

        if self.image_rgb is None:
            messagebox.showerror("Autocalibration", "Import an image first.")
            return

        h, w = self.image_rgb.shape[:2]
        x0 = max(0, min(w - 1, int(round(min(x_min, x_max)))))
        x1 = max(0, min(w, int(round(max(x_min, x_max)))))
        y0 = max(0, min(h - 1, int(round(min(y_min, y_max)))))
        y1 = max(0, min(h, int(round(max(y_min, y_max)))))

        if x1 <= x0 + 2 or y1 <= y0 + 2:
            messagebox.showerror("Autocalibration", "Selection too small for scale bar detection.")
            return

        roi = self.image_rgb[y0:y1, x0:x1]
        try:
            result = self._detect_scale_bar(roi, allow_any_location=True)
        except Exception as exc:
            messagebox.showerror("Autocalibration", f"Scale bar detection failed:\n{exc}")
            return

        if result is None:
            messagebox.showerror("Autocalibration", "Could not detect a scale bar in the selected area.")
            return

        bar_len_px, _bar_bbox = result
        self._finalize_autocalibration(bar_len_px)

    def _finalize_autocalibration(self, bar_len_px: int) -> None:
        if bar_len_px <= 0:
            messagebox.showerror("Autocalibration", "Invalid scale bar length.")
            return
        self._prompt_manual_scale_value_nm(bar_len_px)

    def _prompt_manual_scale_value_nm(self, bar_len_px: int) -> None:
        if bar_len_px <= 0:
            messagebox.showerror("Autocalibration", "Invalid scale bar length.")
            return
        messagebox.showinfo(
            "Autocalibration",
            "Enter the scale value manually in nanometers.",
        )
        value_nm = simpledialog.askfloat(
            "Manual scale value",
            "Scale value (nm):",
            minvalue=0.0,
            parent=self.root,
        )
        if value_nm is None:
            self.status_var.set("Autocalibration cancelled.")
            return
        if value_nm <= 0:
            messagebox.showerror("Autocalibration", "Scale value must be > 0.")
            return
        value_um = float(value_nm) / 1000.0
        um_per_px = float(value_um) / float(bar_len_px)
        self.um_per_px_var.set(f"{um_per_px:.6f}")
        self.status_var.set(f"Autocalibration (manual): {value_nm:.2f} nm -> {um_per_px:.6f} um/px")

    def _detect_scale_bar(
        self,
        image_rgb: np.ndarray,
        allow_any_location: bool = False,
    ) -> Optional[Tuple[int, Tuple[int, int, int, int]]]:
        gray = cv2.cvtColor(image_rgb, cv2.COLOR_RGB2GRAY)
        blur = cv2.GaussianBlur(gray, (3, 3), 0)
        _thr, binary = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        # If the image is mostly bright, invert to keep the scale bar white.
        if float(np.mean(binary)) > 140.0:
            binary = cv2.bitwise_not(binary)

        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 3))
        binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel, iterations=1)

        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            return None

        h, w = gray.shape
        candidates: List[Tuple[float, float, Tuple[int, int, int, int]]] = []
        min_aspect = 6.0 if not allow_any_location else 4.0
        max_height_frac = 0.06 if not allow_any_location else 0.35
        min_width_frac = 0.08 if not allow_any_location else 0.20

        for cnt in contours:
            x, y, bw, bh = cv2.boundingRect(cnt)
            if bw <= 0 or bh <= 0:
                continue
            aspect = float(bw) / float(max(1, bh))
            if aspect < min_aspect:
                continue
            if bh > max_height_frac * h:
                continue
            if bw < min_width_frac * w:
                continue
            if (not allow_any_location) and (y < 0.45 * h):
                continue
            area = float(bw * bh)
            candidates.append((bw, area, (x, y, bw, bh)))

        if not candidates:
            # Fallback: try a line-based detector (Hough) if contours fail.
            result = self._detect_scale_bar_hough(image_rgb, allow_any_location=allow_any_location)
            if result is not None:
                return result
            return self._detect_scale_bar_projection(image_rgb)

        candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
        x, y, bw, bh = candidates[0][2]
        return int(bw), (int(x), int(y), int(bw), int(bh))

    def _detect_scale_bar_hough(
        self,
        image_rgb: np.ndarray,
        allow_any_location: bool = False,
    ) -> Optional[Tuple[int, Tuple[int, int, int, int]]]:
        gray = cv2.cvtColor(image_rgb, cv2.COLOR_RGB2GRAY)
        blur = cv2.GaussianBlur(gray, (3, 3), 0)
        edges = cv2.Canny(blur, 50, 150)

        h, w = gray.shape
        min_len = max(15, int(0.08 * w))
        lines = cv2.HoughLinesP(
            edges,
            rho=1,
            theta=np.pi / 180.0,
            threshold=60,
            minLineLength=min_len,
            maxLineGap=6,
        )
        if lines is None:
            return None

        best = None
        best_len = 0.0
        best_is_horiz = False
        for line in lines:
            x1, y1, x2, y2 = line[0]
            dx = float(x2 - x1)
            dy = float(y2 - y1)
            length = float(math.hypot(dx, dy))
            if length <= 0:
                continue
            is_horiz = abs(dy) <= 0.20 * max(1.0, abs(dx))
            is_vert = abs(dx) <= 0.20 * max(1.0, abs(dy))
            if not (is_horiz or is_vert):
                continue
            if not allow_any_location and is_horiz:
                if max(y1, y2) < 0.45 * h:
                    continue
            if length > best_len or (is_horiz and not best_is_horiz):
                best = (x1, y1, x2, y2, is_horiz)
                best_len = length
                best_is_horiz = is_horiz

        if best is None:
            return None

        x1, y1, x2, y2, _is_horiz = best
        x_min = int(min(x1, x2))
        x_max = int(max(x1, x2))
        y_min = int(min(y1, y2))
        y_max = int(max(y1, y2))
        pad = max(2, int(0.01 * max(h, w)))
        x_min = max(0, x_min - pad)
        y_min = max(0, y_min - pad)
        x_max = min(w - 1, x_max + pad)
        y_max = min(h - 1, y_max + pad)
        bw = max(1, x_max - x_min + 1)
        bh = max(1, y_max - y_min + 1)
        bar_len_px = int(round(best_len))
        return bar_len_px, (x_min, y_min, bw, bh)

    def _detect_scale_bar_projection(self, image_rgb: np.ndarray) -> Optional[Tuple[int, Tuple[int, int, int, int]]]:
        gray = cv2.cvtColor(image_rgb, cv2.COLOR_RGB2GRAY)
        blur = cv2.GaussianBlur(gray, (3, 3), 0)
        _thr, binary = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        if float(np.mean(binary)) > 140.0:
            binary = cv2.bitwise_not(binary)
        white = binary > 0
        h, w = white.shape

        row_thresh = max(6, int(0.4 * w))
        rows = np.where(white.sum(axis=1) >= row_thresh)[0]
        if rows.size > 0:
            groups = np.split(rows, np.where(np.diff(rows) != 1)[0] + 1)
            best = None
            for g in groups:
                if g.size == 0:
                    continue
                y0, y1 = int(g[0]), int(g[-1])
                band = white[y0:y1 + 1, :]
                cols = np.where(band.any(axis=0))[0]
                if cols.size == 0:
                    continue
                x0, x1 = int(cols[0]), int(cols[-1])
                length = x1 - x0 + 1
                height = y1 - y0 + 1
                score = length / max(1, height)
                if best is None or score > best[0]:
                    best = (score, x0, y0, x1, y1)
            if best is not None:
                _, x0, y0, x1, y1 = best
                bw = max(1, x1 - x0 + 1)
                bh = max(1, y1 - y0 + 1)
                return bw, (x0, y0, bw, bh)

        col_thresh = max(6, int(0.4 * h))
        cols = np.where(white.sum(axis=0) >= col_thresh)[0]
        if cols.size == 0:
            return None
        groups = np.split(cols, np.where(np.diff(cols) != 1)[0] + 1)
        best = None
        for g in groups:
            if g.size == 0:
                continue
            x0, x1 = int(g[0]), int(g[-1])
            band = white[:, x0:x1 + 1]
            rows = np.where(band.any(axis=1))[0]
            if rows.size == 0:
                continue
            y0, y1 = int(rows[0]), int(rows[-1])
            length = y1 - y0 + 1
            width = x1 - x0 + 1
            score = length / max(1, width)
            if best is None or score > best[0]:
                best = (score, x0, y0, x1, y1)
        if best is None:
            return None
        _, x0, y0, x1, y1 = best
        bw = max(1, x1 - x0 + 1)
        bh = max(1, y1 - y0 + 1)
        return bh, (x0, y0, bw, bh)


    def _parse_um_per_px(self) -> Optional[float]:
        try:
            value = float(self.um_per_px_var.get().strip())
        except ValueError:
            messagebox.showerror("Scale error", "Scale must be a float value in um/px.")
            return None
        if value <= 0:
            messagebox.showerror("Scale error", "Scale must be > 0.")
            return None
        return value

    def on_run_sam(self) -> None:
        if self.sam_running:
            return
        if not self.image_rgbs and self.image_rgb is None:
            messagebox.showwarning("No image", "Import an image first.")
            return
        if not SAM_AVAILABLE:
            messagebox.showerror("SAM missing", dependency_help_text())
            return
        ckpt = self.checkpoint_var.get().strip()
        model_type = self.model_type_var.get().strip()
        if not ckpt:
            messagebox.showerror("Checkpoint missing", "Provide a checkpoint path.")
            return
        if model_type not in ("vit_b", "vit_l", "vit_h"):
            messagebox.showerror("Model type", "Model type must be vit_b, vit_l, or vit_h.")
            return

        try:
            max_side = int(self.max_side_var.get().strip())
        except ValueError:
            messagebox.showerror("SAM setting", "Max side px must be an integer (e.g. 1280).")
            return
        if max_side < 256:
            messagebox.showerror("SAM setting", "Max side px must be >= 256.")
            return
        fast_mode = bool(self.fast_mode_var.get())

        images_to_run = self.image_rgbs if self.image_rgbs else [self.image_rgb]

        self.sam_running = True
        self.sam_job_id += 1
        job_id = self.sam_job_id
        self.sam_cancel_requested = False
        self.sam_cancel_event = threading.Event()
        self.sam_progress_var.set(1.0)
        self.sam_run_btn.state(["disabled"])
        self.sam_cancel_btn.state(["!disabled"])
        self.sam_started_at = time.time()
        try:
            self.sam_estimated_total_s = sum(
                self._estimate_sam_seconds(
                    image_shape=img.shape,
                    model_type=model_type,
                    max_side=max_side,
                    fast_mode=fast_mode,
                )
                for img in images_to_run
                if img is not None
            )
        except Exception:
            self.sam_estimated_total_s = 0.0
        self.status_var.set(
            f"Running SAM in background... ETA ~{int(self.sam_estimated_total_s)} s."
        )
        try:
            image_copies = [np.asarray(img, dtype=np.uint8).copy() for img in images_to_run if img is not None]
        except Exception:
            self.sam_running = False
            self.sam_run_btn.state(["!disabled"])
            self.sam_cancel_btn.state(["disabled"])
            self.status_var.set("SAM failed.")
            return

        self._start_progress_loop(job_id)
        worker = threading.Thread(
            target=self._sam_worker,
            args=(job_id, image_copies, ckpt, model_type, max_side, fast_mode, self.sam_cancel_event),
            daemon=True,
        )
        worker.start()

    def _sam_worker(
        self,
        job_id: int,
        image_list: List[np.ndarray],
        ckpt: str,
        model_type: str,
        max_side: int,
        fast_mode: bool,
        cancel_event: threading.Event,
    ) -> None:
        def progress_from_worker(message: str, pct: float) -> None:
            self.root.after(
                0,
                lambda: self._on_sam_stage_progress(
                    job_id=job_id,
                    message=message,
                    pct=pct,
                ),
            )

        try:
            total = max(1, len(image_list))
            all_particles: List[List[ParticleMask]] = []
            for idx, img in enumerate(image_list):
                if cancel_event.is_set():
                    raise RuntimeError("SAM cancelled by user.")

                def per_image_progress(message: str, pct: float, i: int = idx) -> None:
                    overall = (float(i) + (pct / 100.0)) / float(total) * 100.0
                    progress_from_worker(f"[{i + 1}/{total}] {message}", overall)

                particles = run_sam(
                    image_rgb=img,
                    checkpoint_path=ckpt,
                    model_type=model_type,
                    max_side=max_side,
                    fast_mode=fast_mode,
                    progress_callback=per_image_progress,
                    cancel_event=cancel_event,
                )
                all_particles.append(particles)

            self.root.after(0, lambda: self._on_sam_done(job_id=job_id, particles=all_particles, err=None))
        except Exception as exc:
            self.root.after(0, lambda: self._on_sam_done(job_id=job_id, particles=None, err=exc))

    def _on_sam_done(self, job_id: int, particles: Optional[List[List[ParticleMask]]], err: Optional[Exception]) -> None:
        if job_id != self.sam_job_id:
            return
        self._stop_progress_loop()
        self.sam_running = False
        self.sam_run_btn.state(["!disabled"])
        self.sam_cancel_btn.state(["disabled"])
        if err is not None:
            if "cancelled by user" in str(err).lower():
                self.sam_progress_var.set(0.0)
                self.status_var.set("SAM cancelled.")
            else:
                messagebox.showerror("SAM error", str(err))
                self.sam_progress_var.set(0.0)
                self.status_var.set("SAM failed.")
            return

        self.sam_progress_var.set(100.0)
        self.particles_by_image = {}
        self.particles = []
        next_id = 1
        if particles:
            for img_idx, plist in enumerate(particles):
                for p in plist:
                    p.mask_id = next_id
                    p.image_index = img_idx
                    p.offset_xy = self.image_offsets.get(img_idx, (0, 0))
                    next_id += 1
                self.particles_by_image[img_idx] = plist
        self.selected_ids.clear()
        self.highlighted_ids.clear()
        self._sync_particles_for_layout()
        self.render_image()
        self.refresh_table()
        self.status_var.set(f"SAM ready: {len(self.particles)} masks.")
        self.root.after(1200, lambda: self.sam_progress_var.set(0.0))

    def on_cancel_sam(self) -> None:
        if not self.sam_running:
            return
        if self.sam_cancel_event is None:
            return
        self.sam_cancel_requested = True
        self.sam_cancel_event.set()
        self.sam_cancel_btn.state(["disabled"])
        self.status_var.set("Cancel requested... waiting for current SAM step to stop.")

    def _on_sam_stage_progress(self, job_id: int, message: str, pct: float) -> None:
        if job_id != self.sam_job_id or not self.sam_running:
            return
        current = float(self.sam_progress_var.get())
        new_value = max(current, min(100.0, float(pct)))
        self.sam_progress_var.set(new_value)
        if not self.sam_cancel_requested:
            self.status_var.set(f"{message} ({int(new_value)}%)")

    def _estimate_sam_seconds(
        self,
        image_shape: Tuple[int, int, int],
        model_type: str,
        max_side: int,
        fast_mode: bool,
    ) -> float:
        h, w = image_shape[:2]
        scale = min(1.0, float(max_side) / float(max(h, w)))
        proc_h = max(1.0, float(h) * scale)
        proc_w = max(1.0, float(w) * scale)
        mpix = (proc_h * proc_w) / 1_000_000.0

        sec_per_mpix = 28.0 if fast_mode else 48.0
        model_factor = {"vit_b": 1.0, "vit_l": 1.45, "vit_h": 1.9}.get(model_type, 1.0)
        estimated = 10.0 + (mpix * sec_per_mpix * model_factor)
        return max(15.0, estimated)

    def _start_progress_loop(self, job_id: int) -> None:
        self._stop_progress_loop()
        self.sam_progress_after_id = self.root.after(250, lambda: self._progress_tick(job_id))

    def _stop_progress_loop(self) -> None:
        if self.sam_progress_after_id is not None:
            try:
                self.root.after_cancel(self.sam_progress_after_id)
            except Exception:
                pass
            self.sam_progress_after_id = None

    def _progress_tick(self, job_id: int) -> None:
        if not self.sam_running or job_id != self.sam_job_id:
            return
        elapsed = max(0.0, time.time() - self.sam_started_at)
        if self.sam_estimated_total_s > 0:
            ratio = min(1.0, elapsed / self.sam_estimated_total_s)
        else:
            ratio = 0.0

        current = float(self.sam_progress_var.get())
        estimated_progress = 5.0 + (ratio * 88.0)  # leaves headroom for final steps
        if not self.sam_cancel_requested and estimated_progress > current:
            self.sam_progress_var.set(min(95.0, estimated_progress))
            remaining = max(0.0, self.sam_estimated_total_s - elapsed)
            self.status_var.set(
                f"Running SAM... {int(self.sam_progress_var.get())}% "
                f"(ETA ~{int(remaining)} s)"
            )
        self.sam_progress_after_id = self.root.after(250, lambda: self._progress_tick(job_id))

    def _build_pick_map(self) -> None:
        if self.image_rgb is None:
            self.mask_pick_map = None
            return
        h, w = self.image_rgb.shape[:2]
        pick_map = np.full((h, w), -1, dtype=np.int32)
        # Paint larger masks first, so smaller ones remain selectable in overlaps.
        by_area = sorted(self.particles, key=lambda p: p.area_px, reverse=True)
        for p in by_area:
            self._apply_mask_to_canvas(pick_map, p, p.mask_id)
        self.mask_pick_map = pick_map

    def set_mode(self, mode: str) -> None:
        self.mode = mode
        self._set_toolbar_pan(mode == "navigate")
        self._clear_roi_artist()
        self.drag_start = None
        self.status_var.set(f"Mode: {mode}")

    def _set_toolbar_pan(self, enable: bool) -> None:
        mode_text = self.toolbar.mode or ""
        pan_active = "pan/zoom" in mode_text.lower()
        if enable and not pan_active:
            self.toolbar.pan()
        elif (not enable) and pan_active:
            self.toolbar.pan()

    def _clear_roi_artist(self) -> None:
        if self.roi_artist is not None:
            try:
                self.roi_artist.remove()
            except Exception:
                pass
            self.roi_artist = None

    def _selection_rgba(self, alpha: Optional[float] = None) -> Tuple[float, float, float, float]:
        r, g, b = self.selection_color_rgb
        a = self.selection_alpha if alpha is None else alpha
        return (r / 255.0, g / 255.0, b / 255.0, a)

    def _particle_offset(self, particle: ParticleMask) -> Tuple[int, int]:
        return particle.offset_xy if hasattr(particle, "offset_xy") else (0, 0)

    def _particle_centroid_global(self, particle: ParticleMask) -> Tuple[float, float]:
        ox, oy = self._particle_offset(particle)
        return (particle.centroid_xy[0] + ox, particle.centroid_xy[1] + oy)

    def _apply_mask_to_canvas(
        self,
        canvas: np.ndarray,
        particle: ParticleMask,
        value: np.ndarray,
    ) -> None:
        if canvas is None:
            return
        ox, oy = self._particle_offset(particle)
        seg = particle.segmentation
        if seg is None or seg.size == 0:
            return
        h, w = seg.shape
        y0, y1 = oy, oy + h
        x0, x1 = ox, ox + w
        if y1 <= 0 or x1 <= 0:
            return
        if y0 >= canvas.shape[0] or x0 >= canvas.shape[1]:
            return
        y0_clamp = max(0, y0)
        x0_clamp = max(0, x0)
        y1_clamp = min(canvas.shape[0], y1)
        x1_clamp = min(canvas.shape[1], x1)
        seg_view = seg[(y0_clamp - y0) : (y1_clamp - y0), (x0_clamp - x0) : (x1_clamp - x0)]
        region = canvas[y0_clamp:y1_clamp, x0_clamp:x1_clamp]
        region[seg_view] = value

    def render_image(self, keep_view: bool = False) -> None:
        if self.image_rgb is None:
            self.ax.clear()
            self.ax.set_title("No image loaded")
            self.ax.axis("off")
            self.canvas.draw_idle()
            return

        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()

        self.ax.clear()
        self.ax.imshow(self.image_rgb, interpolation="nearest")
        if self.image_path:
            title = os.path.basename(self.image_path)
        elif self.image_paths:
            title = f"{len(self.image_paths)} images"
        else:
            title = "Image"
        self.ax.set_title(title)
        self.ax.axis("off")

        if self.selected_ids:
            h, w = self.image_rgb.shape[:2]
            overlay = np.zeros((h, w, 4), dtype=np.float32)
            base_rgba = np.array(self._selection_rgba(), dtype=np.float32)
            for pid in self.selected_ids:
                p = self._particle_by_id(pid)
                if p is None:
                    continue
                if pid in self.highlighted_ids:
                    self._apply_mask_to_canvas(
                        overlay,
                        p,
                        np.array([0.92, 0.17, 0.15, 0.52], dtype=np.float32),
                    )
                else:
                    self._apply_mask_to_canvas(overlay, p, base_rgba)
            self.ax.imshow(overlay, interpolation="nearest")

            # Draw selected particle IDs on top of each selected mask.
            if not (self.correct_params_active and self.correct_param_name in ("length", "width")):
                for pid in sorted(self.selected_ids):
                    p = self._particle_by_id(pid)
                    if p is None:
                        continue
                    cx, cy = self._particle_centroid_global(p)
                    self.ax.text(
                        cx,
                        cy,
                        f"#{pid}",
                        color="white",
                        fontsize=9,
                        fontweight="bold",
                        ha="center",
                        va="center",
                        bbox=dict(
                            boxstyle="round,pad=0.2",
                            facecolor="#381015" if pid in self.highlighted_ids else self._selection_rgba(alpha=0.75),
                            edgecolor="#ff8a80" if pid in self.highlighted_ids else "white",
                            linewidth=0.6,
                            alpha=0.75,
                        ),
                    )

            if self.correct_params_active and self.correct_param_name in ("length", "width"):
                self._build_correct_lines()

        if self.measure_line is not None:
            x0, y0, x1, y1 = self.measure_line
            self.ax.plot([x0, x1], [y0, y1], color="#00FFFF", linewidth=2.2)
            if self.measure_label:
                xm = 0.5 * (x0 + x1)
                ym = 0.5 * (y0 + y1)
                self.ax.text(
                    xm,
                    ym,
                    self.measure_label,
                    color="white",
                    fontsize=9,
                    fontweight="bold",
                    ha="left",
                    va="bottom",
                    bbox=dict(
                        boxstyle="round,pad=0.2",
                        facecolor="black",
                        edgecolor="#00FFFF",
                        linewidth=0.7,
                        alpha=0.8,
                    ),
                )

        if self.angle_points:
            if len(self.angle_points) >= 2:
                p1, p2 = self.angle_points[0], self.angle_points[1]
                self.ax.plot([p1[0], p2[0]], [p1[1], p2[1]], color="#FFD54F", linewidth=2.2)
            if len(self.angle_points) >= 3:
                p2, p3 = self.angle_points[1], self.angle_points[2]
                self.ax.plot([p2[0], p3[0]], [p2[1], p3[1]], color="#FFD54F", linewidth=2.2)
                if self.angle_label:
                    self.ax.text(
                        p2[0],
                        p2[1],
                        self.angle_label,
                        color="black",
                        fontsize=9,
                        fontweight="bold",
                        ha="left",
                        va="bottom",
                        bbox=dict(
                            boxstyle="round,pad=0.2",
                            facecolor="#FFD54F",
                            edgecolor="#AA8B2A",
                            linewidth=0.7,
                            alpha=0.9,
                        ),
                    )
            elif len(self.angle_points) == 2 and self.angle_preview_point is not None:
                p2, p3 = self.angle_points[1], self.angle_preview_point
                self.ax.plot([p2[0], p3[0]], [p2[1], p3[1]], color="#FFD54F", linewidth=1.6, alpha=0.8)
                if self.angle_label:
                    self.ax.text(
                        p2[0],
                        p2[1],
                        self.angle_label,
                        color="black",
                        fontsize=9,
                        fontweight="bold",
                        ha="left",
                        va="bottom",
                        bbox=dict(
                            boxstyle="round,pad=0.2",
                            facecolor="#FFD54F",
                            edgecolor="#AA8B2A",
                            linewidth=0.7,
                            alpha=0.9,
                        ),
                    )

        if self.section_shape and self.section_bbox:
            x_min, y_min, x_max, y_max = self.section_bbox
            cx = 0.5 * (x_min + x_max)
            cy = 0.5 * (y_min + y_max)
            w = max(0.0, x_max - x_min)
            h = max(0.0, y_max - y_min)
            if self.section_shape.lower().startswith("oval"):
                patch = Ellipse((cx, cy), w, h, fill=False, linewidth=2, edgecolor="#FFB74D")
                self.ax.add_patch(patch)
            else:
                patch = Rectangle((x_min, y_min), w, h, fill=False, linewidth=2, edgecolor="#FFB74D")
                self.ax.add_patch(patch)
            if self.section_label:
                self.ax.text(
                    cx,
                    cy,
                    self.section_label,
                    color="black",
                    fontsize=9,
                    fontweight="bold",
                    ha="center",
                    va="center",
                    bbox=dict(
                        boxstyle="round,pad=0.2",
                        facecolor="#FFECB3",
                        edgecolor="#FFB74D",
                        linewidth=0.7,
                        alpha=0.9,
                    ),
                )

        if keep_view and self.image_rgb is not None:
            self.ax.set_xlim(xlim)
            self.ax.set_ylim(ylim)

        self.canvas.draw_idle()

    def _particle_by_id(self, pid: int) -> Optional[ParticleMask]:
        for p in self.particles:
            if p.mask_id == pid:
                return p
        return None

    def on_mouse_press(self, event) -> None:
        if event.inaxes != self.ax:
            return
        if event.xdata is None or event.ydata is None:
            return

        if self._handle_correct_press(event):
            return

        if self.mode == "select_one":
            self._toggle_single_selection(float(event.xdata), float(event.ydata))
            return

        if self.mode == "measure_angle":
            self._handle_angle_click(float(event.xdata), float(event.ydata))
            return

        if self.mode == "measure_line":
            self.drag_start = (float(event.xdata), float(event.ydata))
            self._clear_roi_artist()
            line = Line2D(
                [event.xdata, event.xdata],
                [event.ydata, event.ydata],
                color="#00FFFF",
                linewidth=2.2,
            )
            self.roi_artist = line
            self.ax.add_line(line)
            self.canvas.draw_idle()
            return

        if self.mode in ("oval_section", "rect_section"):
            self.drag_start = (float(event.xdata), float(event.ydata))
            self._clear_roi_artist()
            if self.mode == "oval_section":
                self.roi_artist = Ellipse(
                    (event.xdata, event.ydata),
                    0,
                    0,
                    fill=False,
                    linewidth=2,
                    edgecolor="#FFB74D",
                )
            else:
                self.roi_artist = Rectangle(
                    (event.xdata, event.ydata),
                    0,
                    0,
                    fill=False,
                    linewidth=2,
                    edgecolor="#FFB74D",
                )
            self.ax.add_patch(self.roi_artist)
            self.canvas.draw_idle()
            return

        if self.mode in ("roi_rect", "roi_circle", "deselect_rect", "scale_bar_rect"):
            self.drag_start = (float(event.xdata), float(event.ydata))
            self._clear_roi_artist()
            if self.mode in ("roi_rect", "deselect_rect", "scale_bar_rect"):
                self.roi_artist = Rectangle(
                    (event.xdata, event.ydata),
                    0,
                    0,
                    fill=False,
                    linewidth=2,
                    edgecolor=(
                        "#ff8f00"
                        if self.mode == "scale_bar_rect"
                        else ("red" if self.mode == "deselect_rect" else "cyan")
                    ),
                )
            else:
                self.roi_artist = Circle(
                    (event.xdata, event.ydata),
                    radius=0,
                    fill=False,
                    linewidth=2,
                    edgecolor="yellow",
                )
            self.ax.add_patch(self.roi_artist)
            self.canvas.draw_idle()
            return

    def on_mouse_move(self, event) -> None:
        if self._handle_correct_move(event):
            return
        if self.mode == "measure_angle":
            if self.angle_points and len(self.angle_points) == 2:
                if event.inaxes == self.ax and event.xdata is not None and event.ydata is not None:
                    self.angle_preview_point = (float(event.xdata), float(event.ydata))
                    p1, p2 = self.angle_points[0], self.angle_points[1]
                    angle_deg = self._compute_angle_deg(p1, p2, self.angle_preview_point)
                    if np.isfinite(angle_deg):
                        self.angle_label = f"{angle_deg:.2f}°"
                    else:
                        self.angle_label = "N/A"
                    self.render_image(keep_view=True)
            return
        if self.drag_start is None:
            return
        if self.roi_artist is None:
            return
        if event.inaxes != self.ax:
            return
        if event.xdata is None or event.ydata is None:
            return

        x0, y0 = self.drag_start
        x1, y1 = float(event.xdata), float(event.ydata)

        if isinstance(self.roi_artist, Rectangle):
            x = min(x0, x1)
            y = min(y0, y1)
            w = abs(x1 - x0)
            h = abs(y1 - y0)
            self.roi_artist.set_xy((x, y))
            self.roi_artist.set_width(w)
            self.roi_artist.set_height(h)
        elif isinstance(self.roi_artist, Ellipse):
            cx = 0.5 * (x0 + x1)
            cy = 0.5 * (y0 + y1)
            w = abs(x1 - x0)
            h = abs(y1 - y0)
            self.roi_artist.center = (cx, cy)
            self.roi_artist.width = w
            self.roi_artist.height = h
        elif isinstance(self.roi_artist, Circle):
            radius = float(np.hypot(x1 - x0, y1 - y0))
            self.roi_artist.center = (x0, y0)
            self.roi_artist.set_radius(radius)
        elif isinstance(self.roi_artist, Line2D):
            self.roi_artist.set_data([x0, x1], [y0, y1])

        self.canvas.draw_idle()

    def on_mouse_release(self, event) -> None:
        if self._handle_correct_release(event):
            return
        if self.drag_start is None:
            return
        if self.roi_artist is None:
            self.drag_start = None
            return
        if event.inaxes != self.ax or event.xdata is None or event.ydata is None:
            self._clear_roi_artist()
            self.drag_start = None
            self.canvas.draw_idle()
            return

        x0, y0 = self.drag_start
        x1, y1 = float(event.xdata), float(event.ydata)

        if isinstance(self.roi_artist, Rectangle):
            x_min, x_max = sorted((x0, x1))
            y_min, y_max = sorted((y0, y1))
            if self.mode == "scale_bar_rect":
                self._manual_scale_bar_selected(x_min, y_min, x_max, y_max)
            elif self.mode == "rect_section":
                self._store_section("Rectangle", x_min, y_min, x_max, y_max)
            elif self.mode == "deselect_rect":
                self._deselect_particles_rect(x_min, y_min, x_max, y_max)
            else:
                self._select_particles_rect(x_min, y_min, x_max, y_max)
        elif isinstance(self.roi_artist, Ellipse):
            x_min, x_max = sorted((x0, x1))
            y_min, y_max = sorted((y0, y1))
            self._store_section("Oval", x_min, y_min, x_max, y_max)
        elif isinstance(self.roi_artist, Circle):
            radius = float(np.hypot(x1 - x0, y1 - y0))
            self._select_particles_circle(x0, y0, radius)
        elif isinstance(self.roi_artist, Line2D):
            self._store_measurement_line(x0, y0, x1, y1)

        self._clear_roi_artist()
        self.drag_start = None
        self.render_image(keep_view=True)
        self.refresh_table()

    def on_mouse_scroll(self, event) -> None:
        if self.mode != "navigate":
            return
        if self.image_rgb is None:
            return
        if event.inaxes != self.ax:
            return

        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()
        xdata = event.xdata if event.xdata is not None else 0.5 * (xlim[0] + xlim[1])
        ydata = event.ydata if event.ydata is not None else 0.5 * (ylim[0] + ylim[1])

        if getattr(event, "button", None) == "up":
            scale = 1.0 / 1.2
        elif getattr(event, "button", None) == "down":
            scale = 1.2
        else:
            return

        left = xdata - (xdata - xlim[0]) * scale
        right = xdata + (xlim[1] - xdata) * scale
        bottom = ydata - (ydata - ylim[0]) * scale
        top = ydata + (ylim[1] - ydata) * scale

        self.ax.set_xlim(left, right)
        self.ax.set_ylim(bottom, top)
        self.canvas.draw_idle()

    def _toggle_single_selection(self, x: float, y: float) -> None:
        if self.mask_pick_map is None:
            return
        h, w = self.mask_pick_map.shape
        xi = int(round(x))
        yi = int(round(y))
        if xi < 0 or yi < 0 or xi >= w or yi >= h:
            return
        mask_id = int(self.mask_pick_map[yi, xi])
        if mask_id < 0:
            return
        if mask_id in self.selected_ids:
            self.selected_ids.remove(mask_id)
        else:
            self.selected_ids.add(mask_id)
        self.highlighted_ids &= self.selected_ids
        self.render_image(keep_view=True)
        self.refresh_table()

    def _select_particles_rect(self, x_min: float, y_min: float, x_max: float, y_max: float) -> None:
        if not self.particles:
            return
        added = 0
        for p in self.particles:
            cx, cy = self._particle_centroid_global(p)
            if x_min <= cx <= x_max and y_min <= cy <= y_max:
                if p.mask_id not in self.selected_ids:
                    added += 1
                self.selected_ids.add(p.mask_id)
        self.highlighted_ids &= self.selected_ids
        self.status_var.set(f"Rect ROI selected +{added} particles.")

    def _deselect_particles_rect(self, x_min: float, y_min: float, x_max: float, y_max: float) -> None:
        if not self.particles or not self.selected_ids:
            self.status_var.set("De-selector removed 0 particles.")
            return
        to_remove: List[int] = []
        for p in self.particles:
            if p.mask_id not in self.selected_ids:
                continue
            cx, cy = self._particle_centroid_global(p)
            if x_min <= cx <= x_max and y_min <= cy <= y_max:
                to_remove.append(int(p.mask_id))
        for pid in to_remove:
            self.selected_ids.discard(pid)
        self.highlighted_ids &= self.selected_ids
        self.status_var.set(f"De-selector removed {len(to_remove)} particles.")

    def _select_particles_circle(self, cx0: float, cy0: float, radius: float) -> None:
        if not self.particles:
            return
        r2 = radius * radius
        added = 0
        for p in self.particles:
            cx, cy = self._particle_centroid_global(p)
            if ((cx - cx0) ** 2 + (cy - cy0) ** 2) <= r2:
                if p.mask_id not in self.selected_ids:
                    added += 1
                self.selected_ids.add(p.mask_id)
        self.highlighted_ids &= self.selected_ids
        self.status_var.set(f"Circle ROI selected +{added} particles.")

    def clear_selection(self) -> None:
        self.selected_ids.clear()
        self.highlighted_ids.clear()
        self.render_image(keep_view=True)
        self.refresh_table()
        self.status_var.set("Selection cleared.")

    def clear_measurement(self) -> None:
        self.measure_line = None
        self.measure_label = None
        self.angle_points = []
        self.angle_label = None
        self.angle_preview_point = None
        self.section_shape = None
        self.section_bbox = None
        self.section_label = None
        self.render_image(keep_view=True)
        self.status_var.set("Measurement cleared.")

    def open_flip_image(self) -> None:
        if self.image_rgb is None:
            messagebox.showerror("Error", "Import an image first.")
            return
        if self.flip_window is not None and self.flip_window.winfo_exists():
            self.flip_window.lift()
            return

        self.flip_window = tk.Toplevel(self.root)
        self.flip_window.title("Flip Image")
        self.flip_window.geometry("300x220+930+240")

        outer = ttk.Frame(self.flip_window, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        ttk.Label(outer, text="Flip Image", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 10))

        btn_frame = ttk.Frame(outer)
        btn_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Button(
            btn_frame,
            text="Flip Horizontally",
            command=lambda: self._apply_image_transform(lambda arr: arr[:, ::-1], "Flipped horizontally"),
        ).grid(row=0, column=0, sticky="ew", padx=4, pady=4)

        ttk.Button(
            btn_frame,
            text="Flip Vertically",
            command=lambda: self._apply_image_transform(lambda arr: arr[::-1, :], "Flipped vertically"),
        ).grid(row=0, column=1, sticky="ew", padx=4, pady=4)

        ttk.Button(
            btn_frame,
            text="Flip 90º Left",
            command=lambda: self._apply_image_transform(lambda arr: np.rot90(arr, 1), "Rotated 90º left"),
        ).grid(row=1, column=0, sticky="ew", padx=4, pady=4)

        ttk.Button(
            btn_frame,
            text="Flip 90º Right",
            command=lambda: self._apply_image_transform(lambda arr: np.rot90(arr, -1), "Rotated 90º right"),
        ).grid(row=1, column=1, sticky="ew", padx=4, pady=4)

        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)

        self.flip_window.protocol("WM_DELETE_WINDOW", self._close_flip_window)

    def open_iou_tool(self) -> None:
        if not self.selected_ids:
            messagebox.showerror("AI Evaluation", "Select particles")
            return
        if self.image_rgb is None:
            messagebox.showerror("AI Evaluation", "Import an image first.")
            return
        if self.iou_window is not None and self.iou_window.winfo_exists():
            self.iou_window.lift()
            return

        self.iou_selected_snapshot = set(self.selected_ids)
        self.iou_draw_enabled = False
        self.iou_last_pos = None

        self.iou_window = tk.Toplevel(self.root)
        self.iou_window.title("AI Evaluation")
        self.iou_window.geometry("1180x740+820+120")

        outer = ttk.Frame(self.iou_window, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(outer)
        header.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(
            header,
            text="Fill the selected particles to create ground truth masks.",
            font=("Segoe UI", 10),
        ).pack(side=tk.LEFT, anchor="w")
        ttk.Button(header, text="Save Image", command=self._save_iou_overlay_image).pack(side=tk.RIGHT)
        ttk.Button(header, text="Close", command=self._close_iou_window).pack(side=tk.RIGHT, padx=(0, 6))

        body = ttk.Frame(outer)
        body.pack(fill=tk.BOTH, expand=True)

        left = ttk.Frame(body)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        mid = ttk.Frame(body, width=520)
        mid.pack(side=tk.LEFT, fill=tk.Y, padx=(10, 0))
        mid.pack_propagate(False)

        right = ttk.Frame(body, width=320)
        right.pack(side=tk.LEFT, fill=tk.Y, padx=(10, 0))
        right.pack_propagate(False)

        base = self._build_iou_base_image()

        max_dim = 840
        w, h = base.size
        scale = min(1.0, float(max_dim) / float(max(w, h)))
        disp_w = max(1, int(round(w * scale)))
        disp_h = max(1, int(round(h * scale)))
        self.iou_scale = scale

        self.iou_base_image = base
        self.iou_overlay_image = Image.new("RGBA", (disp_w, disp_h), (0, 0, 0, 0))
        self.iou_mask = Image.new("L", (w, h), 0)

        self.iou_canvas = tk.Canvas(left, width=disp_w, height=disp_h, bg="#111111", highlightthickness=0)
        self.iou_canvas.pack(fill=tk.BOTH, expand=True)
        self._refresh_iou_canvas()
        self.iou_canvas.bind("<ButtonPress-1>", self._on_iou_brush_press)
        self.iou_canvas.bind("<B1-Motion>", self._on_iou_brush_drag)
        self.iou_canvas.bind("<ButtonRelease-1>", self._on_iou_brush_release)

        ba_frame = ttk.LabelFrame(mid, text="Bland-Altman Plot", padding=8)
        ba_frame.pack(fill=tk.BOTH, expand=True)
        self.iou_ba_fig = Figure(figsize=(5.6, 2.6), dpi=100)
        self.iou_ba_ax = self.iou_ba_fig.add_subplot(111)
        self.iou_ba_ax.set_xlabel("Mean area (nm²)")
        self.iou_ba_ax.set_ylabel("Diff (Manual - SAM)")
        self.iou_ba_ax.tick_params(labelsize=PLOT_TEXT_SIZE, pad=2)
        self.iou_ba_ax.grid(True, alpha=0.2)
        self.iou_ba_fig.subplots_adjust(left=0.16, right=0.98, bottom=0.22, top=0.92)
        self.iou_ba_canvas = FigureCanvasTkAgg(self.iou_ba_fig, master=ba_frame)
        self.iou_ba_canvas.draw()
        self.iou_ba_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, pady=(0, 6))
        ttk.Button(ba_frame, text="Save Graph", command=self._save_iou_ba_plot).pack(fill=tk.X)

        mode_frame = ttk.LabelFrame(right, text="Tool", padding=8)
        mode_frame.pack(fill=tk.X, pady=(0, 8))
        brush_btn = ttk.Button(mode_frame, text="Brush", command=lambda: self._set_iou_brush_mode("brush"))
        brush_btn.pack(fill=tk.X, pady=2)
        eraser_btn = ttk.Button(mode_frame, text="Eraser", command=lambda: self._set_iou_brush_mode("eraser"))
        eraser_btn.pack(fill=tk.X, pady=2)
        ttk.Button(mode_frame, text="Clear mask", command=self._clear_iou_mask).pack(fill=tk.X, pady=2)
        self.iou_mode_var = tk.StringVar(value="Active: Brush")
        ttk.Label(
            mode_frame,
            textvariable=self.iou_mode_var,
            font=("Segoe UI", 9, "bold"),
        ).pack(fill=tk.X, pady=(6, 0))
        self.iou_tool_buttons = {"brush": brush_btn, "eraser": eraser_btn}
        self._update_iou_tool_indicator()

        brush_frame = ttk.LabelFrame(right, text="Brush", padding=8)
        brush_frame.pack(fill=tk.X, pady=(0, 8))

        color_row = ttk.Frame(brush_frame)
        color_row.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(color_row, text="Color:").pack(side=tk.LEFT)
        ttk.Button(color_row, text="Pick", command=self._choose_iou_brush_color).pack(side=tk.LEFT, padx=(6, 6))
        self.iou_brush_preview = tk.Canvas(color_row, width=28, height=16, highlightthickness=1, highlightbackground="#999")
        self.iou_brush_preview.pack(side=tk.LEFT)
        self._update_iou_brush_preview()

        size_row = ttk.Frame(brush_frame)
        size_row.pack(fill=tk.X, pady=(4, 0))
        ttk.Label(size_row, text="Diameter (px):").pack(anchor="w")
        self.iou_brush_size_var = tk.StringVar(value="18")
        self.iou_brush_size_scale = ttk.Scale(
            size_row,
            from_=4,
            to=80,
            orient=tk.HORIZONTAL,
            command=self._on_iou_brush_scale,
        )
        self.iou_brush_size_scale.set(18)
        self.iou_brush_size_scale.pack(fill=tk.X, pady=(2, 2))
        size_entry = ttk.Entry(size_row, textvariable=self.iou_brush_size_var, width=8)
        size_entry.pack(anchor="w")
        size_entry.bind("<Return>", lambda _evt: self._on_iou_brush_entry())
        size_entry.bind("<FocusOut>", lambda _evt: self._on_iou_brush_entry())

        compute_frame = ttk.LabelFrame(right, text="AI Evaluation", padding=8)
        compute_frame.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(compute_frame, text="Compute Metrics", command=self._compute_iou).pack(fill=tk.X, pady=(0, 6))

        ttk.Label(compute_frame, text="IoU:").pack(anchor="w")
        self.iou_result_var = tk.StringVar(value="")
        ttk.Entry(compute_frame, textvariable=self.iou_result_var, state="readonly").pack(fill=tk.X, pady=(0, 4))

        ttk.Label(compute_frame, text="Precision:").pack(anchor="w")
        self.iou_precision_var = tk.StringVar(value="")
        ttk.Entry(compute_frame, textvariable=self.iou_precision_var, state="readonly").pack(fill=tk.X, pady=(0, 4))

        ttk.Label(compute_frame, text="Recall:").pack(anchor="w")
        self.iou_recall_var = tk.StringVar(value="")
        ttk.Entry(compute_frame, textvariable=self.iou_recall_var, state="readonly").pack(fill=tk.X, pady=(0, 4))

        ttk.Label(compute_frame, text="F1-score:").pack(anchor="w")
        self.iou_f1_var = tk.StringVar(value="")
        ttk.Entry(compute_frame, textvariable=self.iou_f1_var, state="readonly").pack(fill=tk.X)

        ttk.Separator(compute_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(8, 6))
        ttk.Label(compute_frame, text="Bland-Altman (Area):", font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 4))

        ttk.Label(compute_frame, text="Mean error (nm²):").pack(anchor="w")
        self.iou_ba_mean_var = tk.StringVar(value="")
        ttk.Entry(compute_frame, textvariable=self.iou_ba_mean_var, state="readonly").pack(fill=tk.X, pady=(0, 4))

        ttk.Label(compute_frame, text="Std dev (nm²):").pack(anchor="w")
        self.iou_ba_sd_var = tk.StringVar(value="")
        ttk.Entry(compute_frame, textvariable=self.iou_ba_sd_var, state="readonly").pack(fill=tk.X, pady=(0, 4))

        ttk.Label(compute_frame, text="LoA (nm²):").pack(anchor="w")
        self.iou_ba_loa_var = tk.StringVar(value="")
        ttk.Entry(compute_frame, textvariable=self.iou_ba_loa_var, state="readonly").pack(fill=tk.X)
        ttk.Button(compute_frame, text="Download Data", command=self._save_iou_ba_data).pack(fill=tk.X, pady=(6, 0))

        self.iou_window.protocol("WM_DELETE_WINDOW", self._close_iou_window)

    def _close_iou_window(self) -> None:
        if self.iou_window is not None and self.iou_window.winfo_exists():
            self.iou_window.destroy()
        self.iou_window = None
        self.iou_canvas = None
        self.iou_base_image = None
        self.iou_overlay_image = None
        self.iou_display_tk = None
        self.iou_mask = None
        self.iou_brush_size_var = None
        self.iou_brush_size_scale = None
        self.iou_brush_preview = None
        self.iou_result_var = None
        self.iou_precision_var = None
        self.iou_recall_var = None
        self.iou_f1_var = None
        self.iou_ba_mean_var = None
        self.iou_ba_sd_var = None
        self.iou_ba_loa_var = None
        self.iou_ba_data = None
        self.iou_ba_fig = None
        self.iou_ba_ax = None
        self.iou_ba_canvas = None
        self.iou_mode_var = None
        self.iou_draw_enabled = False
        self.iou_selected_snapshot = set()
        self.iou_last_pos = None

    def _enable_iou_drawing(self) -> None:
        self.iou_draw_enabled = True
        self.status_var.set("AI Evaluation: drawing enabled. Paint the ground truth masks.")

    def _ensure_iou_drawing_enabled(self) -> None:
        if not self.iou_draw_enabled:
            self._enable_iou_drawing()

    def _set_iou_brush_mode(self, mode: str) -> None:
        if mode not in ("brush", "eraser"):
            return
        self.iou_brush_mode = mode
        self._ensure_iou_drawing_enabled()
        self._update_iou_tool_indicator()
        self.status_var.set(f"AI Evaluation tool: {mode}")

    def _update_iou_tool_indicator(self) -> None:
        for mode, btn in self.iou_tool_buttons.items():
            try:
                btn.configure(relief="sunken" if mode == self.iou_brush_mode else "raised")
            except Exception:
                pass
        if self.iou_mode_var is not None:
            label = "Brush" if self.iou_brush_mode == "brush" else "Eraser"
            self.iou_mode_var.set(f"Active: {label}")

    def _on_iou_brush_scale(self, _value: str) -> None:
        if self.iou_brush_size_var is None or self.iou_brush_size_scale is None:
            return
        size = int(round(self.iou_brush_size_scale.get()))
        self.iou_brush_size_var.set(str(size))

    def _on_iou_brush_entry(self) -> None:
        if self.iou_brush_size_var is None or self.iou_brush_size_scale is None:
            return
        try:
            size = int(float(self.iou_brush_size_var.get().strip()))
        except Exception:
            size = 18
        size = max(1, min(200, size))
        self.iou_brush_size_var.set(str(size))
        self.iou_brush_size_scale.set(size)

    def _choose_iou_brush_color(self) -> None:
        color = colorchooser.askcolor(color=self.iou_brush_color, parent=self.iou_window)
        if color is None or color[0] is None:
            return
        r, g, b = color[0]
        self.iou_brush_color = (int(r), int(g), int(b))
        self._update_iou_brush_preview()

    def _update_iou_brush_preview(self) -> None:
        if self.iou_brush_preview is None:
            return
        r, g, b = self.iou_brush_color
        self.iou_brush_preview.delete("all")
        self.iou_brush_preview.create_rectangle(0, 0, 28, 16, fill=f"#{r:02x}{g:02x}{b:02x}", outline="")

    def _refresh_iou_ba_plot(
        self,
        means_nm2: Optional[np.ndarray],
        diffs_nm2: Optional[np.ndarray],
        mean_err: float,
        loa_low: float,
        loa_high: float,
    ) -> None:
        if self.iou_ba_ax is None or self.iou_ba_canvas is None:
            return
        self.iou_ba_ax.clear()
        self.iou_ba_ax.grid(True, alpha=0.2)
        self.iou_ba_ax.set_xlabel("Mean area (nm²)")
        self.iou_ba_ax.set_ylabel("Diff (Manual - SAM)")
        self.iou_ba_ax.tick_params(labelsize=PLOT_TEXT_SIZE, pad=2)

        if means_nm2 is not None and diffs_nm2 is not None and means_nm2.size > 0:
            self.iou_ba_ax.scatter(means_nm2, diffs_nm2, s=14, color="#2b6cb0", alpha=0.8, edgecolors="none")
            if not math.isnan(mean_err):
                self.iou_ba_ax.axhline(mean_err, color="#444", linestyle="--", linewidth=1)
            if not math.isnan(loa_low) and not math.isnan(loa_high):
                self.iou_ba_ax.axhline(loa_low, color="#aa3333", linestyle=":", linewidth=1)
                self.iou_ba_ax.axhline(loa_high, color="#aa3333", linestyle=":", linewidth=1)
        else:
            self.iou_ba_ax.text(
                0.5,
                0.5,
                "No data",
                ha="center",
                va="center",
                transform=self.iou_ba_ax.transAxes,
                fontsize=8,
                color="#666666",
            )

        self.iou_ba_canvas.draw_idle()

    def _label_mask_components(self, mask: np.ndarray) -> Tuple[np.ndarray, int]:
        try:
            from scipy.ndimage import label as ndi_label

            labels, num = ndi_label(mask.astype(np.uint8))
            return labels, int(num)
        except Exception:
            return self._label_mask_components_fallback(mask)

    def _label_mask_components_fallback(self, mask: np.ndarray) -> Tuple[np.ndarray, int]:
        h, w = mask.shape
        labels = np.zeros((h, w), dtype=np.int32)
        current = 0
        for y in range(h):
            row = mask[y]
            for x in range(w):
                if not row[x] or labels[y, x] != 0:
                    continue
                current += 1
                stack = [(y, x)]
                labels[y, x] = current
                while stack:
                    cy, cx = stack.pop()
                    for ny, nx in (
                        (cy - 1, cx),
                        (cy + 1, cx),
                        (cy, cx - 1),
                        (cy, cx + 1),
                    ):
                        if 0 <= ny < h and 0 <= nx < w and mask[ny, nx] and labels[ny, nx] == 0:
                            labels[ny, nx] = current
                            stack.append((ny, nx))
        return labels, current

    def _build_iou_base_image(self) -> Image.Image:
        if self.image_rgb is None:
            return Image.new("RGB", (640, 480), (0, 0, 0))
        base = Image.fromarray(self.image_rgb).convert("RGBA")
        active_ids = self.iou_selected_snapshot if self.iou_selected_snapshot else self.selected_ids
        if not active_ids:
            return base.convert("RGB")
        h, w = self.image_rgb.shape[:2]
        overlay = np.zeros((h, w, 4), dtype=np.uint8)
        r, g, b = self.selection_color_rgb
        alpha = int(round(255 * float(self.selection_alpha)))
        value = np.array([r, g, b, alpha], dtype=np.uint8)
        for pid in sorted(active_ids):
            particle = self._particle_by_id(pid)
            if particle is None:
                continue
            self._apply_mask_to_canvas(overlay, particle, value)
        overlay_img = Image.fromarray(overlay, mode="RGBA")
        merged = Image.alpha_composite(base, overlay_img)
        return merged.convert("RGB")

    def _on_iou_brush_press(self, event) -> None:
        self._ensure_iou_drawing_enabled()
        self.iou_last_pos = (event.x, event.y)
        self._draw_iou_point(event.x, event.y)

    def _on_iou_brush_drag(self, event) -> None:
        self._ensure_iou_drawing_enabled()
        self._draw_iou_line(event.x, event.y)

    def _on_iou_brush_release(self, _event) -> None:
        self.iou_last_pos = None

    def _draw_iou_line(self, x: float, y: float) -> None:
        if self.iou_last_pos is None:
            self.iou_last_pos = (x, y)
        x0, y0 = self.iou_last_pos
        steps = int(max(abs(x - x0), abs(y - y0)) // 2) + 1
        for i in range(steps + 1):
            t = i / max(1, steps)
            xi = x0 + (x - x0) * t
            yi = y0 + (y - y0) * t
            self._draw_iou_point(xi, yi, refresh=False)
        self.iou_last_pos = (x, y)
        self._refresh_iou_canvas()

    def _draw_iou_point(self, x: float, y: float, refresh: bool = True) -> None:
        if self.iou_overlay_image is None or self.iou_mask is None:
            return
        if self.iou_brush_size_var is None:
            size = 18
        else:
            try:
                size = float(self.iou_brush_size_var.get().strip())
            except Exception:
                size = 18.0
        size = max(1.0, min(200.0, size))
        radius_orig = size / 2.0
        radius_disp = radius_orig * self.iou_scale

        # Draw on display overlay.
        draw_disp = ImageDraw.Draw(self.iou_overlay_image)
        if self.iou_brush_mode == "eraser":
            draw_disp.ellipse(
                (x - radius_disp, y - radius_disp, x + radius_disp, y + radius_disp),
                fill=(0, 0, 0, 0),
            )
        else:
            r, g, b = self.iou_brush_color
            draw_disp.ellipse(
                (x - radius_disp, y - radius_disp, x + radius_disp, y + radius_disp),
                fill=(r, g, b, 140),
            )

        # Draw on full-resolution mask.
        scale = self.iou_scale if self.iou_scale > 0 else 1.0
        x_orig = x / scale
        y_orig = y / scale
        draw_mask = ImageDraw.Draw(self.iou_mask)
        fill_val = 0 if self.iou_brush_mode == "eraser" else 255
        draw_mask.ellipse(
            (x_orig - radius_orig, y_orig - radius_orig, x_orig + radius_orig, y_orig + radius_orig),
            fill=fill_val,
        )

        if refresh:
            self._refresh_iou_canvas()

    def _refresh_iou_canvas(self) -> None:
        if self.iou_canvas is None or self.iou_base_image is None or self.iou_overlay_image is None:
            return
        base = self.iou_base_image
        scale = self.iou_scale
        disp = base.resize(
            (self.iou_overlay_image.width, self.iou_overlay_image.height),
            Image.Resampling.BILINEAR,
        )
        merged = disp.convert("RGBA")
        merged = Image.alpha_composite(merged, self.iou_overlay_image)
        self.iou_display_tk = ImageTk.PhotoImage(merged)
        self.iou_canvas.delete("all")
        self.iou_canvas.create_image(0, 0, anchor="nw", image=self.iou_display_tk)

    def _clear_iou_mask(self) -> None:
        if self.iou_base_image is None:
            return
        w, h = self.iou_base_image.size
        self.iou_mask = Image.new("L", (w, h), 0)
        if self.iou_overlay_image is not None:
            self.iou_overlay_image = Image.new("RGBA", self.iou_overlay_image.size, (0, 0, 0, 0))
        self._refresh_iou_canvas()
        if self.iou_result_var is not None:
            self.iou_result_var.set("")
        if self.iou_precision_var is not None:
            self.iou_precision_var.set("")
        if self.iou_recall_var is not None:
            self.iou_recall_var.set("")
        if self.iou_f1_var is not None:
            self.iou_f1_var.set("")
        if self.iou_ba_mean_var is not None:
            self.iou_ba_mean_var.set("")
        if self.iou_ba_sd_var is not None:
            self.iou_ba_sd_var.set("")
        if self.iou_ba_loa_var is not None:
            self.iou_ba_loa_var.set("")
        self.iou_ba_data = None
        self._refresh_iou_ba_plot(None, None, float("nan"), float("nan"), float("nan"))

    def _save_iou_overlay_image(self) -> None:
        if self.iou_base_image is None or self.iou_overlay_image is None:
            messagebox.showerror("AI Evaluation", "No evaluation image to save.")
            return
        save_path = filedialog.asksaveasfilename(
            title="Save Evaluation Image",
            defaultextension=".png",
            initialfile="ai_evaluation_ground_truth.png",
            filetypes=[
                ("PNG", "*.png"),
                ("JPEG", "*.jpg"),
                ("TIFF", "*.tiff"),
            ],
        )
        if not save_path:
            return
        ext = os.path.splitext(save_path)[1].lower()
        if ext not in (".png", ".jpg", ".jpeg", ".tif", ".tiff"):
            save_path = f"{save_path}.png"

        try:
            disp = self.iou_base_image.resize(
                (self.iou_overlay_image.width, self.iou_overlay_image.height),
                Image.Resampling.BILINEAR,
            ).convert("RGBA")
            merged = Image.alpha_composite(disp, self.iou_overlay_image)
            merged.save(save_path)
        except Exception as exc:
            messagebox.showerror("Save error", f"Could not save image:\n{exc}")
            return
        self.status_var.set(f"AI evaluation image saved: {os.path.basename(save_path)}")

    def _save_iou_ba_plot(self) -> None:
        if self.iou_ba_fig is None or self.iou_ba_ax is None:
            messagebox.showerror("AI Evaluation", "No Bland-Altman plot available.")
            return
        self._save_axes_image(self.iou_ba_fig, [self.iou_ba_ax], "Save Bland-Altman Graph", "bland_altman.png")

    def _save_iou_ba_data(self) -> None:
        if self.iou_ba_data is None:
            messagebox.showerror("AI Evaluation", "No Bland-Altman data available.")
            return
        sam_nm2, manual_nm2 = self.iou_ba_data
        if sam_nm2.size == 0 or manual_nm2.size == 0:
            messagebox.showerror("AI Evaluation", "No Bland-Altman data available.")
            return

        file_path = filedialog.asksaveasfilename(
            title="Save Bland-Altman Data",
            defaultextension=".csv",
            filetypes=[
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx"),
            ],
        )
        if not file_path:
            return
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".csv", ".xlsx"):
            file_path = f"{file_path}.csv"
            ext = ".csv"

        try:
            headers = ["SAM area (nm^2)", "Manual area (nm^2)"]
            if ext == ".xlsx":
                if not XLSX_AVAILABLE:
                    raise RuntimeError(f"openpyxl not available: {XLSX_IMPORT_ERROR}")
                wb = Workbook()
                ws = wb.active
                ws.append(headers)
                for x, y in zip(sam_nm2, manual_nm2):
                    ws.append([float(x), float(y)])
                wb.save(file_path)
            else:
                with open(file_path, "w", newline="", encoding="utf-8") as handle:
                    writer = csv.writer(handle)
                    writer.writerow(headers)
                    for x, y in zip(sam_nm2, manual_nm2):
                        writer.writerow([float(x), float(y)])
        except Exception as exc:
            messagebox.showerror("Save error", f"Could not save data:\n{exc}")
            return

        self.status_var.set(f"Saved Bland-Altman data: {os.path.basename(file_path)}")

    def _compute_iou(self) -> None:
        if self.iou_mask is None:
            return
        if not self.iou_selected_snapshot:
            messagebox.showerror("AI Evaluation", "Select particles first.")
            return
        mask_arr = np.asarray(self.iou_mask, dtype=bool)
        if not np.any(mask_arr):
            messagebox.showerror("AI Evaluation", "Paint ground truth masks before computing IoU.")
            return

        if self.image_rgb is None:
            messagebox.showerror("AI Evaluation", "Import an image first.")
            return
        pred = np.zeros(self.image_rgb.shape[:2], dtype=bool)
        for pid in sorted(self.iou_selected_snapshot):
            p = self._particle_by_id(pid)
            if p is None:
                continue
            self._apply_mask_to_canvas(pred, p, True)

        inter = np.logical_and(pred, mask_arr)
        union = np.logical_or(pred, mask_arr)
        tp = float(np.sum(inter))
        fp = float(np.sum(np.logical_and(pred, np.logical_not(mask_arr))))
        fn = float(np.sum(np.logical_and(mask_arr, np.logical_not(pred))))

        iou_val = float(np.sum(inter)) / float(np.sum(union)) if np.any(union) else 0.0
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        f1 = (2.0 * precision * recall) / (precision + recall) if (precision + recall) > 0 else 0.0

        if self.iou_result_var is not None:
            self.iou_result_var.set(f"{iou_val:.4f}")
        if self.iou_precision_var is not None:
            self.iou_precision_var.set(f"{precision:.4f}")
        if self.iou_recall_var is not None:
            self.iou_recall_var.set(f"{recall:.4f}")
        if self.iou_f1_var is not None:
            self.iou_f1_var.set(f"{f1:.4f}")

        # Bland-Altman stats using per-particle areas (SAM vs manual mask components)
        mean_err = float("nan")
        sd_err = float("nan")
        loa_low = float("nan")
        loa_high = float("nan")
        means_nm2 = None
        diffs_nm2 = None
        try:
            labels, num = self._label_mask_components(mask_arr)
            if num > 0:
                comp_area: Dict[int, int] = {}
                for cid in range(1, num + 1):
                    comp_area[cid] = int(np.sum(labels == cid))

                assigned: Set[int] = set()
                sam_areas: List[float] = []
                manual_areas: List[float] = []
                img_h, img_w = labels.shape
                for pid in sorted(self.iou_selected_snapshot):
                    p = self._particle_by_id(pid)
                    if p is None:
                        continue
                    seg = p.segmentation
                    if seg is None or seg.size == 0:
                        continue
                    ox, oy = self._particle_offset(p)
                    y0, x0 = oy, ox
                    y1, x1 = y0 + seg.shape[0], x0 + seg.shape[1]
                    if y1 <= 0 or x1 <= 0 or y0 >= img_h or x0 >= img_w:
                        continue
                    y0c = max(0, y0)
                    x0c = max(0, x0)
                    y1c = min(img_h, y1)
                    x1c = min(img_w, x1)
                    seg_view = seg[(y0c - y0) : (y1c - y0), (x0c - x0) : (x1c - x0)]
                    if seg_view.size == 0:
                        continue
                    labels_view = labels[y0c:y1c, x0c:x1c]
                    overlap_labels = labels_view[seg_view]
                    if overlap_labels.size == 0:
                        continue
                    counts = np.bincount(overlap_labels.ravel())
                    if counts.size <= 1:
                        continue
                    counts[0] = 0
                    order = np.argsort(counts)[::-1]
                    chosen = None
                    for cid in order:
                        if cid == 0:
                            continue
                        if counts[cid] <= 0:
                            break
                        if cid not in assigned:
                            chosen = int(cid)
                            break
                    if chosen is None:
                        continue
                    assigned.add(chosen)
                    sam_areas.append(float(p.area_px))
                    manual_areas.append(float(comp_area.get(chosen, 0)))

                if sam_areas:
                    nm_per_px = self._safe_um_per_px_for_refresh() * 1000.0
                    sam_nm2 = np.asarray(sam_areas, dtype=float) * (nm_per_px ** 2)
                    manual_nm2 = np.asarray(manual_areas, dtype=float) * (nm_per_px ** 2)
                    diffs = manual_nm2 - sam_nm2
                    means = (manual_nm2 + sam_nm2) / 2.0
                    diffs_nm2 = diffs
                    means_nm2 = means
                    self.iou_ba_data = (sam_nm2.copy(), manual_nm2.copy())
                    mean_err = float(np.mean(diffs)) if diffs.size > 0 else float("nan")
                    if diffs.size > 1:
                        sd_err = float(np.std(diffs, ddof=1))
                    else:
                        sd_err = 0.0
                    loa_low = mean_err - 1.96 * sd_err
                    loa_high = mean_err + 1.96 * sd_err
        except Exception:
            pass

        if self.iou_ba_mean_var is not None:
            self.iou_ba_mean_var.set("" if math.isnan(mean_err) else f"{mean_err:.4f}")
        if self.iou_ba_sd_var is not None:
            self.iou_ba_sd_var.set("" if math.isnan(sd_err) else f"{sd_err:.4f}")
        if self.iou_ba_loa_var is not None:
            if math.isnan(loa_low) or math.isnan(loa_high):
                self.iou_ba_loa_var.set("")
            else:
                self.iou_ba_loa_var.set(f"{loa_low:.4f} to {loa_high:.4f}")

        self._refresh_iou_ba_plot(means_nm2, diffs_nm2, mean_err, loa_low, loa_high)
        self.status_var.set(f"AI evaluation computed: IoU {iou_val:.4f}")

    def _close_flip_window(self) -> None:
        if self.flip_window is not None and self.flip_window.winfo_exists():
            self.flip_window.destroy()
        self.flip_window = None

    def _apply_image_transform(self, transform_fn: Callable[[np.ndarray], np.ndarray], status: str) -> None:
        if self.image_rgb is None and not self.image_rgbs:
            messagebox.showerror("Error", "Import an image first.")
            return

        if self.image_rgbs:
            try:
                self.image_rgbs = [np.asarray(transform_fn(img), dtype=np.uint8) for img in self.image_rgbs]
            except Exception as exc:
                messagebox.showerror("Flip Image", f"Could not transform image:\n{exc}")
                return

            if self.particles_by_image:
                for img_idx, plist in self.particles_by_image.items():
                    for p in plist:
                        try:
                            p.segmentation = np.asarray(transform_fn(p.segmentation), dtype=bool)
                        except Exception:
                            p.segmentation = np.asarray(transform_fn(p.segmentation.astype(np.uint8)) > 0, dtype=bool)
                        self._recompute_particle_geometry(p)

            self._compose_active_images()
            self._sync_particles_for_layout()
        else:
            try:
                self.image_rgb = np.asarray(transform_fn(self.image_rgb), dtype=np.uint8)
            except Exception as exc:
                messagebox.showerror("Flip Image", f"Could not transform image:\n{exc}")
                return

            if self.particles:
                for p in self.particles:
                    try:
                        p.segmentation = np.asarray(transform_fn(p.segmentation), dtype=bool)
                    except Exception:
                        p.segmentation = np.asarray(transform_fn(p.segmentation.astype(np.uint8)) > 0, dtype=bool)
                    self._recompute_particle_geometry(p)
                self._build_pick_map()
            else:
                self.mask_pick_map = None

        self.measure_line = None
        self.measure_label = None
        self.angle_points = []
        self.angle_label = None
        self.angle_preview_point = None
        self.section_shape = None
        self.section_bbox = None
        self.section_label = None

        self.render_image()
        self.refresh_table()
        self.status_var.set(status)

    def _recompute_particle_geometry(self, particle: ParticleMask) -> None:
        seg = particle.segmentation
        area_px = int(seg.sum())
        ys, xs = np.nonzero(seg)
        if len(xs) == 0:
            particle.area_px = area_px
            particle.centroid_xy = (0.0, 0.0)
            particle.bbox_xywh = (0, 0, 0, 0)
            return
        cx = float(xs.mean())
        cy = float(ys.mean())
        x_min, x_max = int(xs.min()), int(xs.max())
        y_min, y_max = int(ys.min()), int(ys.max())
        particle.area_px = area_px
        particle.centroid_xy = (cx, cy)
        particle.bbox_xywh = (x_min, y_min, int(x_max - x_min + 1), int(y_max - y_min + 1))
        particle.feret_px = None

    def _store_measurement_line(self, x0: float, y0: float, x1: float, y1: float) -> None:
        d_px = float(np.hypot(x1 - x0, y1 - y0))
        nm_per_px = self._safe_um_per_px_for_refresh() * 1000.0
        d_nm = d_px * nm_per_px
        self.measure_line = (x0, y0, x1, y1)
        self.measure_label = f"{d_px:.2f} px | {d_nm:.2f} nm"
        self.status_var.set(f"Distance: {d_px:.2f} px | {d_nm:.2f} nm")

    def _store_section(self, shape: str, x_min: float, y_min: float, x_max: float, y_max: float) -> None:
        w_px = max(0.0, float(x_max - x_min))
        h_px = max(0.0, float(y_max - y_min))
        nm_per_px = self._safe_um_per_px_for_refresh() * 1000.0
        w_nm = w_px * nm_per_px
        h_nm = h_px * nm_per_px
        self.section_shape = shape
        self.section_bbox = (x_min, y_min, x_max, y_max)
        self.section_label = f"w: {w_nm:.2f} nm\nh: {h_nm:.2f} nm"
        self.status_var.set(f"{shape} section: w {w_nm:.2f} nm | h {h_nm:.2f} nm")

    def _handle_angle_click(self, x: float, y: float) -> None:
        if len(self.angle_points) >= 3:
            self.angle_points = []
            self.angle_label = None
            self.angle_preview_point = None

        self.angle_points.append((x, y))
        if len(self.angle_points) == 2:
            self.status_var.set("Angle: first line defined. Click a third point.")
        elif len(self.angle_points) == 3:
            p1, p2, p3 = self.angle_points
            angle_deg = self._compute_angle_deg(p1, p2, p3)
            if np.isfinite(angle_deg):
                self.angle_label = f"{angle_deg:.2f}°"
                self.status_var.set(f"Angle: {angle_deg:.2f}°")
            else:
                self.angle_label = "N/A"
                self.status_var.set("Angle: N/A")
            self.angle_preview_point = None
        else:
            self.status_var.set("Angle: click second point.")

        self.render_image(keep_view=True)

    def _compute_angle_deg(
        self,
        p1: Tuple[float, float],
        p2: Tuple[float, float],
        p3: Tuple[float, float],
    ) -> float:
        v1 = np.array([p1[0] - p2[0], p1[1] - p2[1]], dtype=float)
        v2 = np.array([p3[0] - p2[0], p3[1] - p2[1]], dtype=float)
        n1 = float(np.linalg.norm(v1))
        n2 = float(np.linalg.norm(v2))
        if n1 == 0.0 or n2 == 0.0:
            return float("nan")
        cosang = float(np.dot(v1, v2) / (n1 * n2))
        cosang = max(-1.0, min(1.0, cosang))
        return float(np.degrees(np.arccos(cosang)))

    def open_histogram(self) -> None:
        if not self.selected_ids:
            messagebox.showwarning("No selection", "Select one or more particles first.")
            return

        if self.hist_window is not None and self.hist_window.winfo_exists():
            self.hist_window.lift()
            self._update_histogram_plot()
            return

        self.hist_window = tk.Toplevel(self.root)
        self.hist_window.title("Particle Histogram")
        self.hist_window.geometry("760x520+1200+180")

        controls = ttk.Frame(self.hist_window, padding=8)
        controls.pack(fill=tk.X)

        ttk.Label(controls, text="Property:").pack(side=tk.LEFT)
        self.hist_property_var = tk.StringVar(value="length")
        combo = ttk.Combobox(
            controls,
            textvariable=self.hist_property_var,
            state="readonly",
            values=("length", "width", "area", "feret diameter", "circularity", "eccentricity"),
            width=12,
        )
        combo.pack(side=tk.LEFT, padx=(8, 0))
        combo.bind("<<ComboboxSelected>>", lambda _evt: self._update_histogram_plot())
        ttk.Button(controls, text="Save Histogram", command=self.on_save_histogram).pack(side=tk.RIGHT)

        self.hist_fig = Figure(figsize=(7.2, 4.6), dpi=100)
        self.hist_ax = self.hist_fig.add_subplot(111)
        self.hist_canvas = FigureCanvasTkAgg(self.hist_fig, master=self.hist_window)
        self.hist_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        self.hist_click_cid = self.hist_canvas.mpl_connect("button_press_event", self._on_histogram_click)

        self.hist_window.protocol("WM_DELETE_WINDOW", self._close_hist_window)
        self._update_histogram_plot()

    def _close_hist_window(self) -> None:
        if self.hist_window is not None and self.hist_window.winfo_exists():
            self.hist_window.destroy()
        self.hist_window = None
        self.hist_fig = None
        self.hist_ax = None
        self.hist_canvas = None
        self.hist_property_var = None
        self.hist_click_cid = None
        self.hist_bins = None
        self.hist_values = None
        self.hist_value_ids = None
        self.hist_patches = []
        self.hist_active_bin_idx = None
        self.highlighted_ids.clear()
        self.render_image(keep_view=True)

    def on_save_histogram(self) -> None:
        if self.hist_fig is None or self.hist_window is None or not self.hist_window.winfo_exists():
            messagebox.showerror("Error", "Open histogram first.")
            return

        save_path = filedialog.asksaveasfilename(
            title="Save Histogram",
            defaultextension=".png",
            initialfile="particle_histogram.png",
            filetypes=[
                ("PNG", "*.png"),
                ("JPEG", "*.jpg"),
                ("TIFF", "*.tiff"),
            ],
        )
        if not save_path:
            return

        ext = os.path.splitext(save_path)[1].lower()
        if ext not in (".png", ".jpg", ".jpeg", ".tif", ".tiff"):
            save_path = f"{save_path}.png"

        try:
            self.hist_fig.savefig(save_path, dpi=300, bbox_inches="tight")
        except Exception as exc:
            messagebox.showerror("Save error", f"Could not save histogram:\n{exc}")
            return

        self.status_var.set(f"Histogram saved: {os.path.basename(save_path)}")

    def _selected_particles(self) -> List[ParticleMask]:
        selected: List[ParticleMask] = []
        for pid in sorted(self.selected_ids):
            p = self._particle_by_id(pid)
            if p is not None:
                selected.append(p)
        return selected

    def _hist_values(self, prop: str) -> Tuple[np.ndarray, str]:
        um_per_px = self._safe_um_per_px_for_refresh()
        nm_per_px = um_per_px * 1000.0
        values: List[float] = []

        for p in self._selected_particles():
            length_px, width_px = self._center_length_width_px(p)
            if prop == "length":
                values.append(length_px * nm_per_px)
            elif prop == "width":
                values.append(width_px * nm_per_px)
            elif prop == "area":
                values.append(float(p.area_px) * (nm_per_px ** 2))
            elif prop == "feret diameter":
                values.append(self._feret_diameter_px(p) * nm_per_px)
            elif prop == "circularity":
                values.append(self._circularity_eccentricity(p)[0])
            elif prop == "eccentricity":
                values.append(self._circularity_eccentricity(p)[1])

        arr = np.asarray(values, dtype=float)
        if prop == "length":
            return arr, "Length (nm)"
        if prop == "width":
            return arr, "Width (nm)"
        if prop == "feret diameter":
            return arr, "Feret Diameter (nm)"
        if prop == "circularity":
            return arr, "Circularity"
        if prop == "eccentricity":
            return arr, "Eccentricity"
        return arr, "Area (nm²)"

    def _hist_values_with_ids(self, prop: str) -> Tuple[np.ndarray, np.ndarray, str]:
        um_per_px = self._safe_um_per_px_for_refresh()
        nm_per_px = um_per_px * 1000.0
        values: List[float] = []
        ids: List[int] = []

        for p in self._selected_particles():
            length_px, width_px = self._center_length_width_px(p)
            if prop == "length":
                values.append(length_px * nm_per_px)
            elif prop == "width":
                values.append(width_px * nm_per_px)
            elif prop == "area":
                values.append(float(p.area_px) * (nm_per_px ** 2))
            elif prop == "feret diameter":
                values.append(self._feret_diameter_px(p) * nm_per_px)
            elif prop == "circularity":
                values.append(self._circularity_eccentricity(p)[0])
            elif prop == "eccentricity":
                values.append(self._circularity_eccentricity(p)[1])
            ids.append(int(p.mask_id))

        arr = np.asarray(values, dtype=float)
        arr_ids = np.asarray(ids, dtype=int)
        if prop == "length":
            return arr, arr_ids, "Length (nm)"
        if prop == "width":
            return arr, arr_ids, "Width (nm)"
        if prop == "feret diameter":
            return arr, arr_ids, "Feret Diameter (nm)"
        if prop == "circularity":
            return arr, arr_ids, "Circularity"
        if prop == "eccentricity":
            return arr, arr_ids, "Eccentricity"
        return arr, arr_ids, "Area (nm^2)"

    def _on_histogram_click(self, event) -> None:
        if self.hist_ax is None or self.hist_canvas is None:
            return
        if event.inaxes != self.hist_ax or event.xdata is None:
            return
        if self.hist_bins is None or self.hist_values is None or self.hist_value_ids is None:
            return
        if len(self.hist_bins) < 2:
            return

        x = float(event.xdata)
        bin_idx = int(np.searchsorted(self.hist_bins, x, side="right") - 1)
        if bin_idx < 0 or bin_idx >= (len(self.hist_bins) - 1):
            return

        # Toggle behavior: second click on the same bin clears subgroup selection.
        if self.hist_active_bin_idx is not None and int(self.hist_active_bin_idx) == int(bin_idx):
            self.hist_active_bin_idx = None
            self.highlighted_ids.clear()
            if self.hist_patches:
                for patch in self.hist_patches:
                    patch.set_facecolor("#1f77b4")
            self.hist_canvas.draw_idle()
            self.render_image(keep_view=True)
            self.status_var.set("Histogram subgroup selection cleared")
            return

        self.hist_active_bin_idx = int(bin_idx)
        if self.hist_patches:
            for i, patch in enumerate(self.hist_patches):
                patch.set_facecolor("#e53935" if i == bin_idx else "#1f77b4")

        per_value_bins = np.digitize(self.hist_values, self.hist_bins, right=False) - 1
        per_value_bins = np.clip(per_value_bins, 0, len(self.hist_bins) - 2)
        self.highlighted_ids = {
            int(pid)
            for pid, b in zip(self.hist_value_ids, per_value_bins)
            if int(b) == bin_idx
        }
        self.highlighted_ids &= self.selected_ids

        self.hist_canvas.draw_idle()
        self.render_image(keep_view=True)
        self.status_var.set(f"Histogram subgroup selected: {len(self.highlighted_ids)} particles")

    def _update_histogram_plot(self) -> None:
        if self.hist_ax is None or self.hist_canvas is None or self.hist_property_var is None:
            return

        prop = self.hist_property_var.get().strip().lower()
        values, ids, xlabel = self._hist_values_with_ids(prop)
        self.hist_values = values
        self.hist_value_ids = ids
        self.hist_bins = None
        self.hist_patches = []
        self.hist_active_bin_idx = None
        self.highlighted_ids.clear()
        self.hist_ax.clear()

        if values.size == 0:
            self.hist_ax.text(0.5, 0.5, "No selected particles", ha="center", va="center")
            self.hist_ax.set_title("Particle Size Histogram")
            self.hist_ax.set_xlabel(xlabel)
            self.hist_ax.set_ylabel("Count")
            self.hist_canvas.draw_idle()
            self.render_image(keep_view=True)
            return

        bins = max(8, min(40, int(np.sqrt(values.size) * 2)))
        _counts, edges, patches = self.hist_ax.hist(
            values,
            bins=bins,
            color="#1f77b4",
            edgecolor="white",
            linewidth=0.6,
            alpha=0.9,
        )
        self.hist_bins = np.asarray(edges, dtype=float)
        self.hist_patches = list(patches)
        self.hist_ax.set_title(f"Particle Size Histogram ({xlabel})")
        self.hist_ax.set_xlabel(xlabel)
        self.hist_ax.set_ylabel("Count")
        self.hist_ax.grid(alpha=0.25, linestyle="--")
        self.hist_canvas.draw_idle()
        self.render_image(keep_view=True)

    def _save_rows_to_csv_or_xlsx(
        self,
        headers: Tuple[str, ...],
        rows: List[Tuple],
        dialog_title: str,
        initialfile: str,
        sheet_name: str,
    ) -> None:
        save_path = filedialog.asksaveasfilename(
            title=dialog_title,
            defaultextension=".csv",
            initialfile=initialfile,
            filetypes=[
                ("CSV", "*.csv"),
                ("Excel workbook", "*.xlsx"),
            ],
        )
        if not save_path:
            return

        ext = os.path.splitext(save_path)[1].lower()
        if ext not in (".csv", ".xlsx"):
            save_path = f"{save_path}.csv"
            ext = ".csv"

        try:
            if ext == ".csv":
                with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(rows)
            else:
                if not XLSX_AVAILABLE:
                    messagebox.showerror(
                        "XLSX unavailable",
                        "Saving as .xlsx requires openpyxl.\n\nInstall with:\npip install openpyxl",
                    )
                    return
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                ws.append(list(headers))
                for row in rows:
                    ws.append(list(row))
                wb.save(save_path)
        except Exception as exc:
            messagebox.showerror("Save error", f"Could not save file:\n{exc}")
            return

        self.status_var.set(f"Data saved: {os.path.basename(save_path)}")

    def on_save_data(self) -> None:
        columns = tuple(self.tree["columns"]) if self.tree is not None else ("id", "area_um2", "length_nm", "width_nm")
        rows = []
        for row_id in self.tree.get_children():
            values = self.tree.item(row_id, "values")
            if values:
                rows.append(tuple(values[: len(columns)]))

        if not rows:
            messagebox.showerror("Error", "Select particles")
            return
        headers = tuple(str(self.tree.heading(col, "text")) for col in columns)

        self._save_rows_to_csv_or_xlsx(
            headers=headers,
            rows=rows,
            dialog_title="Save DATA STREAM",
            initialfile="data_stream.csv",
            sheet_name="DATA STREAM",
        )

    def refresh_table(self) -> None:
        um_per_px = self._safe_um_per_px_for_refresh()
        nm_per_px = um_per_px * 1000.0
        for row in self.tree.get_children():
            self.tree.delete(row)

        total = 0.0
        for pid in sorted(self.selected_ids):
            p = self._particle_by_id(pid)
            if p is None:
                continue
            area_nm2 = float(p.area_px) * (nm_per_px ** 2)
            length_px, width_px = self._center_length_width_px(p)
            length_nm = length_px * nm_per_px
            width_nm = width_px * nm_per_px
            total += area_nm2
            if self.show_more_data:
                circ_val, ecc_val = self._circularity_eccentricity(p)
                feret_nm = self._feret_diameter_px(p) * nm_per_px
                values = (
                    f"#{pid}",
                    f"{area_nm2:.2f}",
                    f"{length_nm:.2f}",
                    f"{width_nm:.2f}",
                    f"{circ_val:.4f}",
                    f"{ecc_val:.4f}",
                    f"{feret_nm:.2f}",
                )
            else:
                values = (f"#{pid}", f"{area_nm2:.2f}", f"{length_nm:.2f}", f"{width_nm:.2f}")
            self.tree.insert("", tk.END, values=values)

        self.summary_var.set(f"Selected: {len(self.selected_ids)} | Total area: {total:.2f} nm²")
        if self.hist_window is not None and self.hist_window.winfo_exists():
            self._update_histogram_plot()
        if self.basic_stats_window is not None and self.basic_stats_window.winfo_exists():
            self._refresh_basic_stats_window()
        if self.gmm_window is not None and self.gmm_window.winfo_exists():
            self._refresh_gmm_window()
        if self.dbscan_window is not None and self.dbscan_window.winfo_exists():
            self._refresh_dbscan_window()
        if self.ttest_window is not None and self.ttest_window.winfo_exists():
            self._refresh_ttest_window()
        if self.violin_window is not None and self.violin_window.winfo_exists():
            self._refresh_violin_plot()
        if self.bivariate_window is not None and self.bivariate_window.winfo_exists():
            self._refresh_bivariate_plot()
        if self.bivariate_gmm_window is not None and self.bivariate_gmm_window.winfo_exists():
            self._refresh_bivariate_gmm_plot()
        if self.bootstrap_window is not None and self.bootstrap_window.winfo_exists():
            self._refresh_bootstrap_window()

    def _center_line_components(self, particle: ParticleMask) -> Tuple[float, float, float, float]:
        seg = particle.segmentation
        h, w = seg.shape
        cx, cy = particle.centroid_xy
        xi = int(round(cx))
        yi = int(round(cy))
        xi = max(0, min(w - 1, xi))
        yi = max(0, min(h - 1, yi))

        row = seg[yi, :]
        col = seg[:, xi]

        horiz = 0.0
        vert = 0.0

        if row.any():
            xs = np.flatnonzero(row)
            horiz = float(xs[-1] - xs[0] + 1)
        if col.any():
            ys = np.flatnonzero(col)
            vert = float(ys[-1] - ys[0] + 1)

        if horiz <= 0.0 and vert <= 0.0:
            bbox_w = float(max(1, particle.bbox_xywh[2]))
            bbox_h = float(max(1, particle.bbox_xywh[3]))
            horiz, vert = bbox_w, bbox_h
        elif horiz <= 0.0:
            horiz = float(max(1, particle.bbox_xywh[2]))
        elif vert <= 0.0:
            vert = float(max(1, particle.bbox_xywh[3]))

        return horiz, vert, float(cx), float(cy)

    def _center_length_width_px(self, particle: ParticleMask) -> Tuple[float, float]:
        horiz, vert, _cx, _cy = self._center_line_components(particle)
        length_px = max(horiz, vert)
        width_px = min(horiz, vert)

        pid = int(particle.mask_id)
        if pid in self.length_overrides_px:
            length_px = float(self.length_overrides_px[pid])
        if pid in self.width_overrides_px:
            width_px = float(self.width_overrides_px[pid])
        return float(length_px), float(width_px)

    def _safe_um_per_px_for_refresh(self) -> float:
        try:
            value = float(self.um_per_px_var.get().strip())
        except Exception:
            return 1.0
        return value if value > 0 else 1.0


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="SAM particle segmentation for SEM/TEM/AFM images.")
    parser.add_argument("--test-image", type=str, default=None, help="Run quick SAM test on one image path.")
    parser.add_argument("--checkpoint", type=str, default=None, help="SAM checkpoint .pth path.")
    parser.add_argument("--model-type", type=str, default="vit_b", choices=["vit_b", "vit_l", "vit_h"], help="SAM model type.")
    parser.add_argument("--um-per-px", type=float, default=0.01, help="Scale for area conversion.")
    return parser.parse_args(argv)


def main(argv: List[str]) -> int:
    args = parse_args(argv)
    if args.test_image:
        if not args.checkpoint:
            print("For --test-image, --checkpoint is required.")
            return 2
        try:
            quick_test(
                image_path=args.test_image,
                checkpoint_path=args.checkpoint,
                model_type=args.model_type,
                um_per_px=float(args.um_per_px),
            )
            return 0
        except Exception as exc:
            print(f"Quick test failed: {exc}")
            return 1

    root = tk.Tk()
    ParticleAIApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))